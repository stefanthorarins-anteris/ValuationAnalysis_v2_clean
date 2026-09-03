"""The deliverable gate: the deck is always built, and a broken deliverable is loud.

TWO PRODUCTION FAILURES UNDER TEST
  1. `generate_presentation.py` was hand-run, so no pipeline run had rendered the deck
     since 2026-07-17.
  2. On 2026-09-01 the presentation XLSX shipped with 20 of 20 pages skipped -- 4,797
     bytes, one empty sheet -- at WARNING severity on line 2209 of a 5,222-line log, and
     the run exited 0.  `AggScoreTop100` lost `price`/`beta`/`sector`/`rating_fmp` on 88 of
     97 rows in the same event.

NO NETWORK ANYWHERE IN THIS FILE.  Every FMP-shaped call is injected.  The one test that
drives the real `postBo.createPresentation` replaces `gdg.safe_json_list` outright, which
is the same seam `test_post_fetch_hardening` uses.

EVERY WRITE GOES TO `tmp_path`.  The repo write guard (conftest RULE T / RULE E) is live
and none of these tests may trip it: `createPresentation` is handed an absolute tmp path
for its workbook, and the audit is pointed at `repo_root=tmp_path`.
"""

import io
import os
import sys

import pandas as pd
import pytest

REPO = os.path.dirname(os.path.abspath(__file__))
if REPO not in sys.path:
    sys.path.insert(0, REPO)

import deliverables as dlv          # noqa: E402
import getData_gen as gdg           # noqa: E402


# --------------------------------------------------------------------------- #
#  helpers                                                                     #
# --------------------------------------------------------------------------- #
def _xlsx(path, tickers):
    """A workbook shaped like `createPresentation`'s: one sheet per ticker plus the
    default `Sheet` openpyxl creates and the producer never removes.  `tickers=[]`
    reproduces the 2026-09-01 artifact exactly."""
    import openpyxl
    wb = openpyxl.Workbook()
    for t in tickers:
        ws = wb.create_sheet(title=t)
        ws['A1'] = t
        ws['A2'] = 'Earnings yield'
        ws['B2'] = 0.1234
    wb.save(str(path))
    wb.close()
    return str(path)


#  THE REAL AggScore COLUMN SET, taken from the header of
#  `AggScoreTop100-2026-08-07_fmp_stock_CUR3K.csv`.
#
#  THE FIXTURE'S WIDTH IS LOAD-BEARING, which is not obvious and cost a round.  The first
#  version of `_aggscore` declared six columns and produced 2,662 bytes for 97 rows -- 27
#  bytes per row against the real artifact's 338 (30 columns, 33,762 bytes for 100 rows).
#  It was therefore a fifth of the width of the deliverable it claimed to stand in for,
#  which made it too small for this kind's byte floor and made every test using it a
#  weaker control than it looked: a narrow fixture cannot exercise a floor calibrated
#  against the real thing.  Same defect class as the S1 it helped hide -- a fixture that
#  is not the shape of the artifact.
_AGGSCORE_COLUMNS = [
    'source', 'price', 'PE-ratio', 'beta', 'sector', 'rating_fmp', 'currentRatio',
    'dividendYield', 'GrahamNumberToPrice', 'GrossProfitMargin_ttm', 'DCF-to-Price',
    'M-Score', 'C-Score', 'CycleHeat', 'moatScore', 'isFinancial', 'financialKind',
    'forensicValid', 'M_flag_gt_-1.78', 'M_drivers', 'C_flag_ge_4', 'C_flags_fired',
    'sloanAccruals', 'sloan_worstQuintile_inShortlist', 'forensicTag', 'volAvg_report',
    'volAvg_asof', 'universe', 'universe_fingerprint',
]


def _aggscore(path, rows, blank=(), sentinel='NaN'):
    """An AggScore CSV of the REAL width.  `blank` names the vendor columns to kill; they
    are filled with the pipeline's own string sentinel, not a real NaN -- which is how the
    2026-09-01 file was actually written."""
    cols = {c: ['x'] * rows for c in _AGGSCORE_COLUMNS}
    cols['source'] = ['T%d' % i for i in range(rows)]
    cols['price'] = [10.0 + i for i in range(rows)]
    cols['beta'] = [1.0] * rows
    cols['sector'] = ['Energy'] * rows
    cols['rating_fmp'] = ['A'] * rows
    cols['universe_fingerprint'] = ['a1b2c3d4e5f6a7b8'] * rows
    cols['M_drivers'] = ['DSRI=1.02,GMI=0.98,AQI=1.01,SGI=1.10'] * rows
    for c in blank:
        cols[c] = [sentinel] * rows
    pd.DataFrame(cols).to_csv(str(path), index=False)
    return str(path)


def _deck(path, pages, pad=True):
    """A rendered deck with `pages` name pages.  Padded past the 50 KB floor so the page
    count -- not the byte floor -- is what the assertion is about."""
    body = ''.join('<div class="name-page" id="T%d"><div class="page-header">'
                   '<h1>T%d</h1></div></div>' % (i, i) for i in range(pages))
    filler = ('<!-- %s -->' % ('x' * 4000)) * 20 if pad else ''
    with io.open(str(path), 'w', encoding='utf-8') as fh:
        fh.write('<html><body>' + filler + body + '</body></html>')
    return str(path)


def _ok_deck_report(path, expected):
    return {'kind': 'deck', 'path': path, 'ok': True, 'reason': None,
            'expected_pages': expected, 'returncode': 0, 'stderr_tail': ''}


# =========================================================================== #
#  1.  THE DECK IS WIRED IN, AND THE WIRED PATH IS OFFLINE                     #
# =========================================================================== #
def test_the_wired_deck_command_shuts_the_ONLY_network_path_off():
    """The deck's sole network path is the Yahoo augment, and `--augment` DEFAULTS TO ON.
    Wiring the deck into every run with the default would put a per-page fetch into a
    pipeline whose operator is at ~90% of his FMP limit -- Yahoo is not FMP, but an
    unattended run that reaches the network at all is not what was authorised.

    BOTH flags are asserted, not either: they are independent kill switches, so a future
    change to one default cannot quietly restore the traffic."""
    cmd = dlv.deck_command('2026-09-01', repo_root='/repo', python_exe='py')
    assert '--augment' in cmd and cmd[cmd.index('--augment') + 1] == 'off', cmd
    assert '--no-augment' in cmd, cmd


def test_the_wired_deck_command_PINS_the_run_date():
    """Unpinned, the deck resolves 'the latest postRank on disk'.  In a run that failed
    before writing its pickle that is YESTERDAY's, and the audit would then certify a
    stale deck as this run's deliverable."""
    cmd = dlv.deck_command('2026-09-01', repo_root='/repo', python_exe='py')
    assert '--run-date' in cmd and cmd[cmd.index('--run-date') + 1] == '2026-09-01', cmd
    assert '--run-dir' in cmd and cmd[cmd.index('--run-dir') + 1] == '/repo', cmd


def test_the_deck_stage_NEVER_raises_when_the_subprocess_dies(tmp_path):
    """A deck failure must not cost a run whose picks are already on disk.  The stage is
    reached AFTER the postRank pickle is written and after two Drive syncs."""
    def _boom(cmd, cwd, timeout_s):
        raise RuntimeError('subprocess machinery exploded')

    rep = dlv.run_deck_stage('2026-09-01', repo_root=str(tmp_path), verbose=False,
                             _run=_boom)
    assert rep['ok'] is False
    assert 'exploded' in rep['reason']


def test_the_deck_stage_reports_a_nonzero_returncode_as_failure(tmp_path):
    def _fail(cmd, cwd, timeout_s):
        return {'returncode': 1, 'stdout': '', 'stderr': 'FileNotFoundError: no postRank'}

    rep = dlv.run_deck_stage('2026-09-01', repo_root=str(tmp_path), verbose=False,
                             _run=_fail)
    assert rep['ok'] is False
    assert 'exited 1' in rep['reason']
    assert 'FileNotFoundError' in rep['stderr_tail']


def test_the_deck_manifest_is_parsed_off_stdout(tmp_path):
    line = ('DECK-MANIFEST expected_pages=45 rendered_pages=45 augment=off '
            'net_clients=NONE run_date=2026-09-01 bytes=2432304 out=x.html')

    def _run(cmd, cwd, timeout_s):
        return {'returncode': 0, 'stdout': 'noise\n' + line + '\nmore noise', 'stderr': ''}

    rep = dlv.run_deck_stage('2026-09-01', repo_root=str(tmp_path), verbose=False,
                             _run=_run)
    assert rep['ok'] is True
    assert rep['expected_pages'] == 45


# =========================================================================== #
#  2.  A MISSING OR EMPTY DELIVERABLE IS IMPOSSIBLE TO MISS                    #
# =========================================================================== #
def test_a_declared_deliverable_that_does_not_exist_FAILS_the_run(tmp_path):
    """The producer declares the filename unconditionally and writes it inside a guard,
    so 'declared but absent' is exactly what a failed stage looks like from outside."""
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables(['AggScoreTop100-2026-09-01_fmp_x.csv'], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 2
    names = [i['name'] for i in res['failed']]
    assert 'AggScoreTop100-2026-09-01_fmp_x.csv' in names
    why = ' '.join(res['failed'][0]['why'])
    assert 'DOES NOT EXIST' in why


def test_the_2026_09_01_XLSX_shape_FAILS(tmp_path):
    """THE EVENT.  20 of 20 pages skipped: openpyxl's default `Sheet` and nothing else.
    4,797 bytes -- above any byte floor that would not also fire on a legitimately small
    run -- so the STRUCTURAL check is what has to catch it."""
    x = _xlsx(tmp_path / 'PresentationTop20-2026-09-01_fmp_x.xlsx', [])
    rep = dlv.xlsx_page_report(x)
    assert rep['pages'] == 0 and rep['sheets'] == 1

    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([os.path.basename(x)], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 2
    why = ' '.join(w for i in res['failed'] for w in i['why'])
    assert 'STRUCTURALLY EMPTY' in why


def test_a_PARTIAL_xlsx_FAILS_and_NAMES_the_missing_pages(tmp_path):
    """3 of 20 is a FAILED, not a warning: a missing page means a name in the CEO's
    top-20 has nothing behind it, so the artifact does not do its job for that name.  The
    banner must also NAME them -- 'something is short' sends the operator back to the
    5,000-line log this gate exists to replace."""
    x = _xlsx(tmp_path / 'PresentationTop20-2026-09-01_fmp_x.xlsx', ['AAA', 'BBB', 'CCC'])
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables(
        [os.path.basename(x)], '2026-09-01', repo_root=str(tmp_path),
        deck_report=_ok_deck_report(deck, 5),
        xlsx_report={'expected': 20, 'written': 3, 'skipped': ['DDD', 'EEE']})
    assert res['exit_code'] == 2
    why = ' '.join(w for i in res['failed'] for w in i['why'])
    assert '3 of 20' in why
    assert 'DDD' in why and 'EEE' in why


def test_a_COMPLETE_xlsx_and_deck_exit_ZERO(tmp_path):
    """The false-positive control.  A gate that fires on a healthy run is worse than no
    gate: it trains the operator to ignore the one signal this change adds."""
    x = _xlsx(tmp_path / 'PresentationTop20-2026-09-01_fmp_x.xlsx',
              ['T%d' % i for i in range(20)])
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 45)
    csv = _aggscore(tmp_path / 'AggScoreTop100-2026-09-01_fmp_x.csv', 97)
    res = dlv.audit_deliverables(
        [os.path.basename(x), os.path.basename(csv)], '2026-09-01',
        repo_root=str(tmp_path), deck_report=_ok_deck_report(deck, 45),
        xlsx_report={'expected': 20, 'written': 20, 'skipped': []})
    assert res['exit_code'] == 0, [i['why'] for i in res['failed'] + res['degraded']]
    assert not res['failed'] and not res['degraded']


def test_a_truncated_file_FAILS_on_the_byte_floor(tmp_path):
    """The floor from BELOW.  On its own this direction cannot falsify a floor -- it
    passes for any threshold above 7 bytes, which is why it coexisted with a floor that
    failed 8 of 13 real deliverables.  It is only meaningful PAIRED with
    `test_the_audit_PASSES_a_REAL_historical_declared_set`, which constrains the floor
    from above over the real population.  Neither test is sufficient alone; do not delete
    one and keep the other."""
    p = tmp_path / 'AggScoreTop100-2026-09-01_fmp_x.csv'
    with io.open(str(p), 'w', encoding='utf-8') as fh:
        fh.write('source\n')
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-01', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 2
    assert 'byte floor' in ' '.join(res['failed'][0]['why'])


# =========================================================================== #
#  3.  THE DECK HALF OF THE GATE                                              #
# =========================================================================== #
def test_a_run_that_NEVER_BUILT_THE_DECK_fails(tmp_path):
    """The 2026-07-17 failure, stated as a test: `deck_report=None` means no stage ran.
    A deck nobody attempted must not be indistinguishable from a deck that succeeded."""
    res = dlv.audit_deliverables([], '2026-09-01', repo_root=str(tmp_path),
                                 deck_report=None)
    assert res['exit_code'] == 2
    why = ' '.join(w for i in res['failed'] for w in i['why'])
    assert 'DID NOT RUN' in why
    assert '2026-07-17' in why


def test_a_deck_with_ZERO_name_pages_fails(tmp_path):
    d = _deck(tmp_path / 'presentation_2026-09-01.html', 0)
    res = dlv.audit_deliverables([], '2026-09-01', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(d, 45))
    assert res['exit_code'] == 2
    assert 'NO NAME PAGES' in ' '.join(res['failed'][0]['why'])


def test_a_PARTIAL_deck_fails_even_though_the_stage_reported_success(tmp_path):
    """The producer's self-report is not the measurement.  A deck that exits 0, prints
    `expected_pages=45` and writes three pages is caught because the page count comes off
    the FILE."""
    d = _deck(tmp_path / 'presentation_2026-09-01.html', 3)
    res = dlv.audit_deliverables([], '2026-09-01', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(d, 45))
    assert res['exit_code'] == 2
    assert '3 of 45' in ' '.join(res['failed'][0]['why'])


def test_the_page_marker_counts_the_real_deck_markup():
    """The count is only meaningful if the marker is the one the real builder emits.
    Read off `build_name_page` itself rather than trusting a fixture."""
    import inspect
    import generate_presentation as gp
    src = inspect.getsource(gp.PresentationBuilder.build_name_page)
    assert dlv._DECK_PAGE_MARKER in src, (
        'the audit counts %r and build_name_page no longer emits it'
        % dlv._DECK_PAGE_MARKER)


def test_the_render_seam_plans_exactly_what_it_builds():
    """`_render_pages` is the one place a name page is produced from a list, so the plan
    and the render cannot disagree.  Duplicates are PRESERVED on purpose: a name in both
    the general pool and a cohort gets a page in each, and de-duplicating the plan (as
    `_page_tickers` does, correctly, for the Yahoo fetch set) would report a healthy deck
    as over-complete -- measured 45 unique tickers against 53 rendered pages on the
    2026-08-04 artifacts."""
    import generate_presentation as gp
    b = object.__new__(gp.PresentationBuilder)
    b._page_plan = []
    b._pages_built = 0
    b.build_name_page = lambda t, i, c: '<page %s %d %s>' % (t, i, c)
    out = b._render_pages(['AAA', 'BBB'], 'general')
    out += b._render_pages(['AAA'], 'Mining')
    assert b._pages_built == 3
    assert b._page_plan == [('AAA', 'general'), ('BBB', 'general'), ('AAA', 'Mining')]
    #  THE COUPLING, ASSERTED DIRECTLY.  The first version of this test checked the plan
    #  and the count separately and a mutation that de-duplicated the plan BY TICKER --
    #  the `_page_tickers` mistake, i.e. the realistic one -- survived it.
    assert len(b._page_plan) == b._pages_built
    assert [t for t, _ in b._page_plan].count('AAA') == 2, (
        'the plan must keep a name that renders under two sections')
    assert out.count('<page') == 3
    assert '<page AAA 1 Mining>' in out          # rank restarts per section


def test_build_html_renders_ONLY_through_the_planning_seam():
    """The seam is only the single source of truth if nothing bypasses it.  A fourth
    section added later that calls `build_name_page` directly would render pages the plan
    never counted, so `expected` would come in BELOW `rendered` and the audit -- which
    only fails on `rendered < expected` -- would wave the shortfall through."""
    import inspect
    import generate_presentation as gp
    src = inspect.getsource(gp.PresentationBuilder.build_html)
    assert 'self.build_name_page(' not in src, (
        'build_html calls build_name_page directly; route it through _render_pages so the '
        'page plan cannot go stale')
    assert src.count('self._render_pages(') >= 3, src.count('self._render_pages(')


# =========================================================================== #
#  4.  THE AGGSCORE VENDOR BLOCK (the other half of the 09-01 event)          #
# =========================================================================== #
def test_the_pipelines_NaN_string_sentinel_reads_as_a_real_NaN(tmp_path):
    """A PREMISE THAT WAS WRONG THE FIRST TIME, recorded rather than quietly corrected.

    This test was written to prove that `isna()` CANNOT see the pipeline's string 'NaN'
    sentinel, which was the justification for a string-matching layer in
    `aggscore_vendor_report`.  It failed on its own premise: `pd.read_csv`'s default
    `na_values` already contains 'NaN'/'nan'/'NA'/'None'/'null'/'', so the sentinel is a
    real NaN by the time the check reads it.  The string layer could never have fired --
    an inert guard -- and was removed.  This test now pins the fact that made it inert, so
    the layer is not re-added on the same wrong hunch."""
    p = _aggscore(tmp_path / 'AggScoreTop100-2026-09-01_fmp_x.csv', 97,
                  blank=('price', 'beta', 'sector', 'rating_fmp'), sentinel='NaN')
    df = pd.read_csv(p)
    assert df['price'].isna().mean() == 1.0, 'read_csv no longer converts the sentinel'
    rep = dlv.aggscore_vendor_report(p)
    assert rep['all_dead'] is True
    assert set(rep['dead']) == set(dlv.AGGSCORE_VENDOR_COLUMNS)


def test_a_WHITESPACE_only_cell_is_the_one_case_isna_misses(tmp_path):
    """The residual the whitespace half of the check exists for, and the only reason it is
    not also inert: `' '` is NOT in `na_values`, so it survives read_csv as a string and a
    column full of it is fully-populated as far as `isna()` is concerned."""
    p = _aggscore(tmp_path / 'AggScoreTop100-2026-09-01_fmp_x.csv', 97,
                  blank=dlv.AGGSCORE_VENDOR_COLUMNS, sentinel=' ')
    df = pd.read_csv(p)
    assert df['sector'].isna().mean() == 0.0, 'premise: isna() cannot see a blank string'
    rep = dlv.aggscore_vendor_report(p)
    assert rep['all_dead'] is True, rep


def test_the_dead_vendor_block_FAILS_the_run(tmp_path):
    p = _aggscore(tmp_path / 'AggScoreTop100-2026-09-01_fmp_x.csv', 97,
                  blank=dlv.AGGSCORE_VENDOR_COLUMNS)
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([os.path.basename(p)], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 2
    assert 'VENDOR BLOCK DEAD' in ' '.join(res['failed'][0]['why'])


def test_ONE_dead_vendor_column_is_DEGRADED_not_FAILED(tmp_path):
    """The alarm-fatigue boundary, and it is a real design line rather than a hedge: one
    flaky endpoint killing one column is a common event with a different action (note it,
    ship), while all four dying together is the throttle signature with the action 'stop
    and re-run'.  Firing the hard gate on the common case is how a gate stops being read."""
    p = _aggscore(tmp_path / 'AggScoreTop100-2026-09-01_fmp_x.csv', 97, blank=('beta',))
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([os.path.basename(p)], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 0
    assert res['degraded'] and not res['failed']
    assert 'beta' in ' '.join(res['degraded'][0]['why'])


# =========================================================================== #
#  5.  CAUSE ATTRIBUTION -- a quota event is not a code defect                #
# =========================================================================== #
def test_a_throttled_run_says_QUOTA_on_the_banner(tmp_path):
    x = _xlsx(tmp_path / 'PresentationTop20-2026-09-01_fmp_x.xlsx', [])
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    tally = {'calls': 578 * 3, 'non_200': 578, 'exceptions': 0, 'by_status': {429: 578}}
    res = dlv.audit_deliverables([os.path.basename(x)], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 http_tally=tally)
    assert res['quota_suspected'] is True
    banner = dlv.format_banner(res)
    assert 'QUOTA event, not a code defect' in banner
    assert '429' in banner


def test_a_MISSING_ARTIFACT_failure_is_NOT_called_a_quota_event(tmp_path):
    """The distinction the directive is about.  A deck that could not find its inputs and
    a deck whose calls were refused need different actions, and a banner that blamed the
    vendor for a missing file would send the operator to wait out a quota that was never
    the problem."""
    res = dlv.audit_deliverables(
        [], '2026-09-01', repo_root=str(tmp_path),
        deck_report={'kind': 'deck', 'ok': False,
                     'path': str(tmp_path / 'presentation_2026-09-01.html'),
                     'reason': 'generate_presentation.py exited 1',
                     'stderr_tail': 'FileNotFoundError: required Boresults pickle not found'},
        http_tally=None)
    assert res['quota_suspected'] is False
    banner = dlv.format_banner(res)
    assert 'MISSING-INPUT' not in banner or 'QUOTA event' not in banner
    assert 'INPUT ARTIFACT' in banner


def test_no_429_means_no_quota_claim(tmp_path):
    """The negative control on the attribution: a shortfall with a clean HTTP window must
    NOT be blamed on the vendor."""
    x = _xlsx(tmp_path / 'PresentationTop20-2026-09-01_fmp_x.xlsx', [])
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    tally = {'calls': 140, 'non_200': 0, 'exceptions': 0, 'by_status': {200: 140}}
    res = dlv.audit_deliverables([os.path.basename(x)], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 http_tally=tally)
    assert res['quota_suspected'] is False
    assert 'QUOTA event' not in dlv.format_banner(res)


# =========================================================================== #
#  6.  THE BANNER AND THE EXIT STATUS                                         #
# =========================================================================== #
def test_the_banner_reaches_BOTH_streams_and_names_the_file():
    """The 2026-09-01 failure WAS reported -- at WARNING severity, on line 2209 of 5,222.
    Loud means: on stdout (which `run_logger` tees into the run log) AND on stderr (what a
    console showing only errors, and a scheduler, capture)."""
    res = {'run_date': '2026-09-01', 'items': [], 'ok': [], 'degraded': [],
           'failed': [{'name': 'PresentationTop20-2026-09-01_fmp_x.xlsx', 'state': 'FAILED',
                       'why': ['STRUCTURALLY EMPTY -- 1 sheet(s), none of them a ticker page'],
                       'bytes': 4797, 'actual': 0, 'expected': 20}],
           'quota_suspected': False, 'http_tally': None, 'exit_code': 2}
    out, err = io.StringIO(), io.StringIO()
    code = dlv.emit_audit(res, out=out, err=err)
    assert code == 2
    for stream in (out.getvalue(), err.getvalue()):
        assert 'PresentationTop20-2026-09-01_fmp_x.xlsx' in stream
        assert 'STRUCTURALLY EMPTY' in stream
        assert '4,797 bytes' in stream


def test_a_clean_run_writes_the_banner_to_stdout_only():
    res = {'run_date': '2026-09-01', 'items': [{'name': 'a.csv', 'state': 'OK',
                                                'why': [], 'bytes': 5000}],
           'ok': [{'name': 'a.csv', 'state': 'OK', 'why': [], 'bytes': 5000}],
           'degraded': [], 'failed': [], 'quota_suspected': False,
           'http_tally': None, 'exit_code': 0}
    out, err = io.StringIO(), io.StringIO()
    assert dlv.emit_audit(res, out=out, err=err) == 0
    assert 'ALL 1 DELIVERABLE(S) PRESENT' in out.getvalue()
    assert err.getvalue() == ''


def test_the_exit_code_is_2_not_1():
    """`Sbocker.main` already exits non-zero on an unhandled exception.  Reusing 1 would
    make 'the run crashed' and 'the run finished with broken deliverables' the same signal
    to anything reading the status -- which is the failure being fixed."""
    res = dlv.audit_deliverables([], '2026-09-01', repo_root=os.path.dirname(__file__),
                                 deck_report=None)
    assert res['exit_code'] == 2


# =========================================================================== #
#  7.  THE HTTP TALLY -- retries are charged requests                         #
# =========================================================================== #
def test_the_tally_counts_every_RETRY_as_a_charged_request():
    """A 429 costs three requests, not one: `safe_http_get` retries three times.  A tally
    of call SITES would under-report the quota spend threefold on exactly the runs where
    the number matters."""
    class _R:
        status_code = 429

        def json(self):
            return []

    base = gdg.http_tally_snapshot()
    gdg.safe_http_get('http://x', _get=lambda *a, **k: _R(), sleep=lambda _s: None)
    d = gdg.http_tally_delta(base)
    assert d['calls'] == 3, d
    assert d['by_status'].get(429) == 3, d
    assert d['non_200'] == 3, d


def test_the_tally_records_a_RAISED_request_as_an_attempt():
    """A read timeout is a request that left the machine."""
    import requests

    def _boom(*a, **k):
        raise requests.ConnectionError('socket died')

    base = gdg.http_tally_snapshot()
    gdg.safe_http_get('http://x', _get=_boom, sleep=lambda _s: None)
    d = gdg.http_tally_delta(base)
    assert d['calls'] == 3 and d['exceptions'] == 3, d


def test_the_tally_is_read_by_DIFFERENCE_so_two_windows_do_not_clobber_each_other():
    class _R:
        status_code = 200

        def json(self):
            return []

    outer = gdg.http_tally_snapshot()
    gdg.safe_http_get('http://x', _get=lambda *a, **k: _R(), sleep=lambda _s: None)
    inner = gdg.http_tally_snapshot()
    gdg.safe_http_get('http://x', _get=lambda *a, **k: _R(), sleep=lambda _s: None)
    assert gdg.http_tally_delta(inner)['calls'] == 1
    assert gdg.http_tally_delta(outer)['calls'] == 2


# =========================================================================== #
#  8.  THE PRODUCER DECLARES -- driven through the real createPresentation     #
# =========================================================================== #
def test_createPresentation_REPORTS_a_fully_skipped_run(tmp_path, monkeypatch):
    """THE 2026-09-01 EVENT REPRODUCED OFFLINE.  Every vendor call returns [] -- which is
    what a throttled call and a genuinely empty response both look like to this stage --
    so every page is skipped, and the workbook is written with only openpyxl's default
    sheet.  Before this change the function returned None and the ONLY record of
    `len(symblist)` was a printed sentence, so no consumer could tell 20-of-20-skipped
    from a legitimately small run.

    NO NETWORK: `gdg.safe_json_list` is replaced wholesale, the same seam
    `test_post_fetch_hardening` uses.  The workbook goes to `tmp_path`, so the repo write
    guard is not tripped."""
    import postBo as pb

    monkeypatch.setattr(pb.gdg, 'safe_json_list', lambda *a, **k: [])
    fname = str(tmp_path / 'PresentationTop20-2026-09-01_fmp_x.xlsx')
    fb_df = pd.DataFrame({'source': ['AAA', 'BBB', 'CCC'],
                          'AggScore': [0.9, 0.8, 0.7]})
    empty = pd.DataFrame(columns=['source'])

    rep = pb.createPresentation(fb_df, empty, empty, 'http://base/', 'KEY', 3, fname,
                                10, flag_df=None, bands=None)

    assert rep is not None, 'the stage must declare what it meant to write'
    assert rep['expected'] == 3
    assert rep['written'] == 0
    assert sorted(rep['skipped']) == ['AAA', 'BBB', 'CCC']
    assert os.path.exists(fname)

    #  And the audit turns that report into a FAILED run rather than a warning.
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([os.path.basename(fname)], '2026-09-01',
                                 repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 xlsx_report=rep)
    assert res['exit_code'] == 2


def test_writeResWrapper_keeps_the_xlsx_report_where_the_audit_can_reach_it():
    """The report has to survive the guard that wraps the stage.  `resdic['xlsx_report']`
    is set to None BEFORE the call, so a stage that raised leaves an explicit 'no
    expectation recorded' rather than an absent key -- which the audit could not tell from
    an older resdic that predates the report."""
    import inspect
    import postBo as pb
    src = inspect.getsource(pb.writeResWrapper)
    i_none = src.index("resdic['xlsx_report'] = None")
    i_call = src.index("resdic['xlsx_report'] = createPresentation(")
    assert i_none < i_call, 'the None must be set before the guarded call, not after'
    assert "resdic['declared_deliverables']" in src
    assert 'http_tally_delta' in src


def test_Sbocker_runs_the_deck_stage_and_audits_with_a_nonzero_exit():
    """The wiring itself.  Parsed, not imported: importing the orchestrator drags the whole
    pipeline import graph into a test run.

    WHAT THIS CANNOT DETECT: that the stage is in the RIGHT PLACE.  It proves the calls
    exist in `main`, not that the deck runs after the postRank pickle it reads."""
    import ast
    with io.open(os.path.join(REPO, 'Sbocker.py'), encoding='utf-8') as fh:
        src = fh.read()
    tree = ast.parse(src)
    main = next(n for n in tree.body
                if isinstance(n, ast.FunctionDef) and n.name == 'main')
    body = ast.dump(main)
    assert 'run_deck_stage' in body, 'the deck stage is not wired into main()'
    assert 'audit_deliverables' in body, 'the deliverable audit is not wired into main()'
    assert 'emit_audit' in body
    assert any(isinstance(n, ast.Raise)
               and isinstance(n.exc, ast.Call)
               and getattr(n.exc.func, 'id', None) == 'SystemExit'
               for n in ast.walk(main)), 'no non-zero exit path in main()'


def test_the_deck_directory_ships_to_Drive():
    """Building the deck on the run machine and not moving it would be the same failure in
    a new place: `presentations/` is git-tracked, but the run machine's git is not a
    channel the CEO pulls.  A DIRECTORY entry, because a root-level `presentation_*.html`
    glob would resolve from CWD and match nothing."""
    import ast
    with io.open(os.path.join(REPO, 'Sbocker.py'), encoding='utf-8') as fh:
        tree = ast.parse(fh.read())
    dirs = None
    for node in ast.walk(tree):
        if isinstance(node, ast.Assign) and any(
                isinstance(t, ast.Name) and t.id == 'allowlist_dirs' for t in node.targets):
            dirs = [e.value for e in node.value.elts
                    if isinstance(e, ast.Constant) and isinstance(e.value, str)]
    assert dirs, 'could not find Sbocker.allowlist_dirs'
    assert 'presentations' in dirs, dirs


# =========================================================================== #
#  9.  AGAINST THE REAL ARTIFACTS ON DISK                                     #
# =========================================================================== #
def _real_declared_set(run_date, datasource, tickerfilter):
    """The declared set a real run of <run_date> produced, GLOBBED off disk in exactly the
    shape `postBo.writeResWrapper` builds it (postBo.py:1236): the three unconditional
    deliverables, then the cohort side-lists, the market-cap band CSVs, the two
    review-reference tables and the provenance sidecar.

    GLOBBED, not hardcoded: a hardcoded list would be a third enumeration of the
    deliverable set, and the whole design rule here is that there is exactly one."""
    import glob as _glob
    suffix = '%s_%s_%s' % (run_date, datasource, tickerfilter)
    names = [
        'AggScoreTop100-%s.csv' % suffix,
        'PresentationTop20-%s.xlsx' % suffix,
        'ForensicFlagsTop100-%s.csv' % suffix,
        'RunProvenance-%s.json' % suffix,
        'RawMetricsTop100-%s_%s.csv' % (run_date, datasource),
        'CohortMetricStats-%s_%s.csv' % (run_date, datasource),
    ]
    for pat in ('SideList_*_Top100-%s.csv' % suffix,
                'MarketCapBand_*-%s.csv' % suffix):
        names += [os.path.basename(f) for f in sorted(_glob.glob(os.path.join(REPO, pat)))]
    return [n for n in names if os.path.exists(os.path.join(REPO, n))]


@pytest.mark.parametrize('run_date,datasource,tickerfilter', [
    ('2026-08-07', 'fmp', 'stock_CUR3K'),
    ('2026-08-04', 'fmp', 'stock_TEST1'),
])
def test_the_audit_PASSES_a_REAL_historical_declared_set(run_date, datasource,
                                                         tickerfilter):
    """THE TEST WHOSE ABSENCE WAS THE DEFECT (review S1, 2026-09-02).

    Every other false-positive control in this file builds its own fixtures, and every
    fixture it built was LARGE -- so a byte floor calibrated only against the three large
    unconditional deliverables passed all of them while failing 8 of 13 real files on the
    known-good 2026-08-07 run and 7 on 2026-08-04.  Every healthy run would have exited 2
    on `809 bytes, below the 1024-byte floor`, and two of those would have taught the
    operator to ignore the banner -- the exact alarm-fatigue failure the partial-page
    argument claims to rule out.  Section 9's other tests call `xlsx_page_report` and
    `aggscore_vendor_report` DIRECTLY and never drive `audit_deliverables`, so nothing ran
    the audit against a real declared set at all.

    This drives the real entry point over the real files: five compact `SideList_*` CSVs
    (187-828 bytes), up to four `MarketCapBand_*` CSVs (217-1,647 bytes) and the large
    ones together.  It must exit 0.

    WHAT IT CANNOT DETECT: a kind that has never been produced on this machine.  It proves
    the floors admit the artifacts that exist, not that they admit every artifact a run
    can legitimately emit."""
    declared = _real_declared_set(run_date, datasource, tickerfilter)
    if len(declared) < 8:
        pytest.skip('only %d of this run\'s artifacts are on this machine' % len(declared))
    #  A real, complete deck of a real run -- so the deck check cannot be what carries or
    #  sinks this test.  Its page count is measured off the file, so `expected` is taken
    #  from the same file: the subject here is the FLOORS over the declared set.
    deck = os.path.join(REPO, 'presentations', 'presentation_2026-07-17.html')
    if not os.path.exists(deck):
        pytest.skip('no real deck on this machine')
    pages = dlv.count_deck_pages(deck)

    res = dlv.audit_deliverables(declared, run_date, repo_root=REPO,
                                 deck_report=_ok_deck_report(deck, pages))
    assert res['exit_code'] == 0, (
        'a KNOWN-GOOD run must not fail its own audit. Failures: %s'
        % [(i['name'], i['bytes'], i['why']) for i in res['failed']])
    assert not res['failed'], [i['name'] for i in res['failed']]
    #  And the compact kinds really are in the set being admitted -- otherwise this test
    #  could pass by not covering them, which is how the original control passed.
    compact = [i for i in res['items']
               if i['name'].startswith(('SideList_', 'MarketCapBand_'))]
    assert len(compact) >= 5, [i['name'] for i in res['items']]
    assert min(i['bytes'] for i in compact) < 1024, (
        'this run has no sub-1KB deliverable, so it cannot exercise the defect')


def test_an_UNMEASURED_deliverable_kind_gets_the_weakest_floor(tmp_path):
    """THE STRUCTURAL HALF OF THE S1 FIX, and the one that matters more than the numbers.
    The old table keyed on EXTENSION, so eight compact CSVs silently inherited a floor
    calibrated for a different kind.  An unrecognised basename must now fall to NONEMPTY:
    a deliverable added to the declared list later has no measured population here, so it
    may not be judged against someone else's."""
    p = tmp_path / 'SomeNewDeliverable-2026-09-01_fmp_x.csv'
    with io.open(str(p), 'w', encoding='utf-8') as fh:
        fh.write('a,b\n1,2\n')                       # 8 bytes: under every old floor
    floor, measured = dlv._min_bytes(str(p))
    assert (floor, measured) == (dlv.NONEMPTY, None)
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-01', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 0, [i['why'] for i in res['failed']]


def test_an_EMPTY_file_still_FAILS_even_with_no_calibrated_floor(tmp_path):
    """The other direction: NONEMPTY is the weakest floor, not the absence of one.  A
    zero-byte write is still a failure, and the message says which check was applied."""
    p = tmp_path / 'SideList_REIT_Top100-2026-09-01_fmp_x.csv'
    io.open(str(p), 'w', encoding='utf-8').close()
    deck = _deck(tmp_path / 'presentation_2026-09-01.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-01', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 2
    why = ' '.join(res['failed'][0]['why'])
    assert 'EMPTY' in why or 'below the' in why, why


def test_every_floor_states_the_population_it_was_calibrated_against():
    """A refusal citing only a threshold leaves the reader unable to judge whether the
    threshold or the file is wrong -- which is how a floor calibrated against three large
    deliverables was applied to eight compact ones with nothing in the message to show it.
    Every calibrated entry carries its kind's measured minimum, and the floor must sit
    UNDER it: a floor at or above the smallest artifact ever seen is a completeness check
    wearing a floor's clothes."""
    for prefix, floor, measured in dlv._KIND_FLOORS:
        assert measured is not None, prefix
        assert floor < measured, (
            '%s: floor %d is not below the smallest observed artifact (%d), so it will '
            'fire on a legitimately small run' % (prefix, floor, measured))
    msg = dlv._floor_message(50, 4096, 11003)
    assert 'smallest observed' in msg and '11003' in msg


def test_a_run_mode_that_CANNOT_build_a_deck_is_DEGRADED_not_FAILED(tmp_path):
    """`-loadboresults` force-disables `saveBoResults` (configuration.py:348-351), so no
    same-date Boresults pickle exists and `resolve_run_artifacts` hard-fails by design.
    That is a property of a supported ZERO-QUOTA replay mode, not a broken run -- and the
    operator reaching for the cheapest mode is exactly the one who must still trust the
    banner.  Loud, named, exit 0."""
    res = dlv.audit_deliverables(
        [], '2026-09-01', repo_root=str(tmp_path),
        deck_report={'kind': 'deck', 'ok': False, 'not_applicable': True, 'path': None,
                     'reason': '-loadboresults disables saveBoResults, so this run wrote '
                               'no same-date Boresults pickle for the deck to read'})
    assert res['exit_code'] == 0, [i['why'] for i in res['failed']]
    assert res['degraded'] and not res['failed']
    banner = dlv.format_banner(res)
    assert 'DECK NOT BUILT THIS RUN' in banner
    assert '-loadboresults' in banner


def _run_shape(d, run_date, with_same_date_boresults):
    """The files a run leaves on disk for the deck to resolve.  The Boresults name follows
    `utils.saveWrapper`: `Bo{type}_dic-{ds}_{tf}_{sf}_{today}_len..._fails....pickle`."""
    names = ['postRank_%s_fmp_stock_CUR6K.pickle' % run_date,
             'AggScoreTop100-%s_fmp_stock_CUR6K.csv' % run_date,
             'ForensicFlagsTop100-%s_fmp_stock_CUR6K.csv' % run_date]
    bo_date = run_date if with_same_date_boresults else '2026-08-31'
    names.append('Boresults_dic-fmp_stock_CUR6K_all_%s_len5900_manelim0_fails430.pickle'
                 % bo_date)
    for n in names:
        io.open(os.path.join(str(d), n), 'w').close()
    return names


def test_a_loadbometric_run_leaves_a_deck_RESOLVABLE_shape(tmp_path):
    """WHY THIS TEST EXISTS: `-loadbometric 1 -bometricfilename <panel>` is the cheap
    re-run mode, and the whole `-loadboresults` skip above rests on the premise that
    `-loadbometric` is DIFFERENT.  It is, and the difference is one line of config:
    `-loadbometric` disables only `saveBoMetric` (configuration.py:344-347), so
    `postBoWrapper` still runs and `utils.saveWrapper('results', resdic)` still writes a
    Boresults pickle stamped with TODAY's date -- which is exactly what
    `resolve_run_artifacts` requires.  Driven over the real resolver on the real filename
    convention, so the premise is measured rather than reasoned."""
    import generate_presentation as gp
    _run_shape(tmp_path, '2026-09-02', with_same_date_boresults=True)
    run_date, postrank, boresults, agg, forensic = gp.resolve_run_artifacts(str(tmp_path))
    assert run_date == '2026-09-02'
    assert '2026-09-02' in os.path.basename(boresults)


def test_a_loadboresults_run_CANNOT_resolve_a_deck_which_is_why_it_is_skipped(tmp_path):
    """The other half of the same premise, and the justification for the skip rather than
    an attempt: `-loadboresults` force-disables `saveBoResults` (configuration.py:348-351),
    so the newest Boresults pickle is a PREVIOUS run's.  `resolve_run_artifacts` refuses
    to pair it with today's postRank -- correctly; that cross-run mixing was a
    publish-blocker -- so attempting the deck in that mode could only ever produce a
    FAILED deliverable and a spurious exit 2 on a zero-quota replay."""
    import generate_presentation as gp
    _run_shape(tmp_path, '2026-09-02', with_same_date_boresults=False)
    with pytest.raises(FileNotFoundError) as e:
        gp.resolve_run_artifacts(str(tmp_path))
    assert 'Boresults' in str(e.value)
    assert 'Refusing to fall back' in str(e.value)


def test_Sbocker_skips_the_deck_on_loadboresults_and_ONLY_on_loadboresults():
    """The mode check must key on `loadBoResults` alone.  `-loadbometric` disables only
    `saveBoMetric` (configuration.py:344-347), so `postBoWrapper` still runs and
    `utils.saveWrapper('results', ...)` still writes a Boresults pickle stamped with
    today's date -- which is what the deck resolves.  Keying the skip on the wrong flag
    would silently stop building the deck on the run the CEO is about to make."""
    import ast
    with io.open(os.path.join(REPO, 'Sbocker.py'), encoding='utf-8') as fh:
        tree = ast.parse(fh.read())
    main = next(n for n in tree.body
                if isinstance(n, ast.FunctionDef) and n.name == 'main')
    guards = [n for n in ast.walk(main)
              if isinstance(n, ast.Call)
              and isinstance(n.func, ast.Attribute) and n.func.attr == 'get'
              and getattr(n.func.value, 'id', None) == 'configdic'
              and n.args and isinstance(n.args[0], ast.Constant)
              and n.args[0].value in ('loadBoResults', 'loadBoMetric')]
    keys = {n.args[0].value for n in guards}
    assert 'loadBoResults' in keys, 'the deck stage does not check loadBoResults'
    assert 'loadBoMetric' not in keys, (
        'the deck stage keys on loadBoMetric; that mode CAN build a deck and must not be '
        'skipped')


@pytest.mark.parametrize('name,pages', [
    ('PresentationTop20-2026-08-04_fmp_stock_TEST1.xlsx', 20),
    ('PresentationTop20-2026-08-07_fmp_stock_CUR3K.xlsx', 20),
])
def test_the_structural_check_reads_the_REAL_shipped_workbooks(name, pages):
    """The fixture-only risk: a check that works on a workbook this file built and not on
    one `createPresentation` built.  Both real artifacts carry 20 ticker sheets plus the
    default `Sheet`, so the filler-sheet subtraction is measured, not assumed."""
    path = os.path.join(REPO, name)
    if not os.path.exists(path):
        pytest.skip('%s not on this machine' % name)
    rep = dlv.xlsx_page_report(path)
    assert rep['pages'] == pages, rep
    assert rep['sheets'] == pages + 1, rep
    assert dlv.XLSX_FILLER_SHEET not in rep['names']


def test_the_vendor_check_passes_a_REAL_healthy_aggscore_csv():
    """The other false-positive control, on real data rather than a fixture."""
    path = os.path.join(REPO, 'AggScoreTop100-2026-08-07_fmp_stock_CUR3K.csv')
    if not os.path.exists(path):
        pytest.skip('no real AggScore CSV on this machine')
    rep = dlv.aggscore_vendor_report(path)
    assert rep['all_dead'] is False, rep
    assert rep['rows'] > 0


# =========================================================================== #
#  10.  THE DECLARE-AFTER-WRITE HOLE (gap 1, closed 2026-09-03)                #
# =========================================================================== #
#  THE HOLE.  `writeResWrapper` appended the side-list, band and sidecar filenames to the
#  declared set ONLY AFTER writing them, and the audit's expected set IS that declaration.
#  So a side-list that failed to write REMOVED ITSELF from the declaration and the audit
#  could not possibly report it: only the three unconditional names, the postRank pickle and
#  the deck could be checked for ABSENCE at all.  Every test in this section drives the REAL
#  `writeResWrapper`, because a source-order assertion is what let the hole ship -- the
#  existing `inspect.getsource` test for `xlsx_report` proves an ordering, not a behaviour.

#  The forensic columns `writeResWrapper` merges into every side-list (Q-66).  THE FIXTURE'S
#  WIDTH IS LOAD-BEARING for the same reason it was for `_aggscore` above: the real 08-07
#  side-lists predate this merge and carry 3 columns, while the CURRENT writer emits 17, and
#  a 3-column stand-in would not be the shape of the artifact under test.
_SIDELIST_FORENSIC_COLUMNS = [
    'carveLabel', 'forensicValid', 'forensicReason', 'forensicNote', 'M_score_mean',
    'M_flag_gt_-1.78', 'M_drivers', 'M_abstain_reason', 'C_score_mean', 'C_flag_ge_4',
    'C_flags_fired', 'sloanAccruals', 'sloan_worstQuintile_inShortlist', 'forensicTag',
]


def _flag_table(sources):
    """A forensic table of the real width, with the free-text columns actually populated:
    `forensicReason` and `M_drivers` carry commas in the shipped artifacts, so the fixture
    must too or the row counter is never asked the question that matters."""
    cols = {'source': list(sources)}
    for c in _SIDELIST_FORENSIC_COLUMNS:
        cols[c] = [''] * len(sources)
    cols['carveLabel'] = ['REIT'] * len(sources)
    cols['forensicReason'] = ['REFUSED: accruals, margins and asset quality are not '
                              'comparable on this cohort'] * len(sources)
    cols['M_drivers'] = ['DSRI=1.02,GMI=0.98,AQI=1.01,SGI=1.10'] * len(sources)
    cols['forensicValid'] = [False] * len(sources)
    return pd.DataFrame(cols)


def _drive_writeResWrapper(tmp_path, monkeypatch, fail_on=None):
    """Run the REAL `postBo.writeResWrapper` in `tmp_path` with every API-touching stage
    stubbed, and return the resdic it declared into.

    `fail_on` -- a substring of a filename whose `to_csv` must raise, i.e. the write-failure
    the audit is supposed to notice.  Patched at `pandas.DataFrame.to_csv` rather than at
    the pipeline, because the point is to fail the WRITE and leave everything else real.

    NO NETWORK: `writeBoAggToCSV`, `createPresentation`, `buildForensicFlagTable`,
    `writeForensicFlagsCSV` and `volavg_report_frame` are all replaced; nothing else in this
    function reaches the vendor.  EVERY WRITE LANDS IN `tmp_path` (chdir), so the repo write
    guard is not tripped.

    The three UNCONDITIONAL deliverables are therefore declared and never written -- their
    writers are the stubs -- so a caller must assert on named files, never on list lengths.
    """
    import carveOut as co
    import postBo as pb
    monkeypatch.chdir(tmp_path)

    fb = pd.DataFrame({'source': ['A', 'B', 'C', 'D', 'E'],
                       'AggScore': [0.9, 0.8, 0.7, 0.6, 0.5],
                       'rankOfRanks_diag': [1.0, 2.0, 3.0, 4.0, 5.0]})
    flags = _flag_table(fb['source'])

    resdic = {
        'ntopagg': 100, 'ntopxlsx': 2,
        'postRank': fb,
        'cdx_df': pd.DataFrame({'source': list('ABCDE'),
                                'date': pd.Timestamp('2025-01-01')}),
        'SLmeanMscore': pd.DataFrame(), 'SLmeanCscore': pd.DataFrame(),
        'baseurl': 'http://x/', 'api_key': 'k',
        'tickerfilter': 'stock_TEST1', 'datasource': 'fmp',
        'universe': 'stock_TEST1', 'universe_fingerprint': 'aaaabbbbccccdddd',
        #  A THREE-NAME cohort and a ONE-NAME cohort.  One member is the real measured
        #  minimum for a compact deliverable (the 2026-08-07 Micro band), so the fixture
        #  carries the case any row FLOOR would have failed.
        'carveout_sidelists': {'REIT': {'postRank': fb.head(3)},
                               'Mining': {'postRank': fb.head(1)}},
        'carveout_labels': {'A': 'REIT'},
    }

    monkeypatch.setattr(pb, 'writeBoAggToCSV', lambda *a, **k: flags)
    monkeypatch.setattr(pb, 'createPresentation', lambda *a, **k: None, raising=False)
    monkeypatch.setattr(pb.ff, 'buildForensicFlagTable', lambda *a, **k: flags)
    monkeypatch.setattr(pb.ff, 'writeForensicFlagsCSV', lambda *a, **k: None)
    monkeypatch.setattr(co, 'partition_by_marketcap', lambda *a, **k: {
        'bands': {'General': fb.head(2), 'Micro_lt_50M': fb.head(1)},
        'band_counts': {'General': 2, 'Micro_lt_50M': 1},
        'unknown_mcap': 0, 'currency_pending': False,
        'band_selective': {'General': True, 'Micro_lt_50M': False},
        'band_note': {'General': '',
                      'Micro_lt_50M': 'NOT A SELECTION: all 1 member(s) of '
                                      'Micro_lt_50M shown (top-5 requested, band has '
                                      'only 1)'}})
    monkeypatch.setattr(co, 'volavg_report_frame', lambda names: pd.DataFrame(
        {'source': list(names), 'volAvg_report': [''] * len(names),
         'volAvg_asof': ['not-captured'] * len(names)}))

    if fail_on is not None:
        _real_to_csv = pd.DataFrame.to_csv

        def _boom(self, path_or_buf=None, *a, **k):
            if isinstance(path_or_buf, str) and fail_on in path_or_buf:
                raise OSError(28, 'No space left on device')
            return _real_to_csv(self, path_or_buf, *a, **k)

        monkeypatch.setattr(pd.DataFrame, 'to_csv', _boom)

    pb.writeResWrapper(resdic)
    return resdic


def test_a_sidelist_that_FAILED_TO_WRITE_IS_STILL_DECLARED(tmp_path, monkeypatch):
    """THE HOLE ITS OWN AUTHOR NAMED AS THE LARGEST ONE.  The REIT side-list raises on
    write.  Before the fix it removed itself from the declaration and the run exited 0 with
    a cohort list silently absent; now it is declared before the attempt, so the audit sees
    a declared file with nothing behind it.

    The second half is the part the old single-try loop got wrong: the Mining cohort AFTER
    the failed one must still be written and still be declared.  One `to_csv` raising used
    to abandon every remaining cohort."""
    resdic = _drive_writeResWrapper(tmp_path, monkeypatch, fail_on='SideList_REIT')
    declared = resdic['declared_deliverables']
    reit = 'SideList_REIT_Top100-%s_fmp_stock_TEST1.csv' % _today()
    mining = 'SideList_Mining_Top100-%s_fmp_stock_TEST1.csv' % _today()

    assert reit in declared, (
        'the failed side-list removed itself from the declaration: %s' % declared)
    assert not os.path.exists(str(tmp_path / reit)), 'the fixture did not fail the write'
    assert mining in declared and os.path.exists(str(tmp_path / mining)), (
        'a cohort AFTER the failed one was abandoned -- the per-file guard is not in place')

    deck = _deck(tmp_path / ('presentation_%s.html' % _today()), 5)
    res = dlv.audit_deliverables([n for n in declared if n.startswith('SideList_')],
                                 _today(), repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations=resdic['deliverable_rows'])
    assert res['exit_code'] == 2
    failed = {i['name']: ' '.join(i['why']) for i in res['failed']}
    assert reit in failed and 'DOES NOT EXIST' in failed[reit], failed
    assert mining not in failed, failed


def test_a_band_CSV_that_FAILED_TO_WRITE_IS_STILL_DECLARED(tmp_path, monkeypatch):
    """The band block had the identical shape and the identical hole.  Its filename depends
    on `band_selective` (the `_ALLMEMBERS_NOT_A_SELECTION` marker), which the partition has
    already decided before any write -- so there is nothing about the name that requires
    waiting for `to_csv` to return."""
    resdic = _drive_writeResWrapper(tmp_path, monkeypatch, fail_on='MarketCapBand_General')
    declared = resdic['declared_deliverables']
    general = 'MarketCapBand_General-%s_fmp_stock_TEST1.csv' % _today()
    micro = ('MarketCapBand_Micro_lt_50M_ALLMEMBERS_NOT_A_SELECTION-%s_fmp_stock_TEST1.csv'
             % _today())

    assert general in declared, declared
    assert not os.path.exists(str(tmp_path / general))
    assert micro in declared and os.path.exists(str(tmp_path / micro)), (
        'the band after the failed one was abandoned')

    deck = _deck(tmp_path / ('presentation_%s.html' % _today()), 5)
    res = dlv.audit_deliverables([n for n in declared if n.startswith('MarketCapBand_')],
                                 _today(), repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations=resdic['deliverable_rows'])
    assert res['exit_code'] == 2
    assert general in [i['name'] for i in res['failed']]


def test_the_provenance_SIDECAR_declares_ITSELF_before_it_is_written(tmp_path, monkeypatch):
    """The sidecar was appended to the declared set after `json.dump` returned, so a sidecar
    that failed to write also vanished from the audit's expected set.  It now declares
    itself first and therefore appears inside its own manifest, which is what a manifest
    that covers the whole run has to do."""
    import json
    resdic = _drive_writeResWrapper(tmp_path, monkeypatch)
    prov_name = 'RunProvenance-%s_fmp_stock_TEST1.json' % _today()
    assert prov_name in resdic['declared_deliverables']
    prov = json.load(io.open(str(tmp_path / prov_name), encoding='utf-8'))
    assert prov_name in prov['deliverables'], (
        'the sidecar does not name itself, so it is still declared only after its write')


def test_the_sidecar_says_WHICH_declared_files_were_NOT_WRITTEN(tmp_path, monkeypatch):
    """`declared` and `written` are now two different facts, and both have to be
    representable or the sidecar becomes a manifest that lies.

    WHY IN THE SIDECAR AND NOT ONLY IN THE AUDIT BANNER: the sidecar is an artifact that
    TRAVELS to Drive, so the discrepancy is diagnosable from the transferred set on the
    receiving machine.  The banner lives in the run log, which is the file the 2026-09-01
    failure was already reported in -- at line 2209 of 5,222.

    The semantics stamp is asserted too.  Two sidecars either side of this change would
    otherwise make the same claim under the same key while meaning different things, which
    is precisely the defect this sidecar exists to end (a universe NAME whose meaning
    changed on 2026-08-02)."""
    import json
    resdic = _drive_writeResWrapper(tmp_path, monkeypatch, fail_on='SideList_REIT')
    prov = json.load(io.open(str(tmp_path /
                                 ('RunProvenance-%s_fmp_stock_TEST1.json' % _today())),
                             encoding='utf-8'))
    reit = 'SideList_REIT_Top100-%s_fmp_stock_TEST1.csv' % _today()
    mining = 'SideList_Mining_Top100-%s_fmp_stock_TEST1.csv' % _today()
    assert prov['deliverables_declaration'] == 'declared-intent-before-write'
    absent = prov['deliverables_declared_not_written']
    assert reit in absent, absent
    assert mining not in absent, absent
    #  and the counts travel with it, so a later offline audit of this run can be given the
    #  producer's own expectations instead of having none
    assert prov['deliverables_expected_rows'][mining] == 1


def test_the_declared_but_absent_file_does_NOT_break_the_transfer_manifest(tmp_path):
    """THE RIPPLE THE FIX HAD TO HANDLE.  `Sbocker.main` passes `writeResWrapper`'s return
    value straight into `transfer_utils.copy_artifacts_to_transfer_dir`, so the declaration
    is also the incremental transfer manifest -- and it may now name a file that does not
    exist.  The copier's contract already covers it ("missing sources are skipped, not an
    error"), and that is asserted here rather than assumed, because a declaration that made
    a healthy run report a transfer ERROR would be the alarm fatigue this gate exists to
    avoid.

    `reconcile_transfer` is NOT affected and is not asserted here: its groups come from
    `Sbocker.allowlist_patterns` globs, which resolve off the filesystem, so a declared name
    with no file behind it never enters reconciliation."""
    import transfer_utils as tu
    src = tmp_path / 'src'
    dest = tmp_path / 'dest'
    src.mkdir()
    dest.mkdir()
    wrote = src / 'SideList_Mining_Top100-2026-09-03_fmp_stock_TEST1.csv'
    io.open(str(wrote), 'w', encoding='utf-8').write('source,AggScore\nA,0.9\n')
    never = str(src / 'SideList_REIT_Top100-2026-09-03_fmp_stock_TEST1.csv')

    res = tu.copy_artifacts_to_transfer_dir(str(dest), [str(wrote), never], verbose=False)
    assert res['status'] == 'success', res
    assert res['copied'] == 1 and res['errors'] == 0, res
    assert os.path.exists(str(dest / wrote.name))


def _today():
    from datetime import datetime
    return datetime.today().strftime('%Y-%m-%d')


# =========================================================================== #
#  11.  COMPLETENESS ON THE COMPACT DELIVERABLES (gap 2, closed 2026-09-03)    #
# =========================================================================== #
#  Five `SideList_*` and up to four `MarketCapBand_*` CSVs are compact BY DESIGN -- 187 to
#  1,647 bytes, 1 to 25 rows measured -- so they got `NONEMPTY` and nothing more, and A
#  HEADER-ONLY SIDE-LIST PASSED THE GATE.  No byte number separates a legitimate five-row
#  cohort list from a truncated one.  The fix is the declare-and-measure split the XLSX
#  check already uses; the risk it must not create is a check that fires on a legitimately
#  tiny cohort, which would be the byte-floor S1 rebuilt in a new place.

def _sidelist(path, rows, declared_cols=None):
    """A side-list of the CURRENT writer's width (source + AggScore + rankOfRanks_diag +
    the 14 merged forensic columns), with the free-text columns populated."""
    n = max(int(rows), 0)
    base = pd.DataFrame({'source': ['T%d' % i for i in range(n)],
                         'AggScore': [0.9 - 0.01 * i for i in range(n)],
                         'rankOfRanks_diag': [float(i + 1) for i in range(n)]})
    flags = _flag_table(base['source'])
    out = base.merge(flags, on='source', how='left') if n else \
        pd.concat([base, flags], axis=1)
    out.to_csv(str(path), index=False)
    return str(path)


def test_a_HEADER_ONLY_sidelist_FAILS_when_the_producer_DECLARED_rows(tmp_path):
    """THE DEFECT.  A cohort of 25 names whose CSV came out with a header and no rows is
    ~150 bytes -- above `NONEMPTY`, below nothing -- so before this change it was reported
    OK.  The producer's own count is what makes it decidable."""
    p = tmp_path / ('SideList_REIT_Top100-2026-09-03_fmp_stock_CUR3K.csv')
    _sidelist(p, 0)
    assert 0 < os.path.getsize(str(p)) < 1024, os.path.getsize(str(p))

    deck = _deck(tmp_path / 'presentation_2026-09-03.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-03', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations={p.name: 25})
    assert res['exit_code'] == 2
    why = ' '.join(res['failed'][0]['why'])
    assert 'INCOMPLETE -- 0 of the 25' in why, why
    assert 'HEADER-ONLY' in why, why
    #  and the banner carries the two numbers, because `150 bytes` is not a readable
    #  statement about a kind whose size says nothing
    assert '0/25 rows' in dlv.format_banner(res)


def test_a_TRUNCATED_sidelist_FAILS(tmp_path):
    """The other half of the same event: some rows written, not all.  A 5-of-25 side-list is
    a cohort the CEO reviews with four fifths of it silently absent."""
    p = tmp_path / 'SideList_Mining_Top100-2026-09-03_fmp_stock_CUR3K.csv'
    _sidelist(p, 5)
    deck = _deck(tmp_path / 'presentation_2026-09-03.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-03', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations={p.name: 25})
    assert res['exit_code'] == 2
    assert 'INCOMPLETE -- 5 of the 25' in ' '.join(res['failed'][0]['why'])


@pytest.mark.parametrize('rows', [1, 2, 5])
def test_a_LEGITIMATELY_TINY_cohort_PASSES(tmp_path, rows):
    """THE FALSE-POSITIVE CONTROL, AT THE REAL MEASURED FLOOR.  A cohort can have ONE
    member: `MarketCapBand_Micro_lt_50M_ALLMEMBERS_NOT_A_SELECTION-2026-08-07_...csv` holds
    exactly one, and the smallest real side-list holds five.  A row FLOOR would have failed
    the first of those on the only run that ever produced one -- the byte-floor S1 in a new
    place, which is what this parameterisation exists to stop.

    Nothing here is a threshold: the expectation IS the producer's frame length, so a
    one-row cohort declares 1, writes 1 and passes."""
    p = tmp_path / ('MarketCapBand_Micro_lt_50M_ALLMEMBERS_NOT_A_SELECTION-2026-09-03_'
                    'fmp_stock_CUR3K.csv')
    _sidelist(p, rows)
    deck = _deck(tmp_path / 'presentation_2026-09-03.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-03', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations={p.name: rows})
    assert res['exit_code'] == 0, [(i['name'], i['why']) for i in res['failed']]
    assert not res['degraded'], [(i['name'], i['why']) for i in res['degraded']]


def test_an_EMPTY_BY_DESIGN_cohort_declares_ZERO_and_PASSES(tmp_path):
    """A cohort with no members is a header-only CSV and that is CORRECT output.  The check
    cannot fire on it, because the producer declares 0 and the file holds 0 -- which is the
    property that separates this from any absolute floor: the same file that PASSES here is
    the file that FAILS in `test_a_HEADER_ONLY_sidelist_...` above, and the only difference
    is what the producer said it had."""
    p = tmp_path / 'SideList_REIT_Top100-2026-09-03_fmp_stock_CUR3K.csv'
    _sidelist(p, 0)
    deck = _deck(tmp_path / 'presentation_2026-09-03.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-03', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations={p.name: 0})
    assert res['exit_code'] == 0, [(i['name'], i['why']) for i in res['failed']]
    assert not res['degraded']


def test_a_compact_deliverable_with_NO_declared_count_DEGRADES_and_never_FAILS(tmp_path):
    """THE DEGRADE PATH THE HISTORICAL RUNS NEED.  An older `resdic`, a loaded run, or an
    offline audit of an artifact from before 2026-09-03 carries no counts.  Failing those
    would fire on every replay -- so the verdict is UNVERIFIED at DEGRADED, exit 0, with the
    reason named.  Silence was the other option and was rejected: a compact deliverable
    whose completeness nobody checked is a fact worth one line."""
    p = tmp_path / 'SideList_REIT_Top100-2026-08-07_fmp_stock_CUR3K.csv'
    _sidelist(p, 0)
    deck = _deck(tmp_path / 'presentation_2026-08-07.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-08-07', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 0, [(i['name'], i['why']) for i in res['failed']]
    assert [i['name'] for i in res['degraded']] == [p.name]
    assert 'UNVERIFIED' in ' '.join(res['degraded'][0]['why'])


def test_a_NON_compact_kind_with_no_declared_count_stays_SILENT(tmp_path):
    """The other side of the same decision.  `AggScoreTop*` has a real byte floor AND a
    vendor-block check, so "no row count was declared" says nothing new about it and must
    not add a line to the banner.  Nine spurious DEGRADED entries per run is how a banner
    stops being read."""
    p = tmp_path / 'AggScoreTop100-2026-09-03_fmp_stock_CUR3K.csv'
    _aggscore(p, 97)
    deck = _deck(tmp_path / 'presentation_2026-09-03.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-03', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5))
    assert res['exit_code'] == 0 and not res['degraded'], (
        [(i['name'], i['why']) for i in res['degraded']])


def test_MORE_rows_than_declared_is_DEGRADED_not_FAILED(tmp_path):
    """Nothing is missing from the file, so refusing the run would cost the operator a
    banner for no recoverable fault -- but the producer's count and the file disagree, which
    means one of the two measurements is wrong, and a disagreement between the only two
    numbers this check has is not something to swallow."""
    p = tmp_path / 'SideList_Mining_Top100-2026-09-03_fmp_stock_CUR3K.csv'
    _sidelist(p, 7)
    deck = _deck(tmp_path / 'presentation_2026-09-03.html', 5)
    res = dlv.audit_deliverables([p.name], '2026-09-03', repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations={p.name: 5})
    assert res['exit_code'] == 0
    assert [i['name'] for i in res['degraded']] == [p.name]
    assert '7 data row(s) against the 5 declared' in ' '.join(res['degraded'][0]['why'])


def test_the_ROW_CHECK_is_keyed_on_the_DECLARATION_not_on_the_KIND(tmp_path):
    """A count declared for a kind that has no entry in `_ROWCOUNT_KINDS` is still honoured,
    so a writer taught to declare one later needs no edit to the auditor.  This is the same
    inversion the floor table uses: the auditor never decides on its own that a file ought
    to have been bigger."""
    p = tmp_path / 'ForensicFlagsTop100-2026-09-03_fmp_stock_CUR3K.csv'
    _aggscore(p, 3)                     # real width, 3 rows, well under this kind's floor
    item = dlv._check_file(str(p), '2026-09-03', expected_rows=100)
    assert item['state'] == 'FAILED'
    assert item['rows'] == 3 and item['expected_rows'] == 100
    assert 'INCOMPLETE -- 3 of the 100' in ' '.join(item['why'])


def test_a_row_expectation_on_a_NON_CSV_is_reported_not_silently_dropped(tmp_path):
    """`count_csv_rows` on a workbook would be nonsense, so the expectation is not applied
    -- but dropping it in silence would make a producer's declaration disappear, which is
    the class of thing this whole change is about.  Named, DEGRADED, exit 0."""
    #  TWENTY sheets, not two: this kind carries an 8 KB floor and a two-sheet workbook
    #  lands under it, so a small fixture would make this test pass on the floor instead of
    #  on the branch it is about (measured: 20 sheets -> 14,725 bytes).
    x = _xlsx(tmp_path / 'PresentationTop20-2026-09-03_fmp_x.xlsx',
              ['T%d' % i for i in range(20)])
    item = dlv._check_file(x, '2026-09-03', expected_rows=20)
    assert item['state'] == 'DEGRADED'
    assert item['rows'] is None and item['expected_rows'] == 20
    assert 'non-CSV' in ' '.join(item['why'])


def test_an_UNREADABLE_csv_with_a_declared_count_FAILS(tmp_path):
    """A directory where a CSV should be, a permission failure, a half-flushed handle: the
    count cannot be taken, and "could not check" must not read as "checked and fine"."""
    d = tmp_path / 'SideList_REIT_Top100-2026-09-03_fmp_stock_CUR3K.csv'
    d.mkdir()
    assert dlv.count_csv_rows(str(d)) == -1
    item = dlv._check_file(str(d), '2026-09-03', expected_rows=5)
    assert item['state'] == 'FAILED'
    assert 'could not be checked' in ' '.join(item['why'])


def test_a_QUOTED_EMBEDDED_NEWLINE_counts_as_ONE_row(tmp_path):
    """THE HAZARD IN THE COUNTER ITSELF.  The side-lists carry free-text columns
    (`forensicReason`, `M_drivers`), and a newline inside a quoted field would make a
    line-count read HIGH -- which on the `<` comparison HIDES a truncation and on the `>`
    comparison manufactures a spurious DEGRADED.  `csv.reader` is why this is one row."""
    p = tmp_path / 'SideList_REIT_Top100-2026-09-03_fmp_stock_CUR3K.csv'
    pd.DataFrame({'source': ['AAA'],
                  'forensicReason': ['REFUSED: accruals are not comparable\n'
                                     'on this cohort, see COHORT_FORENSIC_VALIDITY']
                  }).to_csv(str(p), index=False)
    assert dlv.count_csv_rows(str(p)) == 1
    assert io.open(str(p), encoding='utf-8').read().count('\n') > 2, (
        'the fixture does not actually contain the embedded newline')


def test_the_ROWCOUNT_KINDS_are_DERIVED_from_the_floor_table_not_listed_again():
    """`_EVIDENCE_GLOBS` versus `allowlist_patterns` is the same question asked twice and it
    drifted the moment a writer was added; this module already carries that lesson in its
    docstring.  So the set of kinds needing a row count is DERIVED: a kind with no
    calibrated byte floor is exactly a kind whose completeness a byte floor cannot judge.

    An UNRECOGNISED kind must NOT be in the set -- it matches no prefix in the table -- so a
    deliverable added later is never nagged about a count nobody taught its writer to
    declare."""
    assert dlv._ROWCOUNT_KINDS == tuple(p for p, f, _m in dlv._KIND_FLOORS
                                        if f == dlv.NONEMPTY)
    assert set(dlv._ROWCOUNT_KINDS) == {'SideList_', 'MarketCapBand_'}
    assert not 'SomeNewDeliverable-2026-09-03_fmp_x.csv'.startswith(dlv._ROWCOUNT_KINDS)


@pytest.mark.parametrize('pattern', ['SideList_*.csv', 'MarketCapBand_*.csv'])
def test_the_row_counter_AGREES_WITH_PANDAS_on_every_REAL_compact_artifact(pattern):
    """THE FIXTURE-ONLY RISK, closed on real data.  Every historical artifact of both
    compact kinds in the repo root is counted by `count_csv_rows` and by `pd.read_csv`, two
    independent parsers, and they must agree on all of them -- including the Micro band,
    whose one row carries a quoted `band_selection` sentence containing commas.

    Measured 2026-09-03: SideList_ n=10, 5..25 rows; MarketCapBand_ n=7, 1..20 rows."""
    import glob as _glob
    files = sorted(_glob.glob(os.path.join(REPO, pattern)))
    if len(files) < 5:
        pytest.skip('only %d artifacts of %s on this machine' % (len(files), pattern))
    for f in files:
        assert dlv.count_csv_rows(f) == len(pd.read_csv(f)), f
    assert min(dlv.count_csv_rows(f) for f in files) <= 5, (
        'no artifact here is small enough to exercise the tiny-cohort case')


@pytest.mark.parametrize('run_date,datasource,tickerfilter', [
    ('2026-08-07', 'fmp', 'stock_CUR3K'),
    ('2026-08-04', 'fmp', 'stock_TEST1'),
])
def test_a_REAL_historical_run_PASSES_with_its_rows_counted_INDEPENDENTLY(
        run_date, datasource, tickerfilter):
    """The false-positive control for gap 2 on real artifacts, and the non-circular version
    of it: the expectations handed to the audit come from `pd.read_csv`, a DIFFERENT parser
    from the `csv.reader` the audit uses, so an off-by-one in either would fail this.

    The known-good runs must exit 0 with the row check ARMED, not merely with it absent --
    the absent case is `test_the_audit_PASSES_a_REAL_historical_declared_set` above."""
    declared = _real_declared_set(run_date, datasource, tickerfilter)
    if len(declared) < 8:
        pytest.skip('only %d of this run\'s artifacts are on this machine' % len(declared))
    deck = os.path.join(REPO, 'presentations', 'presentation_2026-07-17.html')
    if not os.path.exists(deck):
        pytest.skip('no real deck on this machine')

    rows = {n: len(pd.read_csv(os.path.join(REPO, n))) for n in declared
            if n.startswith(('SideList_', 'MarketCapBand_'))}
    assert len(rows) >= 5, rows
    assert min(rows.values()) <= 5, ('this run has no small cohort, so it cannot exercise '
                                     'the tiny-cohort case: %s' % rows)

    res = dlv.audit_deliverables(declared, run_date, repo_root=REPO,
                                 deck_report=_ok_deck_report(deck,
                                                             dlv.count_deck_pages(deck)),
                                 row_expectations=rows)
    assert res['exit_code'] == 0, [(i['name'], i['rows'], i['expected_rows'], i['why'])
                                   for i in res['failed']]
    #  SCOPED TO THE COMPACT KINDS.  The XLSX of a historical run legitimately DEGRADES
    #  here -- no `xlsx_report` survives a run that finished weeks ago, so its page
    #  completeness is UNVERIFIED and says so.  That is a pre-existing verdict about a
    #  different kind, and asserting it away would make this test about the wrong thing.
    compact_degraded = [(i['name'], i['why']) for i in res['degraded']
                        if i['name'].startswith(('SideList_', 'MarketCapBand_'))]
    assert not compact_degraded, compact_degraded
    counted = {i['name']: (i['rows'], i['expected_rows']) for i in res['items']
               if i['name'] in rows}
    assert counted and all(a == b for a, b in counted.values()), counted


def test_writeResWrapper_DECLARES_A_ROW_COUNT_for_every_compact_file_it_declares(
        tmp_path, monkeypatch):
    """THE PRODUCER SIDE, END TO END.  A row check the producer never populates is an inert
    guard, and an inert guard is this project's most-logged test defect -- so this drives
    the real `writeResWrapper` and then audits its real output with its own declaration.

    It is also the second false-positive control on the producer side: the writer's actual
    files, judged by the writer's actual counts, must exit 0 -- including the ONE-ROW
    Micro band, which is the case any floor would have failed."""
    resdic = _drive_writeResWrapper(tmp_path, monkeypatch)
    declared = resdic['declared_deliverables']
    rows = resdic['deliverable_rows']
    compact = [n for n in declared if n.startswith(('SideList_', 'MarketCapBand_'))]
    assert len(compact) == 4, compact
    assert set(compact) == set(rows), (
        'a compact deliverable was declared with no row count: %s'
        % sorted(set(compact) - set(rows)))
    assert sorted(rows.values()) == [1, 1, 2, 3], rows

    deck = _deck(tmp_path / ('presentation_%s.html' % _today()), 5)
    res = dlv.audit_deliverables(compact, _today(), repo_root=str(tmp_path),
                                 deck_report=_ok_deck_report(deck, 5),
                                 row_expectations=rows)
    assert res['exit_code'] == 0, [(i['name'], i['rows'], i['expected_rows'], i['why'])
                                   for i in res['failed']]
    assert not res['degraded'], [(i['name'], i['why']) for i in res['degraded']]


def test_Sbocker_HANDS_the_row_declaration_to_the_audit():
    """The wiring.  The producer declaring counts nobody passes to the auditor is a guard
    that cannot fire, which is exactly the failure mode this file has logged four times.
    Parsed, not imported: importing the orchestrator drags the whole pipeline graph in."""
    import ast
    with io.open(os.path.join(REPO, 'Sbocker.py'), encoding='utf-8') as fh:
        tree = ast.parse(fh.read())
    main = next(n for n in tree.body
                if isinstance(n, ast.FunctionDef) and n.name == 'main')
    call = next(n for n in ast.walk(main)
                if isinstance(n, ast.Call)
                and getattr(n.func, 'attr', None) == 'audit_deliverables')
    kw = {k.arg for k in call.keywords}
    assert 'row_expectations' in kw, (
        'the audit is called without the row declaration, so the compact-deliverable '
        'check can never fire in a real run: %s' % sorted(kw))
    src = next(k.value for k in call.keywords if k.arg == 'row_expectations')
    assert ast.dump(src).count('deliverable_rows') == 1, ast.dump(src)


# =========================================================================== #
#  12.  THE POST-PICK ANALYSIS SUITE IS GATED OFF (CEO ruling, 2026-09-03)     #
# =========================================================================== #
#  IN THIS FILE because it is the same `Sbocker.main` wiring block as the deck stage and
#  the deliverable audit above, and because the two rulings pull in opposite directions and
#  the pair only makes sense read together: BUILDING THE DELIVERABLES IS NOT OPTIONAL, and
#  RUNNING THE DIAGNOSTICS IS.  The suite wrote ~3,000 of a 5,222-line run log and re-ran
#  the carve 8+ times over the merged panel, and no per-run action depended on any of it.
#
#  THE FAILURE MODE THESE TESTS GUARD IS NOT THE SKIP, IT IS THE SILENCE.  A stage that
#  quietly stops happening is how the deck went un-rendered from 2026-07-17 to 2026-09-02.

def test_the_analysis_suite_is_OFF_by_default_and_ON_with_the_flag():
    """Driven over the REAL config parser, because "default off" is a claim about parsing.
    `-run_analysis` follows `-run_estimation`'s valued shape exactly, so a bare flag with no
    value raises its own named error rather than defaulting to something."""
    import configuration as cf
    assert cf.getDataFetchConfiguration(['x'])['run_analysis'] == 0
    assert cf.getDataFetchConfiguration(['x', '-run_analysis', '1'])['run_analysis'] == 1
    assert cf.getDataFetchConfiguration(['x', '-run_analysis', '0'])['run_analysis'] == 0
    with pytest.raises(Exception) as e:
        cf.getDataFetchConfiguration(['x', '-run_analysis'])
    assert '-run_analysis requires an integer argument' in str(e.value)


def test_the_analysis_flag_does_not_disturb_the_INNER_estimation_gate():
    """`-run_estimation` is the inner gate on the heavy parameter SEARCH inside the suite
    and must keep its own default and its own meaning.  Two gates that quietly collapsed
    into one would make `-run_estimation 1` silently start running the whole suite."""
    import configuration as cf
    c = cf.getDataFetchConfiguration(['x', '-run_estimation', '1'])
    assert c['run_estimation'] == 1 and c['run_analysis'] == 0, c['run_analysis']


def _main_tree():
    import ast
    with io.open(os.path.join(REPO, 'Sbocker.py'), encoding='utf-8') as fh:
        tree = ast.parse(fh.read())
    return ast, next(n for n in tree.body
                     if isinstance(n, ast.FunctionDef) and n.name == 'main')


def test_Sbocker_GATES_the_analysis_suite_on_run_analysis():
    """The suite call must sit under a `configdic.get('run_analysis')` test.  Parsed, not
    imported: importing the orchestrator drags the whole pipeline import graph in.

    WHAT THIS CANNOT DETECT: that the gate is in the right PLACE.  It proves the call is
    gated, not that the deliverables, the pick-log and the audit around it are untouched --
    the neighbouring `test_Sbocker_runs_the_deck_stage_and_audits_with_a_nonzero_exit` is
    what holds those in place."""
    ast, main = _main_tree()
    gates = [n for n in ast.walk(main)
             if isinstance(n, ast.If)
             and 'run_analysis' in ast.dump(n.test)
             and 'run_analysis_suite' in ast.dump(ast.Module(body=n.body,
                                                             type_ignores=[]))]
    assert len(gates) == 1, (
        'the analysis suite is not gated on run_analysis (%d matching guards)' % len(gates))
    gate = gates[0]
    assert 'configdic' in ast.dump(gate.test), ast.dump(gate.test)
    #  and the suite is called NOWHERE ELSE in main, so the gate cannot be bypassed
    calls = [n for n in ast.walk(main)
             if isinstance(n, ast.Call)
             and getattr(n.func, 'attr', None) == 'run_analysis_suite']
    assert len(calls) == 1, 'run_analysis_suite is called %d times in main()' % len(calls)


def test_the_SKIP_is_announced_and_names_the_flag_that_runs_it():
    """THE POINT OF THE WHOLE SECTION.  A silently-skipped stage is the 2026-07-17 deck
    failure: the operator cannot tell "did not run" from "ran and produced nothing", and
    six weeks of work sat unrendered on that ambiguity.  So the else-branch must print, and
    the line must carry the flag -- a skip notice that does not say how to un-skip sends the
    reader to the source."""
    ast, main = _main_tree()
    gate = next(n for n in ast.walk(main)
                if isinstance(n, ast.If)
                and 'run_analysis' in ast.dump(n.test)
                and 'run_analysis_suite' in ast.dump(ast.Module(body=n.body,
                                                                type_ignores=[])))
    assert gate.orelse, 'the skip path is SILENT -- there is no else branch'
    prints = [n for n in ast.walk(ast.Module(body=gate.orelse, type_ignores=[]))
              if isinstance(n, ast.Call) and getattr(n.func, 'id', None) == 'print']
    assert len(prints) == 1, 'the skip must be ONE line, not %d' % len(prints)
    text = ' '.join(c.value for c in ast.walk(prints[0])
                    if isinstance(c, ast.Constant) and isinstance(c.value, str))
    assert 'SKIPPED' in text, text
    assert '-run_analysis 1' in text, text
    #  and it explains WHY, so the operator is not left guessing whether it broke
    assert 'no per-run action depended' in text, text


def test_the_skip_notice_tells_run_estimation_it_is_the_INNER_gate():
    """`-run_estimation 1` alone now runs nothing, and an operator who set it deserves to be
    told that rather than to watch a silent run.  This is the one branch in the skip line,
    and it exists because a flag that silently stops working is worse than one that is
    rejected."""
    ast, main = _main_tree()
    gate = next(n for n in ast.walk(main)
                if isinstance(n, ast.If)
                and 'run_analysis' in ast.dump(n.test)
                and 'run_analysis_suite' in ast.dump(ast.Module(body=n.body,
                                                                type_ignores=[])))
    dumped = ast.dump(ast.Module(body=gate.orelse, type_ignores=[]))
    assert 'run_estimation' in dumped, (
        'the skip line says nothing to an operator who set -run_estimation')
