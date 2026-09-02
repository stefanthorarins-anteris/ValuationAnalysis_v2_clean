"""The run's own check that it produced what it promised -- and the HTML deck stage.

TWO PRODUCTION FAILURES THIS MODULE EXISTS FOR
----------------------------------------------
1.  THE DECK WAS NEVER BUILT BY A RUN.  `generate_presentation.py` is a hand-run tool.
    The last deck rendered from a real run is dated 2026-07-17; six and a half weeks of
    deck fixes (the ten unearned green Sloan ticks turned amber, absence-outranks-low-
    confidence, the `nan`-string bug, the cohort forensic extension and its twenty stated
    applicability reasons) were all measured OFFLINE against saved artifacts and have
    never appeared on a page the CEO opened.  `run_deck_stage` below makes the deck a
    pipeline stage, so "shipped" and "rendered" stop being different things.

2.  A DELIVERABLE COULD FAIL COMPLETELY AND THE RUN STILL EXITED CLEAN.  On 2026-09-01
    `PresentationTop20-...xlsx` was written with ALL 20 OF 20 PAGES SKIPPED: 4,797 bytes
    and a single empty `Sheet`, against 43,482 bytes and 20 named sheets the day before.
    It was reported at WARNING severity on line 2209 of a 5,222-line log, there was no
    non-zero exit, and the transfer log printed `(0.00 MB)` beside the `0.04 MB` of every
    prior run and said nothing.  `AggScoreTop100` lost `price`, `beta`, `sector` and
    `rating_fmp` on 88 of 97 rows in the same event.  `audit_deliverables` below turns
    that class of event into a banner at the tail of the log and a non-zero exit status.

WHY THE EXPECTED SET IS **DECLARED BY THE PRODUCER**, NOT LISTED HERE
--------------------------------------------------------------------
`postBo.writeResWrapper` already builds the list of files it means to write and already
records it in the RunProvenance sidecar.  This module audits THAT list.  A second
hand-maintained enumeration of the deliverables would be the identical defect the repo
has already logged twice: `conftest._EVIDENCE_GLOBS` and `Sbocker.allowlist_patterns`
are "the same question asked twice" and drifted apart the moment a writer was added.
So the rule here is: the producer declares, this module measures, and the two are never
the same list.

WHAT THE AUDIT CANNOT DETECT -- stated up front, because a guard blind to its own target
is this project's most-logged test defect:
  * A DELIVERABLE NOBODY DECLARED.  The audit measures the declared set.  A writer that
    produces a file and puts it in no list is unaudited (and also unshipped -- see the
    Sbocker manifest note, where the transfer gap is the louder symptom).
  * THE OPTIONALLY-APPENDED FILES.  `writeResWrapper` appends the side-list, band and
    review-reference filenames to its list ONLY AFTER writing them, so their absence
    removes them from the declaration too and this audit can never see it.  Only the
    three unconditional names (AggScore CSV, presentation XLSX, forensic CSV) plus the
    postRank pickle and the deck can be checked for ABSENCE.
  * A DELIVERABLE THAT IS COMPLETE AND WRONG.  Every check here is structural -- the file
    exists, is not a stub, has its sheets/pages/columns.  A full XLSX carrying twenty
    pages of bad numbers passes every check in this module.
  * CONTENT, FOR MOST OF THE SET.  Only THREE deliverables get a structural check: the
    presentation XLSX (per-ticker sheet count), the deck (name-page count) and the
    AggScore CSV (the vendor block).  The forensic CSV, the side-lists, the band CSVs, the
    RunProvenance sidecar and the postRank pickle get existence plus a byte floor and
    nothing more -- so a forensic CSV with a header and no rows is caught (byte floor) but
    one with rows and no verdicts is not.
  * WHETHER CWD IS THE REPO ROOT.  The pipeline writes its top-N deliverables with bare
    relative filenames (so: to CWD) while `transfer_utils.EVIDENCE_DIR` and this module
    resolve to the MODULE's directory.  The two coincide in the only supported
    configuration -- Sbocker launched from the repo root -- and if they ever diverge this
    audit looks in the module directory and reports the declared files absent.  That is
    loud and wrong-in-the-safe-direction, not silent, but it is not a detection of the
    real fault.
  * ANYTHING ABOUT THE EXIT CODE BEING ACTED ON.  Checked 2026-09-02: nothing in this
    repo calls `Sbocker.main()` programmatically and there is no .bat/.cmd/.ps1 wrapper,
    so today the non-zero status is read only by the operator's shell.  Its present value
    is that it cannot be missed the way a mid-log WARNING was; its future value is for a
    scheduler.  The BANNER is doing most of the work right now.
  * WHY a page was skipped, beyond what the producer reported.  The quota attribution
    below reads the run's own HTTP tally; it is EVIDENCE, not proof, and it says so on
    the banner.
"""

import io
import os
import subprocess
import sys

#  The deck's machine-readable self-report.  ONE line on stdout, parsed by
#  `run_deck_stage`.  It carries what the deck INTENDED (`expected_pages`); the page
#  count the audit uses is measured off the written FILE, so a deck that reports success
#  and writes half a document is still caught.  Producer declares, auditor measures.
DECK_MANIFEST_PREFIX = 'DECK-MANIFEST '

#  Byte floors.  THEIR JOB IS NARROW: catch a file too small to contain anything at all
#  -- a truncated write, an exception between `open` and the first row.  A FLOOR IS NOT A
#  COMPLETENESS CHECK.  Completeness is the structural checks below (sheet count, page
#  count, vendor block); those three kinds carry a floor only as a coarse backstop, and
#  the 2026-09-01 XLSX at 4,797 bytes is the standing proof that no defensible floor
#  catches an incompleteness a structural check is meant to catch.
#
#  PER KIND, NOT PER EXTENSION -- and the per-extension version of this table was an S1
#  defect found in review on 2026-09-02.  A single `.csv` floor of 1,024 bytes FAILED 8 OF
#  13 DELIVERABLES on the known-good 2026-08-07 run and 7 on 2026-08-04, so every healthy
#  run would have exited 2 and the operator would have learned to ignore the banner within
#  two runs -- the precise alarm-fatigue failure the partial-page argument in
#  `Sbocker.main` claims to have ruled out.  That argument was about XLSX PAGE COUNTS and
#  did not transfer; the floor had been calibrated by eye against the three large
#  unconditional deliverables and never against the compact ones.
#
#  THE POPULATION EACH FLOOR IS CALIBRATED AGAINST.  Measured 2026-09-02 over every
#  historical artifact of each kind present in the repo root, so the scope of each number
#  is visible rather than remembered:
#
#      kind                       n        min      median         max
#      AggScoreTop*.csv           4     11,003      11,282      33,762
#      PresentationTop*.xlsx      4     34,312      41,408      42,525
#      ForensicFlagsTop*.csv      2      8,107      23,312      23,312
#      SideList_*.csv            10        187         643         828   <- compact BY DESIGN
#      MarketCapBand_*.csv        7        217         504       1,647   <- compact BY DESIGN
#      RawMetricsTop100*.csv      3     22,537      53,273      80,730
#      CohortMetricStats*.csv     3     17,630      18,999      24,082
#      RunProvenance-*.json       2      2,364       3,472       3,472
#      postRank_*.pickle          4    1.55 MB     79.5 MB     88.5 MB
#      presentation_*.html        5    1.49 MB     1.77 MB     2.67 MB
#
#  The sample is 2-10 files per kind drawn from two universes, one of which (TEST1, 126
#  names) is a deliberately tiny test universe -- which is why the SideList minimum is
#  187 bytes and why fitting a floor near any observed minimum would be fitting to noise.
#  So each floor below is set WELL under its kind's minimum: it answers "was anything
#  written at all", not "was enough written".
#
#  THE TWO COMPACT KINDS GET `NONEMPTY` AND NOTHING MORE.  A five-row cohort side-list and
#  a two-row micro-cap band are correct at 187 and 217 bytes, and a run can legitimately
#  produce a smaller one than any yet observed (a cohort with one name).  There is no byte
#  number that separates "compact" from "truncated" for them, so the check is reduced to
#  what is actually decidable -- the file is not empty.
#
#  AN UNRECOGNISED KIND ALSO GETS `NONEMPTY`, DELIBERATELY, and this inversion is the
#  structural fix rather than the numeric one: a deliverable added to the declared list
#  later has NO measured population here, so the safe default is the weakest floor.  The
#  old table did the opposite -- it applied a number calibrated for one kind to every file
#  sharing an extension -- which is exactly how eight compact CSVs inherited a floor
#  nobody had ever checked them against.
NONEMPTY = 1

_KIND_FLOORS = (
    #  (basename prefix, floor, the kind's measured minimum -- for the message)
    ('SideList_',          NONEMPTY,      187),
    ('MarketCapBand_',     NONEMPTY,      217),
    ('AggScoreTop',        4 * 1024,   11003),
    ('ForensicFlagsTop',   2 * 1024,    8107),
    ('RawMetricsTop100',   4 * 1024,   22537),
    ('CohortMetricStats',  4 * 1024,   17630),
    ('RunProvenance-',     512,          2364),
    ('PresentationTop',    8 * 1024,   34312),   # structural sheet count is the real gate
    ('postRank_',          64 * 1024, 1550666),
    ('presentation_',      50 * 1024, 1488054),  # structural page count is the real gate
)

#  The columns whose SIMULTANEOUS death is the quota signature on the AggScore CSV.
#  All four are filled from per-name vendor calls in `writeBoAggToCSV`; on 2026-09-01 all
#  four went null together on 88 of 97 rows.  ALL of them must be dead before this fires:
#  one flaky endpoint killing one column is a different (and much more common) event than
#  the enrichment loop being throttled off, and firing on the former would be the alarm
#  fatigue that turns a hard gate into noise.
AGGSCORE_VENDOR_COLUMNS = ('price', 'beta', 'sector', 'rating_fmp')
AGGSCORE_VENDOR_NULL_FRAC = 0.80

#  openpyxl creates a workbook with one sheet called `Sheet` and `createPresentation`
#  never removes it, so every healthy XLSX carries it alongside the per-ticker sheets
#  (verified on the 08-04 and 08-07 artifacts: 21 sheets, 20 tickers + 'Sheet').  A file
#  whose ONLY sheet is this one is the 2026-09-01 shape exactly.
XLSX_FILLER_SHEET = 'Sheet'

#  One page of the deck opens with this div (`build_name_page`).  Counted rather than
#  taken from the deck's own report for the reason in DECK_MANIFEST_PREFIX above.
_DECK_PAGE_MARKER = 'class="page-header"'


# =========================================================================== #
#  A.  THE DECK STAGE                                                          #
# =========================================================================== #

def deck_out_path(run_date, repo_root=None):
    """Where the deck for <run_date> belongs.  Mirrors `generate_presentation.main`'s
    default so the stage and a hand-run produce the same path."""
    root = repo_root or os.path.dirname(os.path.abspath(__file__))
    return os.path.join(root, 'presentations', 'presentation_%s.html' % run_date)


def deck_command(run_date, repo_root=None, python_exe=None, out_path=None):
    """The exact argv the deck stage runs.

    `--augment off` AND `--no-augment` are BOTH passed, and that redundancy is
    deliberate.  `generate_presentation`'s `--augment` DEFAULTS TO ON, and its `on` path
    imports `yfinance` and fetches per page ticker.  Two flags mean a future change to
    either one's default or spelling cannot silently put network traffic into every
    pipeline run.  MEASURED 2026-09-02 on the 2026-08-04 artifacts: with these flags the
    process makes ZERO requests -- no FMP call, no Yahoo call, and no HTTP client module
    is ever imported.

    `--run-date` is PINNED.  Without it the deck resolves "the latest postRank on disk",
    which in the pipeline is almost always today's -- but "almost always" is how a run
    that failed before writing its pickle would silently re-render YESTERDAY's deck and
    have it audited as today's deliverable.
    """
    root = repo_root or os.path.dirname(os.path.abspath(__file__))
    return [python_exe or sys.executable,
            os.path.join(root, 'generate_presentation.py'),
            '--run-dir', root,
            '--run-date', str(run_date),
            '--augment', 'off',
            '--no-augment',
            '--out', out_path or deck_out_path(run_date, root)]


def run_deck_stage(run_date, repo_root=None, python_exe=None, out_path=None,
                   timeout_s=3600, verbose=True, _run=None):
    """Render the HTML deck for <run_date>.  NEVER RAISES.

    A SUBPROCESS, not an in-process import, for three reasons that are all about not
    letting this stage damage a run that has already succeeded:
      * `generate_presentation` imports `carveOut` and mutates `sys.path`.  In-process it
        would share the pipeline's `carveOut` module state -- including the live FX table
        installed at launch -- so the deck would render differently inside the pipeline
        than it does when the CEO re-runs it by hand.  A deliverable that depends on who
        called it is not reviewable.
      * The deck loads two large pickles.  A MemoryError in-process is the run's
        MemoryError; in a subprocess it is a non-zero returncode.
      * It gives a returncode and a timeout for free.  A hung deck cannot stall a
        12-hour unattended job past `timeout_s`.

    Returns a report dict; `ok` False with `reason` set on every failure path.
    """
    root = repo_root or os.path.dirname(os.path.abspath(__file__))
    out = out_path or deck_out_path(run_date, root)
    cmd = deck_command(run_date, root, python_exe, out)
    report = {'kind': 'deck', 'path': out, 'ok': False, 'reason': None,
              'expected_pages': None, 'returncode': None, 'stderr_tail': ''}

    try:
        os.makedirs(os.path.dirname(out), exist_ok=True)
    except Exception as e:
        report['reason'] = 'could not create %s (%s: %s)' % (
            os.path.dirname(out), type(e).__name__, e)
        if verbose:
            print('[DECK] %s' % report['reason'], flush=True)
        return report

    if verbose:
        #  The api_key is NOT on this command line (the deck has no key and makes no
        #  keyed call), so printing it whole is safe and it is worth printing: the
        #  offline flags are the whole no-network guarantee and they should be visible in
        #  the log rather than asserted in a comment.
        print('[DECK] %s' % ' '.join(cmd), flush=True)

    runner = _run or _default_run
    try:
        proc = runner(cmd, root, timeout_s)
    except Exception as e:
        report['reason'] = '%s: %s' % (type(e).__name__, e)
        if verbose:
            print('[DECK] FAILED to launch: %s' % report['reason'], flush=True)
        return report

    report['returncode'] = proc['returncode']
    report['stderr_tail'] = (proc['stderr'] or '')[-2000:]
    for line in (proc['stdout'] or '').splitlines():
        if line.startswith(DECK_MANIFEST_PREFIX):
            report.update(_parse_deck_manifest(line))

    if verbose:
        #  The deck's own stdout is worth keeping in the run log: it names the four
        #  artifacts it resolved, which is the only record that the deck read THIS run.
        for line in (proc['stdout'] or '').splitlines():
            print('[DECK] %s' % line, flush=True)
        if proc['stderr']:
            print('[DECK] stderr:\n%s' % report['stderr_tail'], file=sys.stderr, flush=True)

    if proc['returncode'] != 0:
        report['reason'] = 'generate_presentation.py exited %s' % proc['returncode']
        return report
    report['ok'] = True
    return report


def _default_run(cmd, cwd, timeout_s):
    p = subprocess.run(cmd, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                       timeout=timeout_s)
    return {'returncode': p.returncode,
            'stdout': p.stdout.decode('utf-8', 'replace'),
            'stderr': p.stderr.decode('utf-8', 'replace')}


def _parse_deck_manifest(line):
    """`DECK-MANIFEST k=v k=v ...` -> dict, ints where they parse.  A malformed manifest
    yields {} rather than raising: it is a self-report, and the audit's page count comes
    off the FILE, so losing it costs one cross-check and not the stage."""
    out = {}
    for tok in line[len(DECK_MANIFEST_PREFIX):].split():
        if '=' not in tok:
            continue
        k, v = tok.split('=', 1)
        if k not in ('expected_pages', 'rendered_pages', 'augment', 'net_clients'):
            continue
        try:
            out[k] = int(v)
        except ValueError:
            out[k] = v
    return out


# =========================================================================== #
#  B.  THE STRUCTURAL CHECKS                                                   #
# =========================================================================== #

def count_deck_pages(path):
    """Per-name pages in a rendered deck.  Counts the marker in the FILE, deliberately
    not the deck's self-report.  -1 when the file cannot be read."""
    try:
        with io.open(path, 'r', encoding='utf-8', errors='replace') as fh:
            return fh.read().count(_DECK_PAGE_MARKER)
    except Exception:
        return -1


def xlsx_page_report(path):
    """{'sheets': n_total, 'pages': n_per_ticker, 'names': [...]} or {'error': ...}.

    `read_only=True` so a 40 KB workbook is not fully materialised, and the workbook is
    closed on every path -- an open read-only workbook holds a file handle, and this runs
    on the machine that is about to copy the same file to Drive."""
    try:
        import openpyxl
    except Exception as e:                                  # pragma: no cover
        return {'error': 'openpyxl unavailable (%s: %s)' % (type(e).__name__, e)}
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        names = [s for s in wb.sheetnames if s != XLSX_FILLER_SHEET]
        return {'sheets': len(wb.sheetnames), 'pages': len(names), 'names': names}
    except Exception as e:
        return {'error': '%s: %s' % (type(e).__name__, e)}
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def aggscore_vendor_report(path, columns=AGGSCORE_VENDOR_COLUMNS,
                           null_frac=AGGSCORE_VENDOR_NULL_FRAC):
    """Is the AggScore CSV's vendor-enrichment block alive?

    Returns {'rows': n, 'null_frac': {col: f}, 'dead': [cols], 'all_dead': bool} or
    {'error': ...}.  A MISSING column counts as dead: on 2026-09-01 the columns were
    present-and-null, but a future failure that drops them entirely is the same event."""
    try:
        import pandas as pd
        df = pd.read_csv(path)
    except Exception as e:
        return {'error': '%s: %s' % (type(e).__name__, e)}
    if df.empty:
        return {'rows': 0, 'null_frac': {}, 'dead': list(columns), 'all_dead': True}
    fracs, dead = {}, []
    for c in columns:
        if c not in df.columns:
            fracs[c] = 1.0
            dead.append(c)
            continue
        #  `isna()` CARRIES THIS, and the first version of this check did not believe it.
        #  The pipeline writes the STRING 'NaN' as its absence sentinel in these columns,
        #  so the obvious worry is that `isna()` reads a degraded column as populated.
        #  MEASURED: it does not.  `pd.read_csv`'s default `na_values` already contains
        #  'NaN', 'nan', 'NA', 'None', 'null' and '', so every sentinel this pipeline
        #  writes is a real NaN by the time it is read here.  A string-matching layer on
        #  top of that is an inert guard -- it can never be the thing that fires -- so it
        #  was removed rather than kept as reassurance.
        #
        #  THE ONE RESIDUAL IS WHITESPACE.  A cell holding a single space is NOT in
        #  `na_values` and survives as a string, so that half is real and is kept.
        col = df[c]
        blank = col.isna() | (col.astype(str).str.strip() == '')
        f = float(blank.mean())
        fracs[c] = f
        if f >= null_frac:
            dead.append(c)
    return {'rows': int(len(df)), 'null_frac': fracs, 'dead': dead,
            'all_dead': len(dead) == len(columns)}


# =========================================================================== #
#  C.  THE AUDIT                                                               #
# =========================================================================== #

def _floor_message(size, floor, measured):
    """Name the population the floor came from.  A refusal that cites only a threshold
    leaves the reader unable to judge whether the threshold or the file is wrong -- which
    is how a floor calibrated against three large deliverables came to be applied to eight
    compact ones without anyone being able to see it from the message."""
    if measured is None:
        return ('%d bytes -- the file is EMPTY. This kind has no calibrated floor here, '
                'so the only check applied was "something was written".' % size)
    return ('%d bytes, below the %d-byte floor for this kind. The floor is calibrated '
            'against every historical artifact of this kind on disk (smallest observed: '
            '%d bytes) and is set well under it, so it means "nothing was written", not '
            '"not enough was written".' % (size, floor, measured))


def _min_bytes(path):
    """(floor, measured_min_for_this_kind) for <path>.  Longest matching prefix wins, so
    `PresentationTop...` cannot be captured by a shorter entry added later.  No match ->
    NONEMPTY: an unmeasured kind gets the weakest floor, never an inherited one."""
    base = os.path.basename(path)
    best = None
    for prefix, floor, measured in _KIND_FLOORS:
        if base.startswith(prefix) and (best is None or len(prefix) > len(best[0])):
            best = (prefix, floor, measured)
    if best is None:
        return NONEMPTY, None
    return best[1], best[2]


def audit_deliverables(declared, run_date, repo_root=None, deck_report=None,
                       xlsx_report=None, http_tally=None):
    """Check every declared deliverable and return a verdict.

    `declared`      -- the producer's list of filenames (postBo.writeResWrapper's return,
                       plus the postRank pickle).  Relative names resolve against
                       `repo_root`, matching where the pipeline writes them.
    `deck_report`   -- `run_deck_stage`'s report, or None when the stage did not run.
    `xlsx_report`   -- `createPresentation`'s report, i.e. how many pages it MEANT to
                       write.  Absent -> the XLSX is checked for emptiness only, and the
                       banner says the expectation was unavailable rather than implying
                       a count was met.
    `http_tally`    -- a `getData_gen.http_tally_delta()` dict covering the deliverable
                       stages, used ONLY to attribute a failure to throttling.

    Verdict: {'failed': [...], 'degraded': [...], 'ok': [...], 'exit_code': 0|2,
              'quota_suspected': bool, 'items': [...]}
    """
    root = repo_root or os.path.dirname(os.path.abspath(__file__))
    items = []

    for name in list(declared or []):
        path = name if os.path.isabs(name) else os.path.join(root, name)
        items.append(_check_file(path, run_date, xlsx_report=xlsx_report))

    #  THE DECK IS AUDITED WHETHER OR NOT THE STAGE REPORTED SUCCESS, and that ordering
    #  matters: the 2026-07-17 failure is precisely a deck that nothing claimed to have
    #  written.  A stage that never ran is a FAILED deliverable, not an absent check.
    items.append(_check_deck(run_date, root, deck_report))

    failed = [i for i in items if i['state'] == 'FAILED']
    degraded = [i for i in items if i['state'] == 'DEGRADED']
    ok = [i for i in items if i['state'] == 'OK']

    quota = _quota_suspected(items, http_tally)
    for i in failed + degraded:
        if quota and i.get('shortfall'):
            i['attribution'] = 'THROTTLE-SUSPECTED'
        elif i.get('missing_input'):
            i['attribution'] = 'MISSING-INPUT'

    return {'run_date': run_date, 'items': items, 'failed': failed,
            'degraded': degraded, 'ok': ok,
            'quota_suspected': quota, 'http_tally': http_tally,
            #  EXIT CODE 2, not 1.  An unhandled exception already leaves this pipeline
            #  with a non-zero status, so reusing 1 would make "the run crashed" and "the
            #  run finished and its deliverables are broken" indistinguishable to
            #  anything reading the status -- which is the whole failure being fixed.
            'exit_code': 2 if failed else 0}


def _check_file(path, run_date, xlsx_report=None):
    base = os.path.basename(path)
    item = {'name': base, 'path': path, 'state': 'OK', 'why': [],
            'bytes': None, 'expected': None, 'actual': None,
            'shortfall': False, 'missing_input': False}

    if not os.path.exists(path):
        item['state'] = 'FAILED'
        item['why'].append('DOES NOT EXIST -- the producer declared it and no file is there')
        return item

    try:
        item['bytes'] = os.path.getsize(path)
    except Exception as e:
        item['state'] = 'FAILED'
        item['why'].append('cannot stat (%s: %s)' % (type(e).__name__, e))
        return item

    floor, measured = _min_bytes(path)
    if item['bytes'] < floor:
        item['state'] = 'FAILED'
        item['why'].append(_floor_message(item['bytes'], floor, measured))

    if base.lower().endswith('.xlsx') and base.startswith('PresentationTop'):
        rep = xlsx_page_report(path)
        if 'error' in rep:
            item['state'] = 'FAILED'
            item['why'].append('unreadable as a workbook (%s)' % rep['error'])
            return item
        item['actual'] = rep['pages']
        if rep['pages'] == 0:
            item['state'] = 'FAILED'
            item['why'].append('STRUCTURALLY EMPTY -- %d sheet(s), none of them a ticker '
                               'page (this is the 2026-09-01 shape exactly)' % rep['sheets'])
            item['shortfall'] = True
        elif xlsx_report and xlsx_report.get('expected'):
            item['expected'] = int(xlsx_report['expected'])
            if rep['pages'] < item['expected']:
                item['state'] = 'FAILED'
                item['why'].append(
                    'PARTIAL -- %d of %d ticker page(s); missing: %s'
                    % (rep['pages'], item['expected'],
                       ', '.join(map(str, xlsx_report.get('skipped') or [])) or 'unnamed'))
                item['shortfall'] = True
        else:
            item['why'].append('%d ticker page(s); the producer reported no expected '
                               'count, so completeness is UNVERIFIED' % rep['pages'])
            if item['state'] == 'OK':
                item['state'] = 'DEGRADED'

    elif base.startswith('AggScoreTop') and base.lower().endswith('.csv'):
        rep = aggscore_vendor_report(path)
        if 'error' in rep:
            item['state'] = 'FAILED'
            item['why'].append('unreadable as a CSV (%s)' % rep['error'])
            return item
        item['actual'] = rep['rows']
        if rep['rows'] == 0:
            item['state'] = 'FAILED'
            item['why'].append('ZERO ROWS')
            item['shortfall'] = True
        elif rep['all_dead']:
            item['state'] = 'FAILED'
            item['why'].append(
                'VENDOR BLOCK DEAD -- %s all >=%.0f%% blank over %d rows (%s). Every one '
                'of these is filled by a per-name vendor call; all four dying together '
                'is the 2026-09-01 signature.'
                % (', '.join(rep['dead']), AGGSCORE_VENDOR_NULL_FRAC * 100, rep['rows'],
                   ', '.join('%s=%.0f%%' % (c, f * 100)
                             for c, f in sorted(rep['null_frac'].items()))))
            item['shortfall'] = True
        elif rep['dead']:
            item['state'] = 'DEGRADED' if item['state'] == 'OK' else item['state']
            item['why'].append('vendor column(s) %s >=%.0f%% blank over %d rows'
                               % (', '.join(rep['dead']),
                                  AGGSCORE_VENDOR_NULL_FRAC * 100, rep['rows']))
            item['shortfall'] = True

    return item


def _check_deck(run_date, root, deck_report):
    path = (deck_report or {}).get('path') or deck_out_path(run_date, root)
    item = {'name': os.path.basename(path), 'path': path, 'state': 'OK', 'why': [],
            'bytes': None, 'expected': None, 'actual': None,
            'shortfall': False, 'missing_input': False}

    if deck_report is None:
        item['state'] = 'FAILED'
        item['why'].append('THE DECK STAGE DID NOT RUN. This is the 2026-07-17 failure: '
                           'no run has rendered the deck since, because nothing in the '
                           'pipeline asked it to.')
        return item

    #  A MODE THAT CANNOT PRODUCE A DECK IS NOT A BROKEN RUN (2026-09-02, review S2).
    #  `-loadboresults` force-disables `saveBoResults` (configuration.py:348-351), so no
    #  same-date Boresults pickle is written, and `resolve_run_artifacts` REFUSES to fall
    #  back to another date's -- correctly, that cross-run mixing was a publish-blocker.
    #  The deck is therefore unavailable in that mode by construction.  Calling it FAILED
    #  would exit 2 on a supported ZERO-QUOTA replay, which is the same alarm-fatigue
    #  error as the byte floor: the operator reaching for the cheapest mode is exactly the
    #  operator who must still trust the banner.  DEGRADED: loud, named, exit 0.
    if deck_report.get('not_applicable'):
        item['state'] = 'DEGRADED'
        item['why'].append('DECK NOT BUILT THIS RUN -- %s. This is a property of the run '
                           'mode, not a failure; the deck can be rendered by hand from '
                           'the loaded run with generate_presentation.py --run-date.'
                           % (deck_report.get('reason') or 'no reason recorded'))
        return item

    if not deck_report.get('ok'):
        item['state'] = 'FAILED'
        item['why'].append('deck stage failed: %s' % (deck_report.get('reason')
                                                      or 'no reason recorded'))
        tail = (deck_report.get('stderr_tail') or '').strip().splitlines()
        if tail:
            item['why'].append('last stderr line: %s' % tail[-1][:300])
        #  A deck failure caused by a MISSING SAME-DATE ARTIFACT is a different event
        #  from a deck failure caused by throttling, and `resolve_run_artifacts` says so
        #  in the exception text it raises.  Read it rather than guessing.
        if 'not found' in (deck_report.get('stderr_tail') or '').lower() \
                or 'FileNotFoundError' in (deck_report.get('stderr_tail') or ''):
            item['missing_input'] = True

    if not os.path.exists(path):
        item['state'] = 'FAILED'
        item['why'].append('no file at %s' % path)
        return item

    item['bytes'] = os.path.getsize(path)
    floor, measured = _min_bytes(path)
    if item['bytes'] < floor:
        item['state'] = 'FAILED'
        item['why'].append(_floor_message(item['bytes'], floor, measured))

    pages = count_deck_pages(path)
    item['actual'] = pages
    if pages <= 0:
        item['state'] = 'FAILED'
        item['why'].append('NO NAME PAGES -- the document rendered with zero per-ticker '
                           'sections')
        item['shortfall'] = True
    else:
        exp = deck_report.get('expected_pages')
        if isinstance(exp, int) and exp > 0:
            item['expected'] = exp
            if pages < exp:
                item['state'] = 'FAILED'
                item['why'].append('PARTIAL -- %d of %d name page(s) rendered'
                                   % (pages, exp))
                item['shortfall'] = True
        else:
            item['why'].append('%d name page(s); the deck reported no expected count, so '
                               'completeness is UNVERIFIED' % pages)
            if item['state'] == 'OK':
                item['state'] = 'DEGRADED'
    return item


def _quota_suspected(items, http_tally):
    """Was this run being throttled while the deliverables were written?

    EVIDENCE, NOT PROOF, and the banner labels it that way.  A 429 count from the
    deliverable window says the vendor was refusing calls; it does not establish that
    THIS page was lost to THAT refusal.  The distinction the CEO asked for is between
    "the vendor said no" and "an artifact was missing", and this is enough to draw it."""
    if not http_tally:
        return False
    by_status = http_tally.get('by_status') or {}
    throttled = int(by_status.get(429, 0) or 0) + int(by_status.get('429', 0) or 0)
    if throttled <= 0:
        return False
    return any(i.get('shortfall') for i in items)


# =========================================================================== #
#  D.  THE BANNER                                                              #
# =========================================================================== #

def format_banner(result):
    """The tail-of-log verdict.  It must be readable without grepping: the 2026-09-01
    failure WAS reported, at WARNING severity, on line 2209 of 5,222."""
    bar = '!' * 78
    lines = []
    failed, degraded, ok = result['failed'], result['degraded'], result['ok']
    n = len(result['items'])

    if not failed and not degraded:
        lines.append('=' * 78)
        lines.append('DELIVERABLE AUDIT: ALL %d DELIVERABLE(S) PRESENT AND STRUCTURALLY '
                     'COMPLETE (run %s)' % (n, result['run_date']))
        for i in result['items']:
            lines.append('    OK       %-52s %s' % (i['name'], _size_note(i)))
        lines.append('=' * 78)
        return '\n'.join(lines)

    lines.append(bar)
    if failed:
        lines.append('!!! DELIVERABLE AUDIT FAILED: %d of %d DELIVERABLE(S) ARE MISSING '
                     'OR BROKEN' % (len(failed), n))
    else:
        lines.append('!!! DELIVERABLE AUDIT: %d of %d DELIVERABLE(S) DEGRADED (none '
                     'broken)' % (len(degraded), n))
    lines.append('!!! run %s   exit status will be %d'
                 % (result['run_date'], result['exit_code']))
    lines.append('!!!')
    for i in failed + degraded:
        lines.append('!!! %-8s %s   %s' % (i['state'], i['name'], _size_note(i)))
        for w in i['why']:
            for chunk in _wrap(w, 68):
                lines.append('!!!            %s' % chunk)
        if i.get('attribution') == 'THROTTLE-SUSPECTED':
            lines.append('!!!            CAUSE: the vendor was REFUSING CALLS during this '
                         'run (see below).')
            lines.append('!!!            This is a QUOTA event, not a code defect. Re-run '
                         'when quota resets.')
        elif i.get('attribution') == 'MISSING-INPUT':
            lines.append('!!!            CAUSE: an INPUT ARTIFACT for this run-date was '
                         'absent, not a failed call.')
    if result.get('quota_suspected'):
        t = result.get('http_tally') or {}
        by = t.get('by_status') or {}
        lines.append('!!!')
        lines.append('!!! HTTP during the deliverable stages: %s call(s), %s x 429, '
                     '%s non-200'
                     % (t.get('calls'), by.get(429, by.get('429', 0)),
                        t.get('non_200')))
        lines.append('!!! Every retry is a separate charged request, so the call count '
                     'above is')
        lines.append('!!! requests spent, not names attempted.')
    if ok:
        lines.append('!!!')
        lines.append('!!! intact: %s' % ', '.join(i['name'] for i in ok))
    lines.append(bar)
    return '\n'.join(lines)


def _size_note(i):
    bits = []
    if i.get('bytes') is not None:
        bits.append('%s bytes' % '{:,}'.format(i['bytes']))
    if i.get('actual') is not None:
        bits.append('%s/%s' % (i['actual'],
                               i['expected'] if i.get('expected') is not None else '?'))
    return '(%s)' % ', '.join(bits) if bits else ''


def _wrap(text, width):
    out, line = [], ''
    for word in str(text).split():
        if line and len(line) + 1 + len(word) > width:
            out.append(line)
            line = word
        else:
            line = (line + ' ' + word).strip()
    if line:
        out.append(line)
    return out or ['']


def emit_audit(result, out=None, err=None):
    """Print the banner to BOTH streams and return the exit code.

    BOTH, deliberately.  `run_logger` tees stdout into the run log (so the banner is in
    the artifact the house inspects) while stderr is what survives a console that only
    shows errors and what a scheduler captures separately.  The 2026-09-01 event was
    visible on neither."""
    banner = format_banner(result)
    (out or sys.stdout).write('\n' + banner + '\n')
    (out or sys.stdout).flush()
    if result['failed'] or result['degraded']:
        (err or sys.stderr).write('\n' + banner + '\n')
        (err or sys.stderr).flush()
    return result['exit_code']
