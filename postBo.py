import calcMetrics as cm
import calcScore as cs
import createDicts as cdic
import meanBars as mb
import nan_policy as npol
import getData_gen as gdg
import postBoRank as pbr
import reporting_period as rp
import forensicFlags as ff
#  Imported for the FORENSIC DATA-GAP charge into the ad-hoc penalty bucket (CEO, 2026-08-16),
#  which runs BEFORE Stage-2 -- the M-score display decoration still happens in `Sbocker` after
#  ranking, and this import does not move it.
import detectManipulation as dm
import pandas as pd
import requests
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import datetime, timedelta
from tqdm import tqdm
from sklearn.linear_model import LinearRegression
import numpy as np
import math
import warnings

# Suppress FutureWarning about DataFrame concatenation with empty/all-NA entries
warnings.filterwarnings('ignore', message='.*concatenation with empty or all-NA entries.*')


def generalTopN(finalBoRank_df, bands, topn, warn=True):
    """The frame the GENERAL top-N is taken from -- `head(topn)` of the result IS that list.

    When currency data is present the general top-N is the General band (>$300M), i.e.
    postRank[marketCap_usd>300e6].head(topn), so the deliverables match the banded partition.
    When banding is absent OR currency is still pending, behaviour is UNCHANGED (byte-identical
    to before): the general top-N stays postRank.head(topn), so nothing wrong ships before the
    field flows.

    Factored out of createPresentation (2026-08-04) because the industry counter has to count
    THE SAME twenty names the deck presents.  A second copy of this selection would be a
    concentration report about a list nobody sees the moment the two drifted -- so there is one
    selection, and the counter reads it with `warn=False` (the same warning printed twice for
    one run is noise, and the reporting caller is not the one that would be shrinking a list).
    """
    if bands and not bands.get('currency_pending', True):
        _gb = (bands.get('bands') or {}).get('General')
        if _gb is not None and not _gb.empty:
            # The General band is pre-capped at the MCAP_BANDS General head_N (=20). If a
            # caller ever requests MORE than that (ntopxlsx > 20), keying the list off the
            # band would SILENTLY shrink the general list -- so fall back to the unbanded
            # head(topn) and warn LOUDLY instead. No effect today (ntopxlsx == 20 == cap).
            if topn > len(_gb):
                if warn:
                    print(f"WARNING: createPresentation topn={topn} exceeds General-band size "
                          f"{len(_gb)} (MCAP_BANDS General cap); using unbanded "
                          f"postRank.head({topn}) to avoid silently shrinking the general "
                          f"list.", flush=True)
            else:
                return _gb
    return finalBoRank_df


def _diag_newest_rows(df, n=3):
    """The most recent `n` rows of a per-source frame, for a PRINT-ONLY diagnostic.

    WHY THIS IS NOT THE STRICT SORT (review L4, 2026-07-31).  The audit-C5 fix replaced
    `head(n)` -- which cannot raise -- with the forensic `_toNewestFirst`, whose
    `pd.to_datetime(s)` has NO `errors='coerce'`.  A single unparseable date in the first
    source's rows therefore raised ValueError inside `postBoWrapper`'s diagnostic block, which
    is unguarded, aborting Stage-2 and every deliverable.  Making a LOG LINE able to kill the
    run is a strictly worse defect than the mislabelling it fixed.

    The strict sort is deliberately NOT loosened: it backs the Beneish/Montier YoY shifts,
    where an unparseable date SHOULD fail loudly rather than sort as NaT.  Coercion belongs
    HERE, in the diagnostic -- and that distinction is now expressed by the policy argument
    to the shared helper (rp.ON_BAD_DATE_COERCE vs rp.ON_BAD_DATE_RAISE) instead of by which
    of two look-alike helpers the site happened to call.

    Falls back to the original `head(n)` if even the coerced sort fails, so the diagnostic
    degrades to "possibly mis-ordered" rather than to "no run".
    """
    try:
        return rp.to_newest_first(df, rp.ON_BAD_DATE_COERCE).head(n)
    except Exception:
        return df.head(n)


def _fmt4(v):
    """`"{:.4f}"` of a number, or the string 'NaN' for anything that is not one.

    Exists because `"{:.4f}".format(None)` raises TypeError and FMP genuinely returns
    `"beta": null` / `"price": null` on many non-US listings (review item 5, 2026-07-31).  bool
    is excluded deliberately: it is an int subclass, and a boolean price is corrupt data, not a
    number.  NaN in, 'NaN' out -- the column's existing sentinel, so nothing downstream changes.
    """
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        return 'NaN'
    try:
        if v != v:                      # NaN
            return 'NaN'
        return "{:.4f}".format(v)
    except (TypeError, ValueError):
        return 'NaN'


def _one_mean_score(df, source, col):
    """The single `col` value for `source`, or None when it is absent / duplicated / NaN.

    Replaces `df[df['source'] == s][col].isna().item()` + `.item()`, which raises ValueError on
    a zero-row OR multi-row selection (issuer clones give multi-row) and KeyError when the
    column is missing.  Returning None lets the caller emit the 'NaN' sentinel instead of
    aborting the stage (review item 5, 2026-07-31).
    """
    try:
        if df is None or col not in getattr(df, 'columns', []):
            return None
        sel = pd.to_numeric(df.loc[df['source'] == source, col], errors='coerce').dropna()
        if len(sel) != 1:
            return None                 # absent (0) or ambiguous (>1) -- both are 'NaN'
        return float(sel.iloc[0])
    except Exception:
        return None


def _resolve_dollarvol_floor(as_of=None, dollarvol_floor='auto'):
    """The $/day floor THIS scoring pass may apply, or None.  See `postBoWrapper`.

    A module-level function rather than four lines inline, so the decision can be exercised
    without standing a scoring run up -- the first cut's guard was a source scan precisely
    because the decision was not reachable, and a source scan is what let the defect through.
    """
    if dollarvol_floor == 'auto':
        import carveOut as _co_floor
        return _co_floor.DOLLAR_VOLUME_FLOOR_USD if as_of is None else None
    return dollarvol_floor


def postBoWrapper(dmdic, as_of=None, dollarvol_floor='auto'):
    """Scoring orchestration.  as_of (default None) threads the point-in-time date D
    from Sbocker through Stage-1 (simpleScore_fromDict) and Stage-2
    (postBoScoreRanking).  as_of=None -> live behaviour, BIT-FOR-BIT unchanged.

    `dollarvol_floor` -- whether THIS run may apply the $1M/day traded-value floor.
    `'auto'` (default) decides from `as_of`; an explicit value, INCLUDING None, always wins.
    See the resolution block below for why this is a parameter and not a test on `as_of`.
    """
    import sys
    import numpy as np

    #  --- WHETHER THIS RUN MAY APPLY THE $1M/DAY TRADED-VALUE FLOOR -------------------
    #  *** THE FIRST CUT PUT THE OPT-IN UNCONDITIONALLY INSIDE THIS FUNCTION AND CALLED THAT
    #  THE LOOKAHEAD GUARD.  IT IS NOT ONE: `postBoWrapper` IS ITSELF A POINT-IN-TIME ENTRY
    #  POINT (reviewer, 2026-08-24). ***  `carveOut.dollar_volume_frame` reads whatever volAvg
    #  profile capture is NEWEST ON DISK -- today's liquidity -- so a re-entrant
    #  scoring a past or date-filtered panel would screen it on a fact from the future, and
    #  the thing that contaminates is the backtest that measures whether the filter works.
    #  `meanBars._prior_streaks` already recorded the re-entrancy in as many words
    #  ("postBoWrapper is the production seam but it is ALSO re-entered by the offline
    #  research"); the fact was available and the first cut failed to carry it into the guard.
    #
    #  THREE REACHABLE PIT PATHS, AND `as_of` ONLY CLOSES THE FIRST:
    #    1. `Sbocker.main` passes `as_of` straight through, and `-asof YYYY-MM-DD` is a
    #       documented production flag (see `configuration.getDataFetchConfiguration`, which
    #       describes it as "run the pipeline as-of that past date (survivorship-safe PIT
    #       universe)").  `as_of is not None` -> OFF.
    #    2. `backtest_ols_analysis.run_ols_analysis` and
    #    3. `portfolio._get_top_symbols` -- the latter is what `Sbocker` invokes when
    #       `portfoliotestyear > 0` -- both build a temp dmdic over a DATE-FILTERED panel and
    #       call with `as_of=None`.  No value of `as_of` can distinguish them from a live run,
    #       so they pass `dollarvol_floor=None` EXPLICITLY.  That is the whole reason this is
    #       a parameter rather than a one-line test on `as_of`.
    #    (SYMBOLS, NOT LINE NUMBERS, and that is an enforced house rule --
    #     `test_structural_guards.test_NO_comment_in_the_scoring_files_cites_a_LINE_NUMBER`.
    #     The first cut of this block cited six line numbers and broke it; a cite corrected
    #     from 209 to 229 was stale within 308 hours, which is the case that set the rule.)
    #
    #  NOT CLOSED HERE, AND DELIBERATELY: `normalized_analysis.run_normalized_pipeline` and the
    #  `baseline_tools` re-entrants (`run_corrected_current`, `industry_attribution`,
    #  `nan_policy_report`) score TODAY's panel and exist to reproduce or A/B the live run, so
    #  today's liquidity is the right question for them and the floor stays ON.
    _dv_floor = _resolve_dollarvol_floor(as_of, dollarvol_floor)
    if dollarvol_floor == 'auto':
        _why = ('as_of is None -> treated as LIVE' if as_of is None
                else 'as_of=%s -> point-in-time, floor withheld' % (as_of,))
    else:
        _why = 'caller passed dollarvol_floor=%r explicitly' % (dollarvol_floor,)
    if _dv_floor:
        print("TRADED-VALUE FLOOR: ON at $%.0f/day (%s). The reading is TODAY's volAvg "
              "capture, so this is a statement about liquidity NOW." % (float(_dv_floor), _why),
              flush=True)
    else:
        print("TRADED-VALUE FLOOR: OFF (%s). This scoring pass is NOT screened on today's "
              "liquidity -- which is correct for a point-in-time or date-filtered panel, and "
              "is NOT a finding that the pool is liquid." % _why, flush=True)
    
    # Diagnostic: Check input data BEFORE any calculations
    print("\n" + "="*60, flush=True)
    print("DIAGNOSTIC: postBoWrapper input data check (BEFORE score calculation)", flush=True)
    print("="*60, flush=True)
    
    bmdf = dmdic['BoMetric_df']
    bmav = dmdic.get('BoMetric_ave', {})
    # UNEXERCISED OPTION, KEPT DELIBERATELY (marked 2026-08-02).  `bmda` is
    # `BoMetric_dateAve` -- the PER-DATE median baseline -- and it is threaded all the way
    # into `calcScore.simpleScore_fromDict` as a positional parameter that the function body
    # NEVER REFERENCES: Stage-1 scores every metric against the FULL-PERIOD median
    # (`BoMetric_ave`) only, so no per-date baseline reaches any score today.  This is
    # capability, not a defect, and it is NOT removed: the CEO has an open item about wiring
    # it up (a per-date baseline is what makes the Stage-1 score point-in-time rather than
    # full-sample), and deleting the plumbing would mean rebuilding it. The marker exists so
    # the next reader does not spend time working out whether it is live -- it is not.
    bmda = dmdic.get('BoMetric_dateAve', pd.DataFrame())
    cdx_df = dmdic.get('cdx_df', pd.DataFrame())
    n = dmdic.get('nrScorePeriods', 8)
    
    # Check BoMetric_df
    if bmdf.empty:
        print("ERROR: BoMetric_df is EMPTY!", flush=True)
    else:
        print(f"BoMetric_df shape: {bmdf.shape} (rows, columns)", flush=True)
        print(f"BoMetric_df unique sources: {bmdf['source'].nunique() if 'source' in bmdf.columns else 'NO SOURCE COLUMN'}", flush=True)
        if 'source' in bmdf.columns:
            first_source = bmdf['source'].iloc[0]
            print(f"BoMetric_df first source: {first_source}", flush=True)
            # Show sample rows for first source -- the MOST RECENT three.  Same audit-C5
            # defect as the cdx_df block below (NOT on the audit's list -- found in the same
            # sweep): BoMetric_df is stored OLDEST-first, so head(3) labelled the three oldest
            # periods "first 3".  Fixed in lockstep so both halves of this diagnostic block
            # mean the same thing; a block where one sample is newest-first and the other
            # oldest-first is worse than either.
            first_source_data = _diag_newest_rows(
                bmdf[bmdf['source'] == first_source], 3)
            print(f"Sample rows for {first_source} (3 MOST RECENT periods, newest first):",
                  flush=True)
            numeric_cols_sample = first_source_data.select_dtypes(include=[np.number]).columns[:5]
            if len(numeric_cols_sample) > 0:
                print(first_source_data[['date', 'source'] + list(numeric_cols_sample)].to_string(), flush=True)
        # Check for NaN in numeric columns
        numeric_cols = bmdf.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            nan_pct = (bmdf[numeric_cols].isna().sum() / len(bmdf) * 100).round(1)
            print(f"\nBoMetric_df NaN percentage in numeric columns (first 5):", flush=True)
            for col, pct in list(nan_pct.head(5).items()):
                print(f"  {col}: {pct}%", flush=True)
    
    # Check cdx_df
    if cdx_df.empty:
        print("\nERROR: cdx_df is EMPTY!", flush=True)
    else:
        print(f"\ncdx_df shape: {cdx_df.shape} (rows, columns)", flush=True)
        print(f"cdx_df unique sources: {cdx_df['source'].nunique() if 'source' in cdx_df.columns else 'NO SOURCE COLUMN'}", flush=True)
        if 'source' in cdx_df.columns:
            first_source = cdx_df['source'].iloc[0]
            print(f"cdx_df first source: {first_source}", flush=True)
            # Show sample rows for first source -- the MOST RECENT three (audit C5, fixed
            # 2026-07-31).  cdx_df is stored OLDEST-first, so head(3) printed the three
            # OLDEST quarters under the label "first 3": on the 07-17 panel that meant
            # ~2020 rows presented as the newest data.  This is the line the CEO reads at
            # 3am to judge whether the fetch actually picked up the current quarter, so
            # showing the oldest rows made it actively misleading rather than merely
            # cosmetic.  _diag_newest_rows sorts by PARSED date (coercing, and never raising --
            # see review L4), so the orientation is right whichever way the rows arrive.
            first_source_data = _diag_newest_rows(
                cdx_df[cdx_df['source'] == first_source], 3)
            print(f"Sample rows for {first_source} (3 MOST RECENT periods, newest first):",
                  flush=True)
            key_cols = ['date', 'source', 'marketCap', 'freeCashFlow', 'price', 'totalAssets']
            available_cols = [col for col in key_cols if col in first_source_data.columns]
            if len(available_cols) > 0:
                print(first_source_data[available_cols].to_string(), flush=True)
    
    # Check BoMetric_ave (could be dict, Series, or DataFrame)
    if bmav is None:
        print("\nWARNING: BoMetric_ave is None!", flush=True)
    elif isinstance(bmav, pd.DataFrame):
        if bmav.empty:
            print("\nWARNING: BoMetric_ave DataFrame is empty!", flush=True)
        else:
            print(f"\nBoMetric_ave DataFrame shape: {bmav.shape}", flush=True)
            print(f"BoMetric_ave columns: {list(bmav.columns[:5])}", flush=True)
    elif isinstance(bmav, pd.Series):
        if bmav.empty:
            print("\nWARNING: BoMetric_ave Series is empty!", flush=True)
        else:
            print(f"\nBoMetric_ave Series length: {len(bmav)}", flush=True)
            print(f"BoMetric_ave sample values (first 10):", flush=True)
            for idx, val in bmav.head(10).items():
                print(f"  {idx}: {val}", flush=True)
    elif isinstance(bmav, dict):
        if not bmav:
            print("\nWARNING: BoMetric_ave dict is empty!", flush=True)
        else:
            print(f"\nBoMetric_ave dict has {len(bmav)} keys", flush=True)
            print(f"BoMetric_ave sample values (first 10):", flush=True)
            for key, val in list(bmav.items())[:10]:
                print(f"  {key}: {val}", flush=True)
    else:
        print(f"\nBoMetric_ave type: {type(bmav)}", flush=True)
    
    print("="*60 + "\n", flush=True)
    sys.stdout.flush()
    
    # FREQUENCY MAP FROM cdx_df, not from BoMetric_df (review S3, 2026-07-26).
    # `period` is the authoritative frequency signal and it lands in cdx_df ONLY --
    # BoMetric_df never carries it by design (createDicts preReq fields do not reach
    # it).  Stage-1 was therefore the ONE consumer still deriving frequency from date
    # cadence while Stage-2, the forensics, the moat and reviewReference all used
    # `period`: two sources of truth for the same per-ticker decision, already
    # disagreeing on 2 names (KXIN, OBI.L) and set to diverge properly from the first
    # fetch that carries the field.  Passing the cdx-derived map makes Stage-1 read the
    # same answer as everything downstream.
    # THE UNIVERSE-WIDE FREQUENCY / CONFLICT BANNER (restored, ship-gate item 3, 2026-07-27).
    # verbose=True is LOAD-BEARING and independent of whether Stage-1 consumes the map:
    # this is the run's ONLY universe-wide frequency readout and the only place the
    # period-vs-cadence conflict list + CSV are emitted.  Both pre-ship reads call it "the
    # first thing to read after the fetch".  It got silenced when the map was threaded into
    # Stage-1 with verbose defaulted off, and the Q2 ruling then made Stage-1's use of the
    # map a no-op on the score -- so the wiring's only surviving effect had been to hide
    # the banner.  Keep the call and keep it loud.
    # CSV WRITING IS ENABLED HERE AND NOWHERE ELSE (fix, 2026-07-31): this is the run's
    # UNIVERSE-WIDE read, so it is the one entitled to write the shared
    # ReportingFrequencyConflicts_<date>.csv artifact.
    # The other verbose callers (postBoRank's top-100, calcScore, detectManipulation) run on
    # narrower pools and would clobber this artifact with a shorter list.  The conflict is now
    # detected even though this call is served from the stored ingest verdict, because
    # frequency_by_source decodes the stamped rp.FREQ_CONFLICT_COLUMN -- before that fix this
    # banner could only ever report zero, whatever the data said.
    _freq_map = rp.frequency_by_source(dmdic.get('cdx_df'), verbose=True, csv=True)

    # ONE RUN = ONE SET OF NaN-POLICY COUNTS (review finding, 2026-08-05).  `POLICY_COUNTS`
    # accumulates ACROSS POOLS on purpose -- postBoScoreRanking runs once per pool -- but it was
    # never cleared, so a process that scores twice (the backtest harness, the tuner, a test
    # session, the two arms of the acceptance report) reported the first run's conversions again
    # in the second.  Cleared HERE because postBoWrapper is exactly once per run.
    npol.reset_counts()

    # PEG's SIGN-CROSSING SUBSTITUTION -- a CROSS-SECTIONAL BASELINE, so it belongs exactly here
    # (CEO ruling, 2026-08-05; see calcMetrics.PEG_CROSSING_SUBSTITUTION).
    #
    # A growth rate computed from a NEGATIVE base is not a growth rate, so a row where earnings
    # crossed zero takes the POOL's median growth rate instead and its P/E has to stand on its
    # own -- neither credit nor penalty, and no tuned constant.  The median cannot be computed in
    # `calc_special`: that sees ONE source, and on the fetch path the panel is still being
    # accumulated, so the crossing rows arrive here as NaN.
    #
    # WHY THIS LINE AND NOT THE BUILDERS.  It is the same class of object as `BoMetric_ave` and it
    # gets the same treatment for the same reason (audit H-1): a cross-sectional baseline is
    # recomputed on the frame ACTUALLY SCORED, never carried stale, and never frozen into the
    # saved panel.  `bmdf` is a LOCAL, so the artifact on disk keeps the honest per-source
    # pre-substitution column while the score uses the pooled one.  It is also the ONE seam every
    # Stage-1 path passes through -- `build_bometric_rows` has four call sites and
    # `fixAfterGetData` four, and a fix applied to three of four is this project's signature
    # defect.
    bmdf, _peg_stats = cm.substitute_peg_crossing(bmdf, dmdic.get('cdx_df'),
                                                 freq_map=_freq_map, verbose=True)

    # bmda is passed but UNUSED inside simpleScore_fromDict -- see the note at its assignment.
    BoScore_df = cs.simpleScore_fromDict(bmdf, bmav, bmda, n, as_of=as_of,
                                        freq_map=_freq_map)

    # --- MEAN-BAR FAILSAFE BAND (register C-12, CEO 2026-08-06) ----------------
    # The seven `mean` criteria are scored against STORED CONSTANTS since C-12, which is
    # what removes the sample-dependence and the pooled-median lookahead -- but a constant
    # cannot notice that the world moved.  So the run MEASURES each bar's realised pass
    # rate on the population Stage-1 actually scored and WRITES IT DOWN.
    # IT NEVER CHANGES A BAR, and that is the design, not a limitation: a bar that re-fits
    # itself to hold a pass rate IS the pooled median with a longer time constant.
    # HERE, and not in the emission block below, because this is the one place the FULL
    # post-manual-elimination panel is in memory -- the cohorts and the top-100 are
    # selected samples, and a bar judged on a selected sample is judged on the selection.
    # ONCE per run: postBoWrapper runs exactly once, the same reason npol.reset_counts()
    # sits above.  Best-effort inside meanBars; it cannot abort the run.
    _mean_dict = cdic.getDicts()[3]
    mb.emit_calibration(
        bmdf,
        {'m' + k[0].upper() + k[1:]: _mean_dict[k]['Sign'] for k in _mean_dict},
        universe=(dmdic.get('universe') or dmdic.get('tickerfilter') or 'unknown'),
        #  THE ONLY OPT-IN IN THE REPO (2026-08-06).  This is the production scoring seam, so
        #  this is the one report allowed to advance or seed the breach-streak hysteresis.
        #  Every offline/research caller reaches `emit_calibration` through its default
        #  `streak_participant=False` and therefore cannot chain a streak off a research panel
        #  -- see meanBars._prior_streaks for the `_unknown`-basename hole this closes.
        streak_participant=True)

    # --- Phase-1 cohort carve-out (BEFORE the head(100) selection) -------------
    # Partition the full BoScore-ranked universe into a GENERAL pool + three
    # disjoint side-cohorts (REITs / Mining / investment vehicles) and apply a
    # gentle $25M market-cap floor uniformly.  The MAIN shortlist is drawn ONLY
    # from the general pool; each cohort is ranked with the SAME machinery and
    # presented as a labeled side-list.  This only changes WHICH names feed each
    # ranking -- the ranking (and its as_of/live behaviour) is unchanged.
    #
    # ROBUSTNESS (critical path): the carve-out must NEVER be able to destroy the
    # main deliverable.  The partition is wrapped so a failure falls back to the
    # legacy BoScore_df.head(100) (original, no-carve behaviour); the MAIN ranking
    # + resdic are built FIRST, fully independent of the cohorts; and the
    # side-lists run afterwards in a per-cohort guarded best-effort block, so a
    # small/degenerate cohort (e.g. qcut on ~10 names) can crash at most its own
    # side-list -- never the general ranking that already succeeded.
    carve = None
    #  Did the PRE-veto short-pool warning already fire?  Read by the post-veto re-check
    #  below so a pool that was short before the veto is not warned about twice with the
    #  same number, while a pool the VETO shortened is always reported.
    _short_warned = False
    try:
        import carveOut as co
        #  `_dv_floor` is resolved at the top of this function, where `as_of` is in scope --
        #  see the long note there.  `partition_universe` itself still defaults the floor OFF,
        #  so the three offline carve callers (refit, the depth grid, the tuner) are untouched.
        carve = co.partition_universe(BoScore_df, cdx_df, dmdic.get('Tickers_df'),
                                      mcap_floor=25e6, cohort_head=25,
                                      dollarvol_floor=_dv_floor)
        general_scores = carve['general']
        gp_count = len(general_scores)
        diag = carve['diagnostics']
        #  Do not CLAIM the floor in the headline unless it actually fired. When
        #  reportedCurrency has not flowed, no name has a known USD market cap and the floor
        #  excludes nothing by design (gating, 2026-08-06); saying "after cohorts + $25M
        #  floor" would then be the banner asserting a filter that was never applied.
        _floor_claim = ("$25M floor" if diag.get('floor_enforced', True)
                        else "$25M floor NOT ENFORCED (currency pending)")
        print(f"CARVE-OUT: general pool = {gp_count} names after cohorts + {_floor_claim} "
              f"(REIT={diag['n_REIT']}, Mining={diag['n_Mining']}, "
              f"FIN1_Vehicle={diag['n_InvestmentVehicle']}, "
              f"FIN2_Manager={diag.get('n_FinManager', 0)}, "
              f"FIN3_BalanceSheet={diag.get('n_BalanceSheetFin', 0)}, "
              f"below_floor={diag['n_below_floor']}, unknown_mcap_kept={diag['n_unknown_mcap']})",
              flush=True)
        if gp_count < 100:
            _short_warned = True
            print(f"CARVE-OUT WARNING: general pool has only {gp_count} names (<100); "
                  f"top-100 selection is short -- top-20 may not fully fill.", flush=True)
    except Exception as e:
        # LOUD FALLBACK. The carve-out/dedup IS the deliverable's integrity guarantee
        # (no issuer duplicates; 0 Basic-Materials / financials in the general pool). A
        # SILENT fallback is itself a defect: on 2026-07-13 it shipped legacy un-carved
        # output that looked like a carved deliverable and nobody noticed. Keep the
        # safety net (never crash the pipeline) but make a fallback IMPOSSIBLE to mistake
        # for success: full traceback + an unmistakable banner on BOTH stdout and stderr.
        import traceback
        banner = (
            "\n" + "!" * 78 + "\n"
            "!!! CARVE-OUT/DEDUP DID NOT RUN -- SHIPPING LEGACY UN-CARVED TOP-100 !!!\n"
            "!!! The general pool is NEITHER de-duped NOR sector-carved this run:  !!!\n"
            "!!!   expect issuer/share-class duplicates AND Basic-Materials /       !!!\n"
            "!!!   financials leaking into the general top-100.                     !!!\n"
            f"!!! Cause: {type(e).__name__}: {e}\n"
            "!!! DO NOT treat this output as a carved deliverable.                  !!!\n"
            + "!" * 78 + "\n")
        # stderr first (survives stdout redirection and tqdm progress-bar noise).
        print(banner, file=sys.stderr, flush=True)
        traceback.print_exc(file=sys.stderr)
        print(banner, flush=True)
        traceback.print_exc(file=sys.stdout)
        carve = None
        general_scores = BoScore_df
        gp_count = len(general_scores)

    # --- STAGE-1 RED-FLAG VETO (CEO, 2026-08-05) -- FLAG DEFAULT OFF -----------
    # Ejects a name that persistently fails ANY of five solvency / earnings-reality flags,
    # BEFORE the head(100) and head(25) cuts, on THE GENERAL POOL ONLY (CEO, 2026-08-07 --
    # two of the five flags are structurally undefined on the carve-out cohorts; the measured
    # per-cohort rates and the overruled all-pools argument are in `stage1_veto`).  The
    # mechanism, the five flags, the `<=1 of 8` fail definition and why it is not `psbrfilter`
    # are in `stage1_veto` too.
    # THE COHORT CALLS BELOW ARE STILL MADE, and deliberately: `stage1_veto.VETO_POOLS` -- not
    # this caller -- decides scope, and each cohort call returns its frame UNCHANGED plus a
    # report carrying `applies=False` and the reason.  Dropping the calls would leave the
    # cohorts absent from `veto_reports`, which reads as "the veto did not run at all" rather
    # than "the veto is out of scope here", and would put the scope decision in two places.
    # WITH THE FLAG OFF THIS IS A NO-OP and the pipeline is bit-identical -- `apply_veto` returns
    # its input frame unchanged, so the `try` below cannot alter a shipped run today.
    # GUARDED like the carve-out, and for the same reason: the veto must never be able to destroy
    # the main deliverable.  A failure (e.g. an older panel with no `uInterestCoverage` column)
    # degrades to the UN-VETOED pool with a loud warning, never to a crash.
    #  --- THE AD-HOC PENALTY BUCKET (CEO, 2026-08-10) ---------------------------------
    #  ONE book per run, filled by the veto layer (which is the only thing that currently
    #  raises contributions) and handed to EVERY Stage-2 pool below.  It is created OUTSIDE
    #  the try so that a veto failure leaves an EMPTY book rather than an undefined name --
    #  a run with no veto has no data-gap findings, which is the honest state, and Stage-2
    #  then scores a zero penalty column rather than no column at all.
    import adhoc_penalty as ap
    penalty_book = ap.PenaltyBook()

    #  --- CHARGE A NAME WE CANNOT PRICE (CEO ruling; wired 2026-08-24) -----------------
    #  CEO, verbatim: *"they should be punished for not having this data ... No metric makes
    #  sense if not relative to a price."*  The backtest half of that ruling is already
    #  shipped; this is the live half.
    #
    #  *** WHAT "CANNOT RELIABLY PRICE" WAS SCOPED TO, AND WHY IT IS THIS AND NOT SOMETHING
    #  WIDER.  This charge is deliberately NARROW: it is exactly the set the $1M/day
    #  traded-value floor COULD NOT JUDGE -- `dollarVolume_usd` is NaN because the name has
    #  no profile price, no trading currency, no resolvable FX rate or no volAvg capture at
    #  all.  Three properties make it the right scope:
    #    * IT IS DISJOINT FROM THE FLOOR BY CONSTRUCTION.  The floor acts on a KNOWN value
    #      below $1M/day; this acts on NO value.  No name can be both, so this cannot become
    #      a second liquidity floor and cannot double-charge the first one.
    #    * IT IS NOT A LIQUIDITY TEST.  It says nothing about how much the name trades -- it
    #      says the price INPUT is absent, which is the CEO's own criterion.  The floor keeps
    #      these names (absence is not illiquidity, see carveOut.partition_universe), so
    #      without this charge a name we cannot price at all ranks EXACTLY as if it had
    #      cleared the floor.  That is the hole this closes, and it is the floor's own hole.
    #    * IT DOES NOT OVERLAP THE IMPUTATION LADDER.  That ladder (postBoRank) measures how
    #      much of a name's Stage-2 WEIGHT was filled with a neutral z; this measures whether
    #      the volavgdic profile capture answered.  Different inputs, different mechanisms.
    #      MEASURED on the 2026-08-22 CUR6K top-100: the one name here (OPLN, rank 93,
    #      `dollarVolume_basis = no-reading`) carries `imputed_weight_share = 0.0`, so the
    #      ladder charges it nothing and the two instruments touch disjoint names.
    #
    #  *** THE WIDER READING OF THE RULING WAS MEASURED AND DECLINED.  A charge for "the
    #  price SERIES is unreliable" -- holes in the panel's `price` column -- was scoped and
    #  found to have an EMPTY population: on the 2026-08-22 CUR6K panel, 0 of 4,954 names
    #  have a single missing or non-positive `price` in the newest 8 quarters, which is the
    #  window the price-bearing Stage-2 metrics read.  The holes that exist are all in the
    #  deep 80-quarter tail the score never looks at, and they are what the derived-price
    #  hole-fill already addresses.  Building a second instrument on an empty population
    #  would ship a latent, uncalibrated charge that first fires on some future run -- so it
    #  was not built.  This is the half of the ruling with a real, measured population. ***
    #
    #  ONE POINT PER NAME, which is the bucket's convention for one named data gap (weight
    #  0.01 -> -0.01 AggScore).  Not scaled: there is nothing to scale BY -- the reading is
    #  absent, and absence has no magnitude.
    CHECK_UNPRICEABLE = 'unpriceable:no_traded_value_reading'
    try:
        _dv_unknown = set((carve or {}).get('diagnostics', {}).get('dollarvol_unknown') or [])
        _dv_enforced = bool((carve or {}).get('diagnostics', {})
                            .get('dollarvol_floor_enforced', False))
        if _dv_unknown and not _dv_enforced:
            #  EVERY name is unknown, i.e. the capture is missing wholesale.  Charging the
            #  entire universe one point each would move nothing (a constant subtracted from
            #  every score cannot reorder) and would fill the evidence CSV with thousands of
            #  rows that say "this run had no volavgdic", which is one fact about the RUN and
            #  not a finding about any name.  Declared instead.
            penalty_book.declare_unmeasured(
                CHECK_UNPRICEABLE,
                'NOT CHARGED: this run has NO traded-value reading for ANY name (no volavgdic '
                'profile capture), so "we cannot price this name" is a statement about the run, '
                'not about %d individual names. The $/day floor did not run either.'
                % len(_dv_unknown), pool='general')
        elif _dv_unknown:
            _pools = {'general': list(general_scores['source'])}
            if carve is not None:
                _pools.update({lab: list(cs['source'])
                               for lab, cs in carve['cohorts'].items()})
            _n_charged = 0
            for _lab, _members in _pools.items():
                for _s in _members:
                    if _s in _dv_unknown:
                        penalty_book.add(
                            _s, CHECK_UNPRICEABLE,
                            'no usable traded-value reading: dollarVolume_usd could not be '
                            'formed (no profile price, no trading currency, no FX rate, or no '
                            'volAvg capture), so the $/day floor could not judge this name and '
                            'kept it. It is charged rather than screened.', 1.0, pool=_lab)
                        _n_charged += 1
            print('UNPRICEABLE CHARGE: %d name(s) across %d pool(s) carry no traded-value '
                  'reading and are charged 1 point each (-%.2f AggScore). The $/day floor '
                  'KEPT them -- absence is not illiquidity -- so this is the only thing that '
                  'acts on them.' % (_n_charged, len(_pools), ap.WEIGHT), flush=True)
    except Exception as _ue:
        #  A charge must never be able to cost the run, and a failed charge must not read as
        #  a clean pool.
        print('WARNING: the unpriceable-name charge did not run (%s: %s) -- names with no '
              'traded-value reading are scored with NO penalty this run. NOT a finding that '
              'every name could be priced.' % (type(_ue).__name__, _ue), flush=True)

    veto_reports = {}
    _gp_pre_veto = gp_count
    try:
        import stage1_veto as sv
        #  `cdx_df` is passed for the AD-HOC BUCKET ONLY -- it decides whether a refused row
        #  is a data problem (`stage1_veto.REFUSAL_CORROBORATOR`) and reaches no verdict.
        #  Without it the bucket cannot judge a refusal and SAYS SO rather than charging zero.
        general_scores, _vrep = sv.apply_veto(general_scores, bmdf, pool_label='general',
                                              penalty_book=penalty_book, cdx_df=cdx_df)
        veto_reports['general'] = _vrep
        if carve is not None and _vrep['enabled']:
            #  ONE call per cohort: it returns (kept, report) together, so the report's `n_in`
            #  is the PRE-veto count.  Calling it twice would report the post-veto pool as the
            #  input and every cohort would look like it ejected nobody.
            _vetoed = {lab: sv.apply_veto(cs, bmdf, pool_label=lab,
                                          penalty_book=penalty_book, cdx_df=cdx_df)
                       for lab, cs in carve['cohorts'].items()}
            carve['cohorts'] = {lab: kept for lab, (kept, _r) in _vetoed.items()}
            veto_reports.update({lab: r for lab, (_k, r) in _vetoed.items()})
        gp_count = len(general_scores)
    except Exception as _e:
        print('WARNING: STAGE-1 VETO DID NOT RUN -- pools are UN-VETOED this run (%s: %s)'
              % (type(_e).__name__, _e), file=sys.stderr, flush=True)
        print('WARNING: STAGE-1 VETO DID NOT RUN -- pools are UN-VETOED this run (%s: %s)'
              % (type(_e).__name__, _e), flush=True)
        veto_reports = {}

    #  --- THE FORENSIC DATA GAP FEEDS THE SAME BUCKET (CEO, 2026-08-16) ------------------
    #  *"We should not be rewarding lack of data. What we want is to have the top 20 be
    #  INFORMED good, rather than perhaps good but we don't know because we lack data."*
    #  A name whose Beneish assessment could not be made now carries points, exactly like a
    #  name the Stage-1 veto could not judge.  THE DERIVATION, THE AMOUNT (max 4 points =
    #  -0.04, half the veto's largest) AND THE MEASURED EFFECT are in `detectManipulation`
    #  beside the guards that produce the gap -- not restated here.
    #  RAISED ONCE, WITH `pool=None`, over the union of every pool Stage-2 will score: the gap
    #  is a property of the name's data, not of the pool, and `penalty_series` sums a source's
    #  contributions across pool tags -- so raising it per pool would charge a name that sits in
    #  both the general pool and a cohort twice for one finding.
    #  ITS OWN `try`, separate from the veto's: a forensic-gap failure must not be able to
    #  un-veto the pools, and a veto failure must not silence this.  Either way the bucket
    #  degrades to FEWER points, never to a wrong one.
    try:
        _gap_sources = list(general_scores['source'])
        if carve is not None:
            for _cs in carve['cohorts'].values():
                _gap_sources += list(_cs['source'])
        _gap_df = dm.contribute_forensic_gap_points(cdx_df, _gap_sources, penalty_book)
    except Exception as _ge:
        print('WARNING: FORENSIC DATA-GAP PENALTY DID NOT RUN -- names whose Beneish '
              'assessment is missing are NOT charged this run (%s: %s)'
              % (type(_ge).__name__, _ge), file=sys.stderr, flush=True)
        print('WARNING: FORENSIC DATA-GAP PENALTY DID NOT RUN (%s: %s)'
              % (type(_ge).__name__, _ge), flush=True)
        #  BOTH checks are declared, not just the first (review L-2): a reader filtering the
        #  evidence CSV on `forensic_gap:no_verdict` would otherwise see nothing AND no caveat,
        #  which is the "charged nothing == found nothing" confusion this section exists to
        #  prevent.  The charge itself is all-or-nothing (see `contribute_forensic_gap_points`),
        #  so "cost nothing" is true here rather than approximately true.
        for _chk in (dm.CHECK_GAP_COMPONENTS, dm.CHECK_GAP_NO_VERDICT):
            penalty_book.declare_unmeasured(
                _chk,
                'the forensic data-gap charge could not be computed this run (%s: %s), so an '
                'ABSENT Beneish assessment cost nothing -- "charged nothing" here does not mean '
                '"no gaps found"' % (type(_ge).__name__, _ge))

    #  SHIP THE BUCKET AS EVIDENCE, at the repo ROOT (CEO, 2026-08-10 -- the root-level
    #  artifacts travel, `output/` demonstrably did not on the 2026-08-10 run).  Written
    #  BEFORE Stage-2 rather than after, so the record of what the bucket charged survives
    #  even if the scoring stage later fails.
    _pen_csv = ap.write_evidence_csv(penalty_book)
    print('AD-HOC PENALTY BUCKET: %d contribution(s) across %d name(s); evidence -> %s'
          % (len(penalty_book), len(penalty_book.sources), _pen_csv or '<not written>'),
          flush=True)

    #  SHIP THE EJECTION LIST TOO (register N-5, CEO 2026-08-13), beside the bucket and for
    #  the same reason: the veto is the single biggest edit this pipeline makes to the pool
    #  and it shipped no record of WHICH names it removed.  Written HERE rather than in
    #  `writeResWrapper` so the record survives a later Stage-2 failure -- the same placement
    #  argument as the bucket CSV one line up -- and written even when `veto_reports` is empty
    #  (the file's presence is the evidence the layer was reached).  Guarded twice over: the
    #  writer cannot raise, and the call is a plain statement in a function whose deliverables
    #  must not depend on it.
    _ej_csv = None
    try:
        import stage1_veto as _sv_csv
        _ej_csv = _sv_csv.write_ejection_csv(veto_reports)
        _n_ej = sum(int(r.get('n_ejected') or 0) for r in (veto_reports or {}).values())
        print('STAGE-1 VETO EJECTION LIST: %d name(s) across %d pool(s); evidence -> %s'
              % (_n_ej, len(veto_reports or {}), _ej_csv or '<not written>'), flush=True)
    except Exception as _ee:
        print('WARNING: Stage-1 veto ejection CSV not written (%s: %s); the run is '
              'unaffected, but this run ships NO list of the names the veto removed.'
              % (type(_ee).__name__, _ee), flush=True)

    # --- SHORT-POOL RE-CHECK, AFTER THE VETO (reviewer, 2026-08-05) -------------
    # THE DEFECT THIS CLOSES: the carve-out's `<100` warning above fires on the PRE-veto
    # count and was never re-checked, so with the flag ON the veto could drop the general
    # pool under 100 and the run would ship a short top-100 SILENTLY -- the one thing the
    # original warning exists to prevent.
    # THE CAUSE IS PART OF THE MESSAGE, not decoration: a pool short because we
    # DELIBERATELY ejected red-flag names is a working veto and may be entirely acceptable;
    # a pool short because the universe was thin is a data problem. A bare count cannot
    # tell those apart, and they call for opposite responses.
    _veto_ejected = _gp_pre_veto - gp_count
    if gp_count < 100 and (_veto_ejected > 0 or not _short_warned):
        if _veto_ejected > 0 and _gp_pre_veto >= 100:
            _cause = (f"THE STAGE-1 VETO CAUSED THE SHORTFALL: it ejected {_veto_ejected} "
                      f"of {_gp_pre_veto} names, which were enough before the veto ran")
        elif _veto_ejected > 0:
            _cause = (f"BOTH causes: the pool was ALREADY short at {_gp_pre_veto} before the "
                      f"veto, and the veto then ejected {_veto_ejected} more")
        else:
            _cause = ("the STAGE-1 VETO ejected 0 names -- the shortfall is the universe/carve, "
                      "not the veto")
        print(f"SHORT-POOL WARNING: general pool has only {gp_count} names (<100) going into "
              f"head(100); top-100 selection is short -- top-20 may not fully fill. {_cause}.",
              flush=True)

    # --- MAIN ranking + resdic: built first, independent of the cohorts --------
    BoS_dftop100 = general_scores.head(100)
    BoM_dftop100 = bmdf[bmdf['source'].isin(list(BoS_dftop100.source))].reset_index(drop=True)
    cdx_dftop100 = cdx_df[cdx_df['source'].isin(list(BoS_dftop100.source))].reset_index(drop=True)

    n= 16
    # issuer names for the emission-time issuer-dedup (edge B: name+shares). cdx-based
    # edges (A/C) work without names; passing them widens dual-listing coverage.
    _tdf = dmdic.get('Tickers_df')
    _names = (dict(zip(_tdf['symbol'], _tdf['name']))
              if _tdf is not None and 'symbol' in getattr(_tdf, 'columns', [])
                 and 'name' in getattr(_tdf, 'columns', []) else {})
    rankdic = pbr.postBoScoreRanking(BoM_dftop100, BoS_dftop100, cdx_dftop100, dmdic['baseurl'], dmdic['api_key'],
                                     dmdic['period'],n,as_of=as_of,names=_names,
                                     pool_label='general', penalty_book=penalty_book)

    # UNWINNED FILTER: -1.5 z-score pass-filter on six metrics (earnYield, grahamNumberToPrice,
    # RoA, EPStoEPSmean, freeCashFlowYield, revenueGrowth). Computed and stored in resdic but
    # NOT wired into any shipped deliverable — the shortlist is built from resdic['postRank']
    # only, so psbrfilter currently filters zero names. Left in place per CEO decision
    # (2026-07-14) pending a future decision to either wire it in (would require a soundness
    # review of the -1.5 cutoff on these 6 metrics) or remove it.
    # NOT REMOVED AS DEAD CODE (2026-08-02), and the reason is mechanical rather than a
    # judgement call: `psbrfilter` is STORED IN resdic, so deleting it changes the contents of
    # a saved artifact that other tools read. That makes it a CEO decision, not a cleanup.
    # NOTE the cutoff is applied to postRank's metric columns, which are z x w, NOT z --
    # so -1.5 is -1.5 WEIGHTED units, i.e. a different threshold per metric. That is part of
    # the soundness review this filter is waiting for, not something to "fix" while it is
    # inert (fixing it would change a stored artifact's contents).
    metricList = ['earnYield', 'grahamNumberToPrice', 'RoA', 'EPStoEPSmean', 'freeCashFlowYield', 'revenueGrowth']
    cutoff = 1.5
    psbrfilter = pbr.postBoRankingPassFilter(rankdic['postRank'],metricList,-cutoff,np.inf)

    regressMetricsOnROR(rankdic)

    resdic = {**rankdic, **{'BoS_dftop100': BoS_dftop100, 'BoM_dftop100': BoM_dftop100, 'cdx_dftop100': cdx_dftop100,
                          'BoScore_df': BoScore_df, 'psbrfilter': psbrfilter,  # NOT WIRED — see above comment
                          'general_pool_count': gp_count,
                          # per-pool veto report (see stage1_veto). {} when the flag is off or
                          # the layer did not run -- so a reader can tell "no ejections" from
                          # "the veto was not applied", which a bare count could not.
                          'stage1_veto': veto_reports,
                          #  The AD-HOC PENALTY BUCKET as data (CEO, 2026-08-10), beside the
                          #  dated CSV.  Itemised, so a penalised name in a saved resdic can be
                          #  explained without the CSV and without re-running the veto.
                          'adhoc_penalty': penalty_book.itemised(),
                          'adhoc_penalty_weight': ap.WEIGHT}}

    # --- Side-lists: guarded best-effort, AFTER resdic is complete -------------
    # Only runs if the partition succeeded.  Per-cohort try/except so one
    # degenerate cohort degrades to "no side-list for that cohort" (visible
    # warning) without touching the others or the already-built main output.
    carveout_sidelists = {}
    if carve is not None:
        for label, cohort_scores in carve['cohorts'].items():
            try:
                head = cohort_scores.head(25)
                if head.empty:
                    carveout_sidelists[label] = None
                    continue
                bm = bmdf[bmdf['source'].isin(list(head.source))].reset_index(drop=True)
                cd = cdx_df[cdx_df['source'].isin(list(head.source))].reset_index(drop=True)
                # per-cohort weight vector (general/main pool keeps the default)
                wov = co.COHORT_WEIGHTS.get(label)
                print(f"CARVE-OUT side-list '{label}': ranking {len(head)} names"
                      f"{' with per-cohort weights' if wov else ''}", flush=True)
                # RULE UNM's OTHER HALF, reported at the point of use (E-2, 2026-08-04).
                # A block whose question APPLIES but has no instrument keeps its budget and
                # spends it on nothing; the spendable weights then renormalise to 1, which
                # silently converts "we cannot measure this" into "this does not matter"
                # unless the residue is printed.  So it is printed, per cohort, beside the
                # ranking it qualifies.
                import scoringWeights as sw
                _unpriced = sw.COHORT_UNPRICED_RISK.get(label, 0.0)
                if _unpriced > 1e-9:
                    print(f"  UNPRICED RISK in '{label}': {_unpriced:.2%} of the cohort's "
                          f"weight budget is HELD AND UNSPENT (Rule UNM -- the question "
                          f"applies, no instrument exists). This cohort's score does NOT "
                          f"price that risk; the human review must.", flush=True)
                # pool_label=label so the missing-data fill report names THIS cohort: the
                # cohorts are the part of that calibration nobody has measured.
                carveout_sidelists[label] = pbr.postBoScoreRanking(bm, head, cd, dmdic['baseurl'], dmdic['api_key'],
                                                                   dmdic['period'], n, as_of=as_of,
                                                                   weight_override=wov,
                                                                   pool_label=label,
                                                                   penalty_book=penalty_book)
            except Exception as e:
                print(f"CARVE-OUT side-list '{label}' FAILED ({type(e).__name__}: {e}); "
                      f"skipping this side-list (main output unaffected).", flush=True)
                carveout_sidelists[label] = None
        resdic['carveout_sidelists'] = carveout_sidelists
        resdic['carveout_labels'] = carve['labels']
        resdic['carveout_diagnostics'] = carve['diagnostics']
        # FULL carved + size-floored membership per cohort (source lists, BEFORE the
        # head(25)/head(100) selection).  Data-only: consumed by the READ-ONLY
        # review-reference artifacts (reviewReference) to build cohort distribution
        # stats / percentiles over EVERY peer, not just the Stage-2-scored subset.
        # Never read back into scoring.
        resdic['carve_full_membership'] = {
            'general': list(general_scores['source']),
            **{lab: list(cs['source']) for lab, cs in carve['cohorts'].items()}}
    else:
        resdic['carveout_sidelists'] = {}
        resdic['carve_full_membership'] = None

    return resdic

def regressMetricsOnROR(rankdic):
    # 'rankOfRanks' -> 'rankOfRanks_diag' (postBoRank.ROR_COLUMN): the emitted column was
    # renamed to mark it a DIAGNOSTIC, not a competing ranking (audit M1).
    ror = pbr.ROR_COLUMN
    # DETERMINISTIC column order (2026-08-02).  `list(set(...))` here made the OLS design
    # matrix's column order hash-seed dependent, so the printed R^2 and coefficients moved
    # in their last digits between runs of the same data -- and the reader has no way to tell
    # that from a real change.  PRINT-ONLY (this function returns None and touches no
    # artifact), so unlike getAggScore nothing downstream depends on it; it is fixed because
    # it is the same pattern and the fix is free.  The `zip(regressors, coef)` pairing below
    # was always internally consistent -- the order was arbitrary, not mismatched.
    regressors = pbr.deterministic_column_order(
        rankdic['postRank'].columns, exclude=(ror, 'rankOfRanks', 'AggScore', 'source'))
    regressant = [ror]
    df = rankdic['postRank']
    X = df[regressors]
    y = df[regressant]
    
    # Guard: skip regression if insufficient samples (need at least 1 sample to fit)
    if X.shape[0] < 1:
        print("Warning: Insufficient samples for regression (found 0 samples). Skipping regression.")
        return None
    
    model = LinearRegression()
    model.fit(X, y)
    r_squared = model.score(X, y)
    print("R-Squared:", r_squared)
    coef = model.coef_
    intercept = model.intercept_
    print("Coefficients:", coef)
    print("regressors:", regressors)
    print("Intercept:", intercept)
    print("coefreg:", tuple(zip(regressors, coef.tolist()[0])))
    return None

def _fx_provenance(resdic=None):
    """The `fx_rates` block of the run-provenance sidecar.  NEVER raises, ALWAYS returns.

    Mirrors `_veto_provenance` below in shape and in purpose: it records WHICH FX BASIS
    produced this run's USD market caps, so a run whose FX feed was healthy and a run whose
    FX feed was dead cannot ship indistinguishable artifacts.  `state` is the important
    field -- 'live' means dated rates from v3/quotes/forex, 'failed' means every currency
    was unknown and the $25M floor did not run, and 'unset' means no feed was ever
    attempted and the UNDATED FX_TO_USD constants were used (offline tooling; production
    should never emit this)."""
    try:
        import carveOut as _co
        meta = _co.live_fx_meta() or {}
        out = {'state': _co.fx_source_state()}
        out.update(meta)
        if out['state'] == 'unset':
            out['warning'] = ('no live FX feed was installed: USD market caps were derived '
                              'from the UNDATED carveOut.FX_TO_USD constants')
        elif out['state'] == 'failed':
            out['warning'] = ('the FX feed produced no usable rate: NO reportedCurrency '
                              'resolved, so the $25M universe floor did NOT run and the '
                              'market-cap bands were skipped')
        #  PANEL COVERAGE IN THE SHIPPED ARTIFACT (F-2, reviewer 2026-08-08).  The
        #  supported-set coverage above describes the FEED; this describes what the feed
        #  actually reached on THIS universe, which is the number that qualifies any
        #  `floor_enforced` label. `resdic` is the right source here (unlike the FX stamp
        #  itself) because it is the panel this run scored. Best-effort: a coverage read
        #  must never cost the sidecar.
        try:
            _cdx = (resdic or {}).get('cdx_df')
            if _cdx is not None and len(_cdx):
                _n, _tot, _frac = _co.currency_coverage(_cdx)
                out['panel_coverage'] = round(_frac, 4)
                out['panel_sources_with_usd_mcap'] = _n
                out['panel_sources'] = _tot
                import fx_rates as _fxr
                out['panel_coverage_ok'] = bool(_frac >= _fxr.FX_MIN_PANEL_COVERAGE)
                if not out['panel_coverage_ok']:
                    out['coverage_warning'] = (
                        'the $25M floor reached only %.1f%% of this universe (%d of %d '
                        'names have a resolvable USD market cap). The rest were KEPT and '
                        'UNBANDED -- nothing was wrongly deleted, but this run is NOT '
                        'floor-filtered end to end and must not be labelled as if it '
                        'were.' % (100.0 * _frac, _n, _tot))
        except Exception as _ce:
            out['panel_coverage'] = None
            out['panel_coverage_error'] = '%s: %s' % (type(_ce).__name__, _ce)
        return out
    except Exception as _e:
        return {'state': 'unknown', 'error': '%s: %s' % (type(_e).__name__, _e)}


def _veto_provenance(resdic):
    """The `stage1_veto` block of the run-provenance sidecar.  NEVER raises, ALWAYS returns.

    THE DEFECT THIS CLOSES (CEO, 2026-08-07).  The veto is now ON by default for the general
    pool, and it ejects 58.4% of that pool.  Nothing in the artifacts said whether it ran: two
    runs -- one vetoed, one not -- produced deliverables that were INDISTINGUISHABLE on the one
    axis that changed the pool, so a reader had no way to know which regime produced a top-100.
    Turning the flag on is exactly what makes that unacceptable: while the default was OFF, the
    flip itself was the visible event; with the default ON, the ARTIFACT has to carry it.

    READ FROM THE RUN'S OWN REPORTS (`resdic['stage1_veto']`, one per pool, produced by
    `apply_veto` while the pool was being gated), NOT from `stage1_veto.ENABLED` at write time.
    Reading the module constant here would stamp what the code says NOW rather than what THIS
    RUN did -- and on a `-loadboresults` run, where the scoring happened in a different process
    (possibly with a different constant), that stamp would be a fabrication.  The parameters
    (`WINDOW_ROWS` etc.) ARE read from the module, because they are not carried in the report;
    they are labelled as such and are only meaningful alongside `status`.

    THREE STATUSES, DELIBERATELY DISTINCT -- the whole point is that these must not collapse:
      `applied` / `off`  the layer ran and reported; `by_pool` says what it did to each pool.
      `did_not_run`      the guarded block in `postBoWrapper` caught an exception, so the pools
                         are UN-VETOED. `n_ejected = 0` everywhere would say the same thing as a
                         clean pool, so it is spelled out instead.
      `unknown`          no report on `resdic` at all -- an older pickle re-emitted through
                         `-loadboresults`, i.e. the veto state of that run is NOT establishable.
                         An honest `unknown` beats a plausible default (see the universe stamp's
                         `unknown-not-stamped`, same rule).
    """
    out = {}
    try:
        import stage1_veto as sv
        out['params'] = {'window_rows': sv.WINDOW_ROWS,
                         'fail_max_passes': sv.FAIL_MAX_PASSES,
                         'eject_min_flags': sv.EJECT_MIN_FLAGS,
                         'flags': sorted(sv.FLAGS),
                         'params_read_from': 'stage1_veto module at write time, NOT from the run'}
    except Exception as _e:
        out['params'] = {'error': f'{type(_e).__name__}: {_e}'}

    reports = resdic.get('stage1_veto')
    if reports is None:
        out.update(status='unknown', enabled=None, pools=None, by_pool=None,
                   note=('this run carried NO veto report (pre-veto code, or a resdic loaded '
                         'from an older pickle): whether a Stage-1 veto shaped these '
                         'deliverables is NOT establishable from this run'))
        return out
    if not reports:
        out.update(status='did_not_run', enabled=None, pools=None, by_pool={},
                   note=('the guarded veto block RAISED: no pool was gated and every pool is '
                         'UN-VETOED this run. See the WARNING in logs/. This is NOT the same '
                         'as a veto that ran and ejected nobody'))
        return out
    try:
        _gen = reports.get('general') or next(iter(reports.values()))
        enabled = bool(_gen.get('enabled'))
        #  Per pool: the four counts and the by-flag ejection breakdown the report already
        #  computed, so this costs nothing to carry.  `applies` is kept per pool because
        #  "out of scope here" and "gated and found clean" are different facts and both
        #  show up as `n_ejected = 0`.  The per-source `ejected` / `short_window` lists are
        #  NOT copied -- they live in the postRank pickle (also transferred), and the
        #  sidecar is meant to stay readable.
        by_pool = {}
        for lab, r in sorted(reports.items()):
            by_pool[lab] = {'applies': bool(r.get('applies')),
                            'n_in': r.get('n_in'), 'n_ejected': r.get('n_ejected'),
                            'n_out': r.get('n_out'),
                            'ejected_by_flag': r.get('by_flag') or {},
                            'abstained_by_flag': r.get('n_short_window') or {}}
            if not r.get('applies'):
                by_pool[lab]['not_applicable_reason'] = r.get('not_applicable_reason')
        out.update(status='applied' if enabled else 'off', enabled=enabled,
                   pools=sorted(l for l, r in reports.items() if r.get('applies')),
                   pools_reported=sorted(reports), by_pool=by_pool)
        if not enabled:
            out['note'] = ('the veto was OFF this run: every pool was returned UNCHANGED and '
                           'these deliverables are un-vetoed')
    except Exception as _e:
        #  The key is still emitted, carrying the failure -- dropping it would read as
        #  "no veto", which is the one thing this block exists to prevent.
        out.update(status='unknown', enabled=None, pools=None, by_pool=None,
                   note=f'veto report present but unreadable ({type(_e).__name__}: {_e})')
    return out


def writeResWrapper(resdic):
    ntopagg = resdic['ntopagg']
    ntopxlsx = resdic['ntopxlsx']
    fidag = datetime.today().strftime('%Y-%m-%d')
    fb_df = resdic['postRank']

    # --- market-cap band segmentation (ADDITIVE size axis over the general pool) ---
    # GROUP the existing general ranking (postRank) by USD market cap: General (>$300M)
    # -> top-20; Mid ($150-300M)/Small ($50-150M)/Micro (<$50M) -> top-5 each. No re-score,
    # no re-rank -- the ordering is only PARTITIONED. Best-effort + fully guarded so it can
    # NEVER touch the main deliverables. Degrades gracefully when reportedCurrency has not
    # yet flowed (currency_pending): sub-band CSVs are then SKIPPED rather than emitting
    # misbanded output (CEO 2026-07-17: nothing wrong ships now).
    marketcap_bands = None
    try:
        import carveOut as co
        _tdf_b = resdic.get('Tickers_df')
        _cols_b = getattr(_tdf_b, 'columns', [])
        _names_b = (dict(zip(_tdf_b['symbol'], _tdf_b['name']))
                    if _tdf_b is not None and 'symbol' in _cols_b and 'name' in _cols_b else {})
        marketcap_bands = co.partition_by_marketcap(fb_df, resdic.get('cdx_df'), _names_b)
        resdic['marketcap_bands'] = marketcap_bands['bands']
        resdic['marketcap_band_counts'] = marketcap_bands['band_counts']
        resdic['marketcap_currency_pending'] = marketcap_bands['currency_pending']
        resdic['marketcap_band_selective'] = marketcap_bands.get('band_selective', {})
        resdic['marketcap_band_note'] = marketcap_bands.get('band_note', {})
        _pend = marketcap_bands['currency_pending']
        print("MARKET-CAP BANDS: " + ("CURRENCY PENDING (sub-bands suppressed this run) -- " if _pend else "")
              + ", ".join(f"{lab}={marketcap_bands['band_counts'].get(lab, 0)}"
                          for lab, *_ in co.MCAP_BANDS)
              + f" (unknown_mcap->General={marketcap_bands['unknown_mcap']})", flush=True)
        # MIN-N: name the bands whose "top-N" selects nothing, so a zero-selectivity list
        # is never read as a shortlist (audit M5).
        _nonsel = [lab for lab, sel in (marketcap_bands.get('band_selective') or {}).items()
                   if not sel and marketcap_bands['band_counts'].get(lab, 0) > 0]
        if _nonsel:
            print("MARKET-CAP BANDS -- NOT A SELECTION (band size <= requested top-N, "
                  "every member is listed): "
                  + "; ".join(marketcap_bands['band_note'][lab] for lab in _nonsel),
                  flush=True)
    except Exception as _be:
        print(f"WARNING: market-cap banding skipped ({type(_be).__name__}: {_be}); "
              f"main deliverables unaffected.", flush=True)
        marketcap_bands = None

    mscore = resdic['SLmeanMscore']
    cscore = resdic['SLmeanCscore']
    baseurl = resdic['baseurl']
    api_key = resdic['api_key']
    tickerfilter = resdic['tickerfilter']
    datasource = resdic['datasource']
    years = 6

    # Build the per-name forensic-flag table ONCE (offline, no API calls) and route
    # it into every top-N output. Promotes the already-computed M/C forensic signals
    # + Sloan accruals + financial indicator from decoration to decision-support.
    # Covers ntopagg so both the CSV (ntopagg) and the presentation (ntopxlsx) can
    # index into it. FLAGS, NOT VERDICTS: nothing here drops a name.
    flag_df = ff.buildForensicFlagTable(resdic, ntopagg)
    fname_forensic = f'ForensicFlagsTop{ntopagg}-{fidag}_{datasource}_{tickerfilter}.csv'

    # create csv listing the ntopagg stocks. writeBoAggToCSV fetches the API
    # `sector` per name and cross-checks it against the pickle-derived financial
    # classification (ff.applySectorFallback), returning the reconciled flag_df.
    fname_AggScoretop = f'AggScoreTop{ntopagg}-{fidag}_{datasource}_{tickerfilter}.csv'
    # GUARDED (fix, 2026-07-31), matching the convention every other block in this function
    # already follows.  writeBoAggToCSV makes ~4-5 API calls x ntopagg names, and it ran
    # UNGUARDED here: anything it raised took out not only the AggScore CSV but every
    # deliverable AFTER it -- the forensic CSV, the presentation XLSX, the side-lists, the band
    # CSVs and the pick-log -- i.e. the entire output of a 12-hour fetch, for a transient
    # post-processing error. The calls themselves are now hardened (gdg.safe_json_list), so
    # this is the second line of defence rather than the first.
    # flag_df is DELIBERATELY left at its pre-call value on failure: writeBoAggToCSV returns
    # the sector-reconciled table, so on failure we publish the unreconciled one (which is
    # what the forensic CSV was built from) rather than None, which would cascade.
    try:
        flag_df = writeBoAggToCSV(fb_df, mscore, cscore, baseurl, api_key, ntopagg,
                                  fname_AggScoretop, flag_df,
                                  raw_df=resdic.get('postScoreMetric_raw'),
                                  cdx_df=resdic.get('cdx_dftop100'),
                                  missing_fill_df=resdic.get('missing_fill_by_name'),
                                  run_date=fidag,
                                  universe_stamp={
                                      'universe': resdic.get('universe'),
                                      'universe_fingerprint':
                                          resdic.get('universe_fingerprint')})
    except Exception as _e:
        print(f'WARNING: AggScore CSV stage failed ({type(_e).__name__}: {_e}); '
              f'{fname_AggScoretop} may be missing or partial and the API sector '
              f'cross-check did not run. Every LATER deliverable (forensic CSV, XLSX, '
              f'side-lists, band CSVs, pick-log) still runs.', flush=True)

    # Write the standalone forensic decision-support CSV AFTER the API-sector
    # cross-check, so it carries the reconciled (conservative) financial classification.
    ff.writeForensicFlagsCSV(flag_df, fname_forensic)
    print(f'Forensic-flag table written to: {fname_forensic}')

    # create presentation xlsx of ntopxlsx stocks. When currency data is present the
    # general top-N is drawn from the General band (>$300M); pending currency -> unchanged.
    fname_presentationtop= f'PresentationTop{ntopxlsx}-{fidag}_{datasource}_{tickerfilter}.xlsx'
    # GUARDED (fix, 2026-07-31), same reasoning as the AggScore CSV above: ~7 API calls x
    # ntopxlsx names, previously unguarded, and everything after it (side-lists, band CSVs,
    # reviewReference, pick-log) died with it.
    try:
        createPresentation(fb_df, mscore, cscore, baseurl, api_key, ntopxlsx, fname_presentationtop, years, flag_df,
                           bands=marketcap_bands)
    except Exception as _e:
        print(f'WARNING: presentation XLSX stage failed ({type(_e).__name__}: {_e}); '
              f'{fname_presentationtop} may be missing or partial. Every LATER deliverable '
              f'(side-lists, band CSVs, reviewReference, pick-log) still runs.', flush=True)

    # Phase-1 carve-out: write each labeled side-list (REIT / Mining / investment
    # vehicles) as its own compact CSV alongside the main deliverables. Best-effort
    # and self-contained: never raises, and is a no-op when no side-lists are
    # present (e.g. an older resdic), so the main path is unchanged.
    sidelist_fnames = []
    try:
        sidelists = resdic.get('carveout_sidelists') or {}
        for label, sdic in sidelists.items():
            if not sdic or 'postRank' not in sdic:
                continue
            sl_df = sdic['postRank'].head(ntopagg).copy()
            keep = [c for c in ['source', 'AggScore', pbr.ROR_COLUMN] if c in sl_df.columns]
            fname_sidelist = f'SideList_{label}_Top{ntopagg}-{fidag}_{datasource}_{tickerfilter}.csv'
            sl_df[keep].to_csv(fname_sidelist, index=False)
            sidelist_fnames.append(fname_sidelist)
            print(f'Carve-out side-list written to: {fname_sidelist}')
    except Exception as _e:
        print(f'WARNING: carve-out side-list writing skipped ({_e})')

    # Market-cap band CSVs -- one compact CSV per band, MIRRORING the SideList block
    # above. General (>$300M) top-20 + Mid/Small/Micro top-5. Best-effort + self-
    # contained (never raises; no-op when banding was skipped or absent). SKIPPED
    # entirely when currency data is pending, so no misbanded file ever ships (the
    # General-band CSV would just duplicate the top-20 and the sub-bands would be
    # misbanded) -- the feature becomes correct automatically once reportedCurrency flows.
    band_fnames = []
    try:
        if marketcap_bands and not marketcap_bands.get('currency_pending', True):
            _sel_map = marketcap_bands.get('band_selective') or {}
            _note_map = marketcap_bands.get('band_note') or {}
            for label, band_df in (marketcap_bands.get('bands') or {}).items():
                if band_df is None or band_df.empty:
                    continue
                keep = [c for c in ['source', 'AggScore', pbr.ROR_COLUMN] if c in band_df.columns]
                out_band = band_df[keep].copy()
                # MIN-N LABEL (audit M5): carry the selectivity statement INTO the file --
                # a "top-5" over a 4-member band selected nothing, and a CSV read out of
                # context gives the reader no way to know that. Also marked in the filename
                # so it is visible before the file is even opened.
                _note = _note_map.get(label, '')
                out_band['band_selection'] = _note
                # AVERAGE VOLUME -- REPORTED, NEVER SCREENED ON (register J-1, CEO
                # 2026-08-06).  Same two columns as the AggScore CSV, for the same reason:
                # these band CSVs ARE the per-band shortlist a human reads.  Appended to an
                # already-selected, already-ordered frame, so it cannot affect membership or
                # order; absence reads as absence (see carveOut.volavg_report_frame).
                try:
                    _bvol = co.volavg_report_frame(list(out_band['source']))
                    out_band['volAvg_report'] = _bvol['volAvg_report'].values
                    out_band['volAvg_asof'] = _bvol['volAvg_asof'].values
                except Exception as _bve:
                    print(f'WARNING: volAvg report columns not added to the {label} band CSV '
                          f'({type(_bve).__name__}: {_bve}); the band selection itself is '
                          f'unaffected (the columns are REPORT-ONLY).', flush=True)
                _selective = _sel_map.get(label, True)
                _marker = '' if _selective else '_ALLMEMBERS_NOT_A_SELECTION'
                fname_band = (f'MarketCapBand_{label}{_marker}-{fidag}_'
                              f'{datasource}_{tickerfilter}.csv')
                out_band.to_csv(fname_band, index=False)
                band_fnames.append(fname_band)
                print(f'Market-cap band CSV written to: {fname_band}'
                      + (f'   [{_note}]' if _note else ''))
        elif marketcap_bands and marketcap_bands.get('currency_pending', True):
            print('Market-cap band CSVs SKIPPED this run: reportedCurrency not yet in the '
                  'data (correct automatically from the next full fetch).', flush=True)
    except Exception as _e:
        print(f'WARNING: market-cap band CSV writing skipped ({_e})')

    # READ-ONLY review-reference DATA artifacts (RawMetricsTop100 + CohortMetricStats).
    # Computed AFTER scoring from the RAW metrics captured before normalizeAndDropNA
    # (postBoRank -> rankdic['postScoreMetric_raw']); NEVER read back into scoring or the
    # ranking (see reviewReference module docstring -- feeding cohort stats into the score
    # would be cross-sectional sector-neutralization, CEO-ratified OFF). Gated internally
    # on carve labels present (pre-carve run -> skipped with a note) and fully guarded so
    # it can never crash the deliverables. Filenames travel with the other deliverables.
    reviewref_fnames = []
    try:
        import reviewReference as rr
        reviewref_fnames = rr.emit_live(resdic, fidag, datasource, tickerfilter)
    except Exception as _e:
        print(f'WARNING: review-reference artifacts skipped ({type(_e).__name__}: {_e})')

    # ------------------------------------------------------------------------------ #
    #  UNIVERSE-PROVENANCE SIDECAR (2026-08-03).                                       #
    #                                                                                #
    #  The universe stamp reached the three PICKLES but none of the human-readable      #
    #  deliverables -- and those are what actually get opened, mailed and compared.     #
    #  Every one of them is named `<kind>-<date>_<ds>_<tickerfilter>`, i.e. by a         #
    #  universe NAME whose meaning changed on 2026-08-02, so two files could carry       #
    #  identical names while describing different pools.                                #
    #                                                                                #
    #  A SIDECAR rather than in-file stamping, deliberately:                            #
    #    * the XLSX is produced by a 160 KB generator with its own sheet layout, and     #
    #      threading a stamp through it risks the deck for no extra information;         #
    #    * the CSVs have consumers (baseline_tools, the backtest) that select columns    #
    #      by name, and a sidecar cannot break any of them;                             #
    #    * it covers EVERY deliverable this run wrote, including ones added later,       #
    #      because it lists them by filename.                                           #
    #  Guarded: a sidecar failure must never cost a deliverable.                         #
    # ------------------------------------------------------------------------------ #
    deliverables = ([fname_AggScoretop, fname_presentationtop, fname_forensic]
                    + sidelist_fnames + band_fnames + reviewref_fnames)
    try:
        import json as _json
        _prov = {k: resdic.get(k) for k in (
            'universe', 'universe_label', 'universe_fingerprint', 'universe_exchanges',
            'universe_every_exchange', 'universe_expected_count',
            'universe_definition_changed', 'universe_previous_exchanges',
            'universe_codes_verified', 'universe_note', 'universe_resolved_members')}
        # The explicit member list is recorded for a curated universe -- it IS the
        # definition there, and 144 symbols is cheap.
        if resdic.get('universe_symbols'):
            _prov['universe_symbols'] = list(resdic['universe_symbols'])
        _prov.update({'run_date': fidag, 'datasource': datasource,
                      'tickerfilter_flag': tickerfilter,
                      #  WHICH VETO REGIME PRODUCED THESE FILES (CEO, 2026-08-07).  The
                      #  universe keys above say WHICH NAMES WENT IN; this says which of them
                      #  a gate then removed -- 58.4% of the general pool with the veto on.
                      #  Without it, a vetoed and an un-vetoed run are indistinguishable on
                      #  the one axis that changed the pool.  Never raises (see
                      #  `_veto_provenance`), so it cannot cost the universe stamp.
                      'stage1_veto': _veto_provenance(resdic),
                      #  WHICH FX PRODUCED THESE USD NUMBERS (2026-08-08).  A stale-FX run
                      #  and a clean-FX run must NEVER produce identical artifacts -- that
                      #  is the 2026-08-07 post-mortem rule (the allowlist note in
                      #  Sbocker.transfer_outputs_to_drive), and FX is
                      #  now a live input that can differ run to run or fail outright.
                      #  `fx_rates_as_of` + the dated FxRates_*.csv are what make
                      #  the two distinguishable on the receiving machine.
                      #
                      #  READ FROM MODULE STATE, NEVER FROM resdic: on the -loadbometric /
                      #  -loadboresults paths resdic is rebuilt from a SAVED pickle, which
                      #  can carry a PREVIOUS run's FX stamp. carveOut's state always
                      #  describes the process that is writing this file.
                      'fx_rates': _fx_provenance(resdic),
                      'deliverables': list(deliverables)})
        if not _prov.get('universe_fingerprint'):
            _prov['universe_fingerprint'] = 'unknown-not-stamped'
            _prov['warning'] = ('this run carried no universe stamp; its membership is '
                                'NOT establishable from the artifacts')
        fname_prov = f'RunProvenance-{fidag}_{datasource}_{tickerfilter}.json'
        with open(fname_prov, 'w') as _f:
            _json.dump(_prov, _f, indent=1, default=str)
        deliverables = deliverables + [fname_prov]
        _vp = _prov.get('stage1_veto') or {}
        print(f'Universe provenance sidecar written to: {fname_prov} '
              f"(universe={_prov.get('universe')} "
              f"fingerprint={_prov.get('universe_fingerprint')} "
              f"stage1_veto={_vp.get('status')} pools={_vp.get('pools')})", flush=True)
    except Exception as _e:
        print(f'WARNING: universe-provenance sidecar skipped '
              f'({type(_e).__name__}: {_e}); deliverables unaffected.', flush=True)

    # ------------------------------------------------------------------------------ #
    #  INDUSTRY COUNTER -- top-100 and top-20 (CEO, 2026-08-04).                       #
    #                                                                                #
    #  The 07-17 corrected top-100 holds 11 Marine Shipping (7 of the top-20) and NO   #
    #  deliverable said so.  The CEO reviews the shortlist by hand, so concentration is #
    #  something he has to SEE -- it is NOT something the filter may act on (standing   #
    #  ruling: no hard gates in the filtering logic).                                   #
    #                                                                                #
    #  Printed LAST, after every deliverable, so it is the composition summary of the   #
    #  files just listed and so nothing downstream of it exists to be influenced.       #
    #  Counts THE SAME twenty names the deck presents (generalTopN, shared with         #
    #  createPresentation) rather than a second guess at the general top-N.             #
    #  Read-only + guarded: no resdic key is written, no frame is mutated, and a failure #
    #  here costs a text block, never a deliverable.                                    #
    # ------------------------------------------------------------------------------ #
    try:
        import industry_concentration as ic
        _cdx_ind = resdic.get('cdx_df')
        _uni = (sorted(set(_cdx_ind['source']))
                if _cdx_ind is not None and 'source' in getattr(_cdx_ind, 'columns', [])
                else None)
        _top100 = list(fb_df['source'].head(ntopagg))
        _top20 = list(generalTopN(fb_df, marketcap_bands, ntopxlsx,
                                  warn=False)['source'].head(ntopxlsx))
        print("\n" + "\n".join(ic.report_lines(
            _top100, _top20, universe_sources=_uni,
            labels=(f'top-{ntopagg}', f'top-{ntopxlsx}'))), flush=True)
    except Exception as _e:
        print(f'WARNING: industry counter skipped ({type(_e).__name__}: {_e}); '
              f'deliverables unaffected.', flush=True)

    # Return the human-readable top-N deliverables just written (same pattern as
    # utils.saveWrapper returning its pickle name) so Sbocker.main can copy them to
    # the Drive-synced transfer dir at the pre-ingestion phase boundary. Data-only:
    # nothing here changes scoring/ranking/forensic output.
    return deliverables

def _current_ratio_panel_table(cdx_df):
    """{source: currentRatio} for the published `currentRatio` column, COMPUTED from the
    run's OWN panel.  Newest row per source, taken by DATE (same reason as
    `_pe_panel_table`).  Returns {} on any failure -- this feeds a REPORT column, so it
    degrades to the vendor fallback rather than costing the CSV.

    *** THE DEFECT THAT FORCED THIS (2026-08-24). ***  The column was FMP's `currentRatio`
    read straight off a SEPARATE live `v3/ratios/<symb>?period=quarter&limit=4` call and
    printed.  The consumer TYPE-CHECKED the value and never VALUE-CHECKED it, so a vendor
    `0` formatted as "0.0000" while a missing key correctly produced the 'NaN' sentinel.
    On the shipped 2026-08-22 CUR6K top-100 that put 0.0000 against `092730.KQ` at **RANK
    7** and `041830.KQ` at rank 27 -- and `generate_presentation` flags `currentRatio < 1.0`
    as a solvency red flag, so a rank-7 pick shipped a red flag it does not have, in the CSV
    the CEO's manual review reads.  Both names are solvent by a wide margin: the panel's own
    balance sheet gives 184,604,415,000 / 32,089,653,000 KRW = **5.75** and
    230,256,128,000 / 43,462,233,000 = **5.30**.

    IT IS NOT A COLUMN-WIDE ERROR, AND THAT WAS CHECKED BEFORE CHANGING ANYTHING.  Across
    the same 100 names the vendor value and this computed one agree to a relative difference
    below 4e-5 on 97 of them; the only other deviation is `067160.KQ` at 7.8%, a
    reporting-vintage difference (the live call's newest quarter is later than the panel's),
    not a defect.  The two zeros are the whole population, both `.KQ`, 2 of 18 Korean names
    and 0 of 82 others.

    WHY COMPUTE RATHER THAN VALUE-CHECK THE CELL.  The house rule: the vendor supplies RAW
    INPUTS, we compute the derived quantity -- the same rule that moved `PE-ratio`,
    `dividendYield` and `GrahamNumberToPrice` off the vendor's own fields.  The
    reconstruction is free: `totalCurrentAssets` and `totalCurrentLiabilities` are already
    in the panel, and the ratio IS their quotient (verified: the panel's own `currentRatio`
    column equals the quotient to floating point on both zero cases).  A value-check alone
    would still leave the published number coming from a second, later, unreconciled live
    call -- a whole class of divergence, not one bad cell.

    *** THIS DOES NOT REMOVE AN API CALL, AND AN EARLIER FRAMING SAID IT WOULD. ***  The
    same `v3/ratios` response still supplies `grossProfitMargin` and the
    `priceEarningsRatio` vendor fallback further down the loop, so the call stays.  What is
    removed is the DIVERGENCE, not the request.

    SCORING IS UNAFFECTED AND THIS IS MEASURED, NOT ASSUMED.  `uCurrentRatio` is
    `FIELD_EVIDENCE='counts'`, so a 0 behaves in Stage-1 exactly as a NaN does, and of the
    407 general-pool `uCurrentRatio` ejections **0 of 407 flip** when the impossible zeros
    are repaired.  This is a DISPLAY change.
    """
    try:
        if cdx_df is None:
            return {}
        cols = getattr(cdx_df, 'columns', [])
        if 'totalCurrentAssets' not in cols or 'totalCurrentLiabilities' not in cols:
            return {}
        df = cdx_df.copy()
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df = df.sort_values(['source', 'date'], ascending=[True, False])
        newest = df.groupby('source', sort=False).head(1)
        out = {}
        for _, r in newest.iterrows():
            ca = pd.to_numeric(r.get('totalCurrentAssets'), errors='coerce')
            cl = pd.to_numeric(r.get('totalCurrentLiabilities'), errors='coerce')
            ratio = (ca / cl) if (pd.notna(ca) and pd.notna(cl) and cl != 0) else None
            out[r['source']] = _current_ratio_value(ratio)
        return out
    except Exception:
        return {}


def _current_ratio_value(v):
    """`v` as a float if it is a POSSIBLE current ratio, else None.

    THE ONE RULE, and it is the rule the old type-check was missing: a current ratio of
    exactly ZERO is not a solvency reading, it is a vendor null wearing a number.  It would
    mean a company with current assets of exactly nothing while carrying current
    liabilities -- and the two names it fired on hold 184.6bn and 230.3bn KRW of current
    assets.  A NEGATIVE ratio is impossible on the same argument (neither side of the
    quotient can be negative), and a non-finite one is arithmetic debris.  All three become
    None, which the caller renders as the 'NaN' sentinel a MISSING key already produced --
    so "the vendor said nothing" and "the vendor said something impossible" finally read the
    same, which is the whole defect.

    WHAT THIS CANNOT DETECT: a WRONG-BUT-POSSIBLE ratio.  A vendor 1.0 on a company whose
    balance sheet says 5.75 passes this guard untouched and would still ship a red flag one
    side of the deck's 1.0 line.  Nothing here reconciles the two sources -- reading the
    panel is what does that, and this function is only the backstop for the path where the
    panel cannot answer.
    """
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(f) or f <= 0:
        return None
    return f


def _pe_panel_table(cdx_df):
    """{source: (newest earningsYield, rows-per-year)} from the run's OWN panel.

    Newest row per source, taken by DATE rather than by arrival order, because nothing on
    this path guarantees the ingestion order (the same reason Stage-1 and Stage-2 both
    re-sort at their own boundary).  Returns {} on any failure -- this feeds a REPORT column,
    so it degrades to the vendor fallback rather than costing the CSV.
    """
    try:
        if cdx_df is None or 'earningsYield' not in getattr(cdx_df, 'columns', []):
            return {}
        import reporting_period as _rp
        df = cdx_df.copy()
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df = df.sort_values(['source', 'date'], ascending=[True, False])
        newest = df.groupby('source', sort=False).head(1)
        try:
            freq = _rp.frequency_by_source(df, verbose=False)
        except Exception:
            freq = None
        out = {}
        for _, r in newest.iterrows():
            s = r['source']
            try:
                rpy = _rp.rows_per_year(freq, s) if freq is not None else _rp.DEFAULT_ROWS_PER_YEAR
            except Exception:
                rpy = _rp.DEFAULT_ROWS_PER_YEAR
            out[s] = (pd.to_numeric(r.get('earningsYield'), errors='coerce'), float(rpy))
        return out
    except Exception:
        return {}


def _pe_ratio_from_panel(table, symb):
    """The published `PE-ratio` for `symb`, COMPUTED, or None if the panel cannot answer.

    THE DEFECT THAT FORCED THIS (2026-08-10).  The column was FMP's `priceEarningsRatio`
    read straight off `v3/ratios/<symb>` and printed.  On the 2026-08-10 shipped top-100
    `086280.KS` displayed **66.28 at RANK 4** while the panel's own newest row gives
    `earningsYield = 0.021864` per quarter -> an annualised P/E of **11.43**, and
    `price / epsTTM = 207,500 / 22,306 = 9.30`.  Two independent readings off our own data
    agree to within a third; the vendor's is 5.8x either of them.

    IT IS NOT A COLUMN-WIDE ERROR, AND THAT WAS CHECKED BEFORE CHANGING ANYTHING: across the
    same 100 names the displayed value tracks the panel's earnings yield with a MEDIAN RATIO
    OF 1.000.  So the column is right about the population and wrong about individual cells --
    `086280.KS` at 5.80x and `281820.KS` at 3.94x, with the next-largest deviation 1.43x.
    (The brief that raised this called it a single isolated cell; there are TWO past 1.5x.)

    WHY COMPUTE RATHER THAN PATCH THE CELL.  A per-name correction would be a hard-coded
    number with no rule behind it, and the next bad cell would arrive unannounced.  The house
    rule is the general one: the vendor supplies RAW INPUTS, we compute the derived quantity.
    Here the reconstruction is free -- the panel already carries `earningsYield`, which is the
    same FIELD that Stage-2's `earnYield` is built from.

    THEY ARE NOT THE SAME OBJECT, AND AN EARLIER VERSION OF THIS NOTE SAID THEY WERE
    (corrected, reviewer S3).  `earnYield` is a SIXTEEN-QUARTER windowed mean of that field
    (`stage2_metrics.STAGE2_METRIC_SPEC` declares it WINDOW_SCORING); this is the NEWEST ROW
    only.  So the published P/E and the scored cheapness share an input and a sign convention,
    not a value -- a name whose newest quarter is unrepresentative can read cheap here and
    ordinary in the score.  Newest-row is the right basis for a REPORT column (it is the P/E a
    reader would compute from today's statements) and the wrong one for a score, which is why
    the two differ on purpose.

    THE BASIS, STATED: `P/E = 1 / (rpy * earningsYield)` on the newest row.  `rpy` is the
    source's own rows-per-year (4 quarterly, 2 semi-annual), so a semi-annual filer's
    per-period yield is annualised by 2 rather than by a hard-coded 4 -- the same treatment
    every other flow quantity in this pipeline gets.

    REFUSED (None -> the vendor fallback, then 'NaN') when the yield is missing or NOT
    POSITIVE.  A loss-maker has no meaningful P/E; publishing a negative one invites it to be
    read as "cheap", which is the sign-inversion class this repo keeps finding.  `price /
    epsTTM` is deliberately NOT used as a second fallback: it is a DIFFERENT basis, and
    silently mixing two bases in one column is how a column stops meaning one thing.
    """
    try:
        ey, rpy = table.get(symb, (None, None))
        if ey is None or not np.isfinite(ey) or ey <= 0 or not rpy:
            return None
        return 1.0 / (float(rpy) * float(ey))
    except Exception:
        return None


# =========================================================================== #
#  THE DISPLAY COLUMNS THAT MIXED TWO CURRENCIES  (register N-3, CEO 2026-08-13) #
# =========================================================================== #
#  THE DEFECT.  `GrahamNumberToPrice` was `v3/key-metrics.grahamNumber / v3/profile.price`
#  -- a STATEMENT-currency numerator over a TRADING-currency denominator, with no FX
#  anywhere.  Every cross-listed name in the shortlist therefore published a ratio off by
#  its own exchange rate.  MEASURED on the shipped 2026-08-13 top-100:
#
#      SKHY        6763.0518     (KRW graham / USD price)     true, panel-computed: 3.99
#      SMSN.L        51.7596     (KRW graham / USD price)     true:                 0.23
#      SHEL.L         0.0113     (USD graham / GBp price)     true:                 1.47
#      column median  0.5275                                  true median:        0.9545
#
#  and the direction matters as much as the size: SKHY read as a 6,763x Graham bargain and
#  SHEL.L as 88x overvalued, from the same bug.  The RANKING WAS NEVER AFFECTED -- Stage-2
#  builds `grahamNumberToPrice` from the panel, where numerator and denominator are both in
#  `reportedCurrency` (`createDicts`: Upper='grahamNumber', Lower='price', and panel `price`
#  IS `marketCap / weightedAverageShsOut`, i.e. a statement-currency quantity).  This was a
#  DISPLAY defect only -- but the display is what the shortlist is read from.
#
#  THE FIX IS THE HOUSE RULE, NOT AN FX PATCH.  We could have multiplied by the run's live
#  rate; instead the column is now sourced from the metric the pipeline ALREADY COMPUTED for
#  the score.  The vendor supplies raw inputs; we compute the derived quantity.  An FX patch
#  would have left two definitions of one ratio in the codebase, and the whole class of
#  defect this repo keeps finding is two definitions drifting apart.
#
#  A SECOND CURRENCY DEFECT IS **NOT** FIXED HERE AND MUST NOT BE READ AS FIXED: the `price`
#  column itself is a raw quote in the LINE'S OWN currency (000660.KS = 1,616,000 KRW next
#  to TNK = 76.21 USD).  Converting it would be wrong -- the CEO reads it as the price he
#  would pay -- so it is LABELLED instead, by the new `priceCurrency` column beside it.
GRAHAM_BASIS_SCORED = 'scored'          # the Stage-2 metric that ranked this name
GRAHAM_BASIS_PANEL = 'panel-latest'     # newest panel row, for a name Stage-2 could not score


def _graham_to_price_panel_latest(cdx_df):
    """{source: grahamNumber / price} from the NEWEST panel row per source.

    FX-FREE BY CONSTRUCTION, which is the entire point: both columns are `reportedCurrency`
    quantities off the same row of the same panel (`grahamNumber` is stamped in-pipeline by
    `getData_fmp.stamp_frequency_and_graham`; `price` is `marketCap / weightedAverageShsOut`),
    so their ratio has no currency in it at all and no rate can be stale or wrong.

    This is the FALLBACK basis, not the primary -- see `_graham_to_price_published`.  Newest
    row taken by DATE, not by arrival order, for the same reason `_pe_panel_table` does it.
    Returns {} on any failure: this feeds a report column and must not cost the CSV.
    """
    try:
        cols = getattr(cdx_df, 'columns', [])
        if cdx_df is None or 'grahamNumber' not in cols or 'price' not in cols:
            return {}
        df = cdx_df.copy()
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df = df.sort_values(['source', 'date'], ascending=[True, False])
        newest = df.groupby('source', sort=False).head(1)
        gn = pd.to_numeric(newest['grahamNumber'], errors='coerce')
        px = pd.to_numeric(newest['price'], errors='coerce')
        ratio = (gn / px.where(px > 0)).replace([np.inf, -np.inf], np.nan)
        return {s: float(v) for s, v in zip(newest['source'], ratio) if np.isfinite(v)}
    except Exception:
        return {}


def _graham_to_price_published(raw_df, cdx_df):
    """({source: value}, {source: basis}) for the published `GrahamNumberToPrice`.

    PRIMARY = THE SCORED METRIC (`resdic['postScoreMetric_raw']['grahamNumberToPrice']`).
    Deliberately the literal number that ranked the name rather than a fresh point-in-time
    reading, because publishing a DIFFERENT number under a scored metric's own name is how
    `CycleHeat` came to ship a column whose correlation with the truth was -1.0000.  One name,
    one number.

    FALLBACK = the newest panel row, used ONLY where the scored value is NaN -- i.e. where
    Stage-2's window could not compute the metric and `normalizeAndDropNA` imputed it.  On
    the 2026-08-13 top-100 that is exactly THREE names -- STRT, ENS and PET.TO -- and for all
    three the newest row DOES compute (1.2613 / 0.5086 / 0.3024).  Publishing that is more
    informative than a blank.
    (An earlier version of this note said "exactly two names (STRT, PET.TO)" and UNDERSTATED
    its own evidence, reviewer L-1: the fallback set is EXACTLY the >=90%-imputed set found by
    register N-4 -- all three names, no more and no fewer.  That coincidence is not decoration,
    it is the mechanism: Stage-2 imputes the metric precisely when its window cannot compute
    it, which is the same condition that drives `imputed_weight_share` up.)

    THE MIX IS NAMED, NOT SILENT -- and from 2026-08-13 it is a COLUMN, not just a log line
    (`GrahamNumberToPrice_basis`, reviewer H-2).  The basis map is returned so the caller can
    publish and log which names fell back, as `_pe_vendor_fallback` does for the P/E.  A column
    that quietly carries two bases is a column that has stopped meaning one thing.

    NO VENDOR FALLBACK AT ALL, unlike the P/E.  The vendor's value here IS the defect: it is
    the FX-mixed ratio this whole block exists to remove, so falling back to it would restore
    the bug for precisely the names we know least about.
    """
    vals, basis = {}, {}
    try:
        cols = getattr(raw_df, 'columns', [])
        if raw_df is not None and 'source' in cols and 'grahamNumberToPrice' in cols:
            v = pd.to_numeric(raw_df['grahamNumberToPrice'], errors='coerce')
            for s, x in zip(raw_df['source'], v):
                if np.isfinite(x):
                    vals[s] = float(x)
                    basis[s] = GRAHAM_BASIS_SCORED
    except Exception:
        pass
    for s, x in _graham_to_price_panel_latest(cdx_df).items():
        if s not in vals:
            vals[s] = x
            basis[s] = GRAHAM_BASIS_PANEL
    return vals, basis


def _dividend_yield_ttm_from_panel(cdx_df):
    """{source: TRAILING-TWELVE-MONTH dividend yield, in PERCENT} from the run's own panel.

    THE DEFECT THIS REPLACES, AND IT IS NOT THE ONE THE BRIEF NAMED.  The column was
    `v3/key-metrics?period=quarter.dividendYield * 100`, and the brief that raised it read
    `TCL-A.TO = 259.87%` as a vendor fraction that was already a percent being multiplied by
    100 again.  IT IS NOT.  Our OWN panel, computed independently, puts TCL-A.TO's trailing
    yield at 272.6%: its newest row pays CAD 1.1536bn of dividends against a CAD 444m market
    cap after the price fell 23.10 -> 5.31 in one quarter.  That is a real distribution (or
    real broken vendor data about one), not a unit error, and `*100` is arithmetically
    correct.  The prescription was wrong and is recorded here so the next reader does not
    "fix" a scaling bug that never existed.

    THE DEFECT THAT IS REAL: `period=quarter` makes it a SINGLE QUARTER's yield published
    under a name every reader takes as ANNUAL.  Across the shipped 2026-08-13 top-100 the
    median ratio of the true trailing-twelve-month yield to the published number is 3.09
    across the 57 names that pay anything at all -- i.e. the column understated income by
    roughly the number of periods in a year, for every dividend payer in the shortlist.  SHEL.L shipped 0.9957% against a real trailing 3.87%; TNK shipped 1.9184%
    against 3.07%.  A 4x understatement of yield is not cosmetic on a value screen.

    THE BASIS: `-sum(dividendsPaid over rpy rows) / marketCap(newest row) * 100`.
      * `dividendsPaid` is a cash OUTFLOW and therefore negative in FMP's cash-flow
        statement, hence the leading minus; a POSITIVE reported value is refused rather than
        sign-flipped, because we do not know what it would mean.
      * `rpy` is the source's own rows-per-year, so a semi-annual filer sums 2 rows for a
        real twelve months instead of 4 rows for twenty-four -- the same treatment
        `stamp_frequency_and_graham` gives EPS.
      * numerator and denominator are BOTH `reportedCurrency`, so this is FX-free by
        construction -- the same property that makes the Graham ratio above safe.
      * the window must be COMPLETE (all `rpy` rows present and numeric).  A 3-of-4 sum
        masquerading as a year is the failure mode `ttm_aligned_sums` exists to refuse, and
        it would UNDERSTATE the yield, which is the direction that reads as safe.
    Returns {} on any failure.
    """
    try:
        cols = getattr(cdx_df, 'columns', [])
        if cdx_df is None or 'dividendsPaid' not in cols or 'marketCap' not in cols:
            return {}
        import reporting_period as _rp
        df = cdx_df.copy()
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df = df.sort_values(['source', 'date'], ascending=[True, False])
        try:
            freq = _rp.frequency_by_source(df, verbose=False)
        except Exception:
            freq = None
        out = {}
        for s, g in df.groupby('source', sort=False):
            try:
                rpy = int(_rp.rows_per_year(freq, s) if freq is not None
                          else _rp.DEFAULT_ROWS_PER_YEAR)
            except Exception:
                rpy = int(_rp.DEFAULT_ROWS_PER_YEAR)
            dp = pd.to_numeric(g['dividendsPaid'], errors='coerce').head(rpy)
            mc = pd.to_numeric(g['marketCap'], errors='coerce').iloc[0]
            if len(dp) < rpy or not dp.notna().all() or not np.isfinite(mc) or mc <= 0:
                continue
            total = float(dp.sum())
            if total > 0:
                continue                    # sign convention violated -- refuse, never flip
            #  `+ 0.0` NORMALISES NEGATIVE ZERO, and it is not a cosmetic (reviewer H-3).
            #  A non-payer sums `dividendsPaid` to exactly 0.0, and IEEE-754 makes `-0.0 / mc`
            #  a NEGATIVE zero, which `_fmt4` renders as `'-0.0000'`.  That hit 20 of the 100
            #  names on the shipped 2026-08-13 top-100 -- a fifth of the shortlist showing a
            #  negative-signed dividend yield -- and it was a REGRESSION against the vendor
            #  column this replaced, which printed a plain `0.0`.  `-0.0 + 0.0` is `+0.0` by
            #  the standard, and no other value is touched.
            out[s] = -total / float(mc) * 100.0 + 0.0
        return out
    except Exception:
        return {}


#  A TRAILING YIELD THIS LARGE IS A FINDING, NOT AN INCOME OPPORTUNITY, and the display layer
#  has to say so without inventing a verdict.  25% is the "no ordinary dividend policy pays
#  this" line: on the 2026-08-13 top-100 exactly one name (TCL-A.TO, 272.6%) is past it and
#  the 75th percentile is 3.87%.  Consumed by the HTML report to mark the value NEUTRALLY
#  (verify-this), never as a positive.
DIVIDEND_YIELD_IMPLAUSIBLE_PCT = 25.0


def writeBoAggToCSV(fb_df, mscore, cscore, baseurl, api_key, ntopagg, fname_AggScoretop,
                    flag_df=None, raw_df=None, universe_stamp=None, cdx_df=None,
                    missing_fill_df=None, run_date=None):
    """raw_df : resdic['postScoreMetric_raw'] -- the UNWEIGHTED, UN-NORMALISED metric
    frame.  Required to publish any metric under its own name: `fb_df` is postRank,
    whose metric columns are all `z x w` (see the CycleHeat note below).  ALSO the primary
    source of the published `GrahamNumberToPrice` (register N-3).

    cdx_df : resdic['cdx_dftop100'] -- the run's OWN fundamentals panel for these names.
    Used to COMPUTE the published `PE-ratio`, `dividendYield` and the fallback
    `GrahamNumberToPrice` instead of consuming FMP's -- see `_pe_ratio_from_panel`,
    `_dividend_yield_ttm_from_panel` and `_graham_to_price_published`.

    missing_fill_df : resdic['missing_fill_by_name'] -- the run's own per-name imputation
    audit (postBoRank.missing_data_fill_report).  Supplies `imputed_weight_share`, so a
    reader can see that a name was scored almost entirely on fills (register N-4).

    run_date : the run's date stamp.  Used ONLY to find THIS run's
    `VendorContaminationFlags_<date>.csv`, so the clone markers on `dollarVolume_basis` come
    from the same run as the numbers they qualify -- never from whichever dated file happens
    to be newest on disk, which is the cross-run mixing the deck's `resolve_run_artifacts`
    refuses for the same reason."""
    fbdf_tocsv = fb_df.head(ntopagg)
    symblist = list(fbdf_tocsv['source'])
    #BoComp_tocsv = pd.DataFrame(columns=['source','currentRatio','dividendYield','grahamNumberToPrice','price','beta',
    #                                    'sector','fmpRating','PEratio','M_score','C_score'])
    BoComp_tocsv = pd.DataFrame()
    BoComp_tocsv['source'] = symblist
    #quote_full = pd.DataFrame(requests.get(f'{baseurl}v3/quote/{symblist}?&apikey={api_key}').json())
    crVec = []
    priceVec = []
    margin = []
    dcf2p = []
    betaVec = []
    sectorVec = []
    ratingVec_fmp = []
    pEratioVec = []
    mscoreVec = []
    cscoreVec = []
    #  THE TRADING CURRENCY OF THE `price` COLUMN (register N-3).  Captured from the SAME
    #  profile response the price comes out of, so the label and the number cannot disagree
    #  -- an exchange suffix is NOT a substitute (SHEL.L quotes GBp and reports USD; the LSE
    #  IOB `0*.L` lines are foreign issuers).  `carveOut.trading_currency` is the offline
    #  fallback for a name whose profile call degraded.
    ccyVec = []
    #  The published P/E is COMPUTED from this table, not read off the vendor -- see
    #  `_pe_ratio_from_panel`.  Built ONCE for the whole CSV rather than per name: it is a
    #  groupby over the panel, and doing it inside the loop would be 100 passes over it.
    _pe_panel = _pe_panel_table(cdx_df)
    _pe_vendor_fallback = []
    #  COMPUTED, NOT CONSUMED -- both built once, for the same reason as the P/E table.  See
    #  the register N-3 block above `_graham_to_price_panel_latest` for what these replace and
    #  why an FX patch on the vendor's numbers was the wrong fix.
    _gtp_vals, _gtp_basis = _graham_to_price_published(raw_df, cdx_df)
    _gtp_panel_fallback = sorted(s for s, b in _gtp_basis.items()
                                 if b == GRAHAM_BASIS_PANEL and s in set(symblist))
    _div_ttm = _dividend_yield_ttm_from_panel(cdx_df)
    #  COMPUTED, NOT CONSUMED -- see `_current_ratio_panel_table` for the rank-7 false red
    #  flag that forced it.  Built once, for the same reason the P/E table is.
    _cr_panel = _current_ratio_panel_table(cdx_df)
    _cr_vendor_fallback = []
    _cr_refused = []
    #  Resolved ONCE, offline, from the run's volavgdic capture -- it is only the FALLBACK
    #  for a name whose live profile response degraded, so a per-row lookup would re-read the
    #  pickle 100 times for a value the loop usually does not need.
    try:
        import carveOut as _co_ccy
        _trading_ccy = {s: _co_ccy.trading_currency(s) for s in symblist}
    except Exception as _ce:
        print('WARNING: offline trading-currency fallback unavailable (%s: %s); the '
              '`priceCurrency` column falls back to NaN for any name whose profile call '
              'degrades.' % (type(_ce).__name__, _ce), flush=True)
        _trading_ccy = {}
    # Note: Bulk endpoints require higher subscription tier, using individual API calls only
    profile_bulk_dict = {}
    rating_bulk_dict = {}
    dcf_bulk_dict = {}
    
    print(f'Writing top {ntopagg} stocks to .csv')
    # GRANULARITY: THREE LAYERS, because the first two are not enough (review item 5,
    # 2026-07-31 -- this supersedes an earlier note here that argued a per-row guard was
    # impossible).
    #   1. THE CALL degrades: every request goes through gdg.safe_json_list, which cannot raise
    #      and returns [] -- so a throttled/hung endpoint costs that field, not the stage.
    #   2. THE VALUE degrades: `_fmt4` / `_one_mean_score` absorb the present-but-awkward values
    #      (`beta: null`, a duplicated issuer) that a healthy 200 can carry and that layer 1
    #      cannot help with.
    #   3. THE ROW degrades: the pad-to-length guard below, as a backstop for anything neither
    #      layer anticipated.
    # WHY LAYER 3 IS SAFE, since the obvious form is not: the twelve vectors are appended once
    # per name and then assigned as columns of a fixed-length frame, so a plain
    # `try/except: continue` leaves them RAGGED and the assignment raises "Length of values does
    # not match length of index" -- a one-name fault becoming total loss.  Padding in `finally`
    # to a per-row target length cannot ragged them, which is what makes the guard workable.
    _row_vectors = (priceVec, ccyVec, pEratioVec, betaVec, sectorVec, ratingVec_fmp, crVec,
                    margin, dcf2p, mscoreVec, cscoreVec)
    assert len(_row_vectors) == 11, len(_row_vectors)
    _rows_degraded = []
    # TOTAL TAKEN FROM THE FRAME THE LOOP ITERATES, not from the REQUESTED count: `fb_df` can
    # be shorter than `ntopagg` (a small universe, or a carve-out cohort with fewer members
    # than the ask), and the bar then hangs short of its total forever.  `desc` names the
    # stage because three bars run back-to-back after the fetch and 'the second one' is not
    # an identification.  Display only.
    pbar = tqdm(total=len(BoComp_tocsv), desc='AggScore CSV', unit='name',
                smoothing=0.05, dynamic_ncols=True)
    for _row_i, row in enumerate(BoComp_tocsv.itertuples(), start=1):
        # PAD-TO-LENGTH GUARD (review item 5, 2026-07-31).  My earlier reasoning that a
        # per-row guard was impossible here was wrong: only the NAIVE form is defeated by
        # the parallel per-row vectors (eleven of them since 2026-08-13; it was twelve when
        # this note was written).  A `finally` that pads every vector to this row's
        # target length CANNOT ragged them, so the column assignment after the loop can
        # never hit 'Length of values does not match length of index'.  A partially-
        # appended row is completed with the 'NaN' sentinel the columns already use.
        _want = _row_i                      # this row's target length for every vector
        try:
            symb = row.source
            # HARDENED (fix, 2026-07-31).  These were bare `requests.get(...).json()` -- no
            # timeout, no retry, and `.json()` chained to the call inside a loop body with NO
            # try/except anywhere in it.  ~4-5 calls x 100 names, fired immediately after 12+
            # hours of sustained API load, i.e. exactly when a throttle is most likely.  A
            # throttled 200 with an HTML body raised JSONDecodeError straight out of the stage and
            # cost the CSV, the XLSX, the forensic CSV, the postRank pickle AND the pick-log; with
            # no timeout, a hung socket stalled the unattended run indefinitely.
            # safe_json_list returns [] on any failure, so a FAILED CALL degrades that column for
            # that name.
            #
            # CORRECTION (review item 5, 2026-07-31).  An earlier version of this comment claimed
            # "every consumer below already guards on `len(...) == 0` -> 'NaN'".  That is FALSE, and
            # it was the false premise used to argue no per-row guard was needed.  Three raise paths
            # remained, all confirmed by execution, and NONE of them is a failed call -- each is a
            # healthy 200 with an awkward VALUE, which safe_json_list cannot help with:
            #   * `"{:.4f}".format(None)` -> TypeError on `price` and `beta`.  FMP returns
            #     `"beta": null` for plenty of non-US listings, and these two checked key PRESENCE
            #     but not None -- while the grahamNumberToPrice block a few lines down does check
            #     None explicitly, so the omission was an oversight, not a convention.
            #   * `grossProfitMargin` summed rows [1..3] after type-checking only row [0], so a
            #     None in any later row raised TypeError on `float + None`.  (NOT named in the
            #     review -- found applying its lesson to the rest of the loop.)
            #   * `.item()` on a zero-row (KeyError) or duplicated (ValueError) mscore/cscore
            #     selection -- issuer clones are a known phenomenon in this pipeline.
            # All three are now closed at the point of use, AND a pad-to-length guard backstops the
            # whole row (see the loop preamble) so no future raise here can ragged the vectors.
            temp_resp_km = gdg.safe_json_list(
                f'{baseurl}v3/key-metrics/{symb}?period=quarter&limit=4&apikey={api_key}',
                label='key-metrics %s' % symb)
            temp_resp_fr = gdg.safe_json_list(
                f'{baseurl}v3/ratios/{symb}?period=quarter&limit=4&apikey={api_key}',
                label='ratios %s' % symb)

            # Use bulk data for profile, rating, and DCF if available, otherwise fallback to individual calls
            if symb in profile_bulk_dict:
                temp_resp_pr = [profile_bulk_dict[symb]]  # Convert dict to list format
            else:
                temp_resp_pr = gdg.safe_json_list(
                    f'{baseurl}v3/profile/{symb}?apikey={api_key}', label='profile %s' % symb)

            _dcf_status = 'bulk'
            if symb in dcf_bulk_dict:
                temp_resp_dcf = [dcf_bulk_dict[symb]]  # Convert dict to list format
            else:
                # safe_json_list already normalises the dict/Error-Message body the block below
                # used to handle by hand, and never raises on an unparseable one.
                temp_resp_dcf = gdg.safe_json_list(
                    f'{baseurl}v3/discounted-cash-flow/{symb}?apikey={api_key}',
                    label='dcf %s' % symb)
                _dcf_status = 'ok' if temp_resp_dcf else 'empty-or-failed'

            # Diagnostic: Check what the DCF API actually returns (only for first ticker)
            if len(crVec) == 0:  # Only print for first ticker to avoid spam
                # Emitted through gdg.bar_print: this block fires on the FIRST name, i.e.
                # one line after the bar appears, and a multi-line dump into a live '\r'
                # render strands the bar on screen for the rest of the stage.  Same text.
                gdg.bar_print(f"\nDEBUG: DCF API response for {symb}:")
                gdg.bar_print(f"  Status: {_dcf_status}")
                gdg.bar_print(f"  Response type: {type(temp_resp_dcf)}")
                if isinstance(temp_resp_dcf, list):
                    gdg.bar_print(f"  Response length: {len(temp_resp_dcf)}")
                    if len(temp_resp_dcf) > 0:
                        gdg.bar_print(f"  First element type: {type(temp_resp_dcf[0])}")
                        if isinstance(temp_resp_dcf[0], dict):
                            gdg.bar_print(f"  First element keys: {list(temp_resp_dcf[0].keys())}")
                elif isinstance(temp_resp_dcf, dict):
                    gdg.bar_print(f"  Dict keys: {list(temp_resp_dcf.keys())}")
                    gdg.bar_print(f"  Dict content: {temp_resp_dcf}")
                else:
                    gdg.bar_print(f"  Response content: {temp_resp_dcf}")
        
            # Handle case where API returns a dict instead of a list (API might have changed)
            if isinstance(temp_resp_dcf, dict):
                # If it's a dict, try to convert to list format or extract error
                if 'Error Message' in temp_resp_dcf or 'error' in str(temp_resp_dcf).lower():
                    temp_resp_dcf = []  # Treat as empty
                else:
                    # Try to wrap in list if it's a single DCF object
                    temp_resp_dcf = [temp_resp_dcf] if temp_resp_dcf else []
        
            # Check if API responses are empty before accessing
            # currentRatio -- THE PANEL FIRST, the vendor only as a fallback, and BOTH
            # value-checked.  See `_current_ratio_panel_table`: this column used to be the
            # vendor's field behind a TYPE check with no VALUE check, so a vendor 0 shipped
            # as "0.0000" and the deck flagged a rank-7 pick as insolvent.
            _cr = _cr_panel.get(symb)
            if _cr is None:
                _cr_vendor_fallback.append(symb)
                _cr_raw = (temp_resp_fr[0].get('currentRatio')
                           if len(temp_resp_fr) and isinstance(temp_resp_fr[0], dict)
                           else None)
                _cr = _current_ratio_value(_cr_raw)
                #  A value the vendor SUPPLIED and this refused -- recorded by name, because
                #  a refusal that shows up only as 'NaN' is indistinguishable from the vendor
                #  never answering, and those are different facts about the run.
                if _cr is None and _cr_raw is not None:
                    _cr_refused.append('%s=%r' % (symb, _cr_raw))
            crVec.append('NaN' if _cr is None else "{:.4f}".format(_cr))
            
            #  `dividendYield` and `GrahamNumberToPrice` USED TO BE APPENDED HERE and are
            #  now assigned AFTER the loop -- see the block above their column assignments.
            #  They no longer need a vendor response, so they must not sit inside the row
            #  guard that nulls a row when one does.

            # The TRADING CURRENCY of the price on this same row -- from THIS response, so a
            # 1,616,000 (KRW) and a 76.21 (USD) can no longer sit in one column unlabelled.
            _cur = temp_resp_pr[0].get('currency') if len(temp_resp_pr) else None
            if not isinstance(_cur, str) or not _cur.strip():
                _cur = _trading_ccy.get(symb)             # offline fallback (volavgdic)
            ccyVec.append(_cur.strip() if isinstance(_cur, str) and _cur.strip() else 'NaN')


            # Check price -- _fmt4, because these two checked key PRESENCE but not None and
            # FMP returns `"price": null` / `"beta": null` on many non-US listings, so
            # `"{:.4f}".format(None)` raised TypeError here and cost the whole stage
            # (review item 5, 2026-07-31; the grahamNumberToPrice block above always
            # checked None explicitly, so these two were simply missed).
            if len(temp_resp_pr) == 0 or 'price' not in temp_resp_pr[0]:
                priceVec.append('NaN')
            else:
                priceVec.append(_fmt4(temp_resp_pr[0]['price']))

            # Check beta
            if len(temp_resp_pr) == 0 or 'beta' not in temp_resp_pr[0]:
                betaVec.append('NaN')
            else:
                betaVec.append(_fmt4(temp_resp_pr[0]['beta']))
            
            # Check sector
            if len(temp_resp_pr) == 0 or 'sector' not in temp_resp_pr[0]:
                sectorVec.append('NaN')
            else:
                sectorVec.append(temp_resp_pr[0]['sector'])
            
            # PE-ratio -- COMPUTED FROM OUR OWN PANEL, NOT READ OFF THE VENDOR.
            # See `_pe_ratio_from_panel` for the defect that forced this and the basis chosen.
            # The vendor's `priceEarningsRatio` is still READ, but only as a FALLBACK for a
            # name the panel cannot answer for, and it is labelled as such in the log.
            _pe_ours = _pe_ratio_from_panel(_pe_panel, symb)
            if _pe_ours is not None:
                pEratioVec.append("{:.4f}".format(_pe_ours))
            elif len(temp_resp_fr) == 0 or 'priceEarningsRatio' not in temp_resp_fr[0]:
                pEratioVec.append('NaN')
            else:
                perat = temp_resp_fr[0]['priceEarningsRatio']
                #  THE FALLBACK TAKES THE SAME SIGN TEST AS THE COMPUTED VALUE (reviewer S3,
                #  2026-08-10).  Without it the refusal was defeated by its own fallback: our
                #  own P/E is refused precisely when `earningsYield <= 0`, and the vendor's
                #  P/E on that same name is then NEGATIVE for the same reason -- so 100% of
                #  the refusing population got published exactly what the refusal exists to
                #  prevent.  MEASURED on the shipped 2026-08-10 top-100: one name refuses,
                #  `NEXN` (`earningsYield = -0.013804`), and its published `PE-ratio` was
                #  **-18.1111**.  A negative P/E invites "cheap"; a loss-maker has none.
                if (type(perat) == int or type(perat) == float) and perat > 0:
                    pEratioVec.append("{:.4f}".format(perat))
                    _pe_vendor_fallback.append(symb)
                else:
                    pEratioVec.append('NaN')
                
            # Check rating
            # Use bulk rating data if available, otherwise fallback to individual call
            if symb in rating_bulk_dict:
                temp_resp_rating = [rating_bulk_dict[symb]]
            else:
                temp_resp_rating = gdg.safe_json_list(
                    f'{baseurl}v3/rating/{symb}?apikey={api_key}', label='rating %s' % symb)
            if len(temp_resp_rating) == 0 or 'ratingRecommendation' not in temp_resp_rating[0]:
                ratingVec_fmp.append('NaN')
            else:
                ratingVec_fmp.append(temp_resp_rating[0]['ratingRecommendation'])
            
            # Check M_Score / C_Score -- via _one_mean_score, because the old
            # `[...].isna().item()` + `.item()` raised ValueError on a DUPLICATED source
            # (issuer clones are a known phenomenon here) and KeyError on an ABSENT one, either
            # of which cost the whole stage (review item 5, 2026-07-31).  Semantics are
            # unchanged for the healthy single-row case; absent/duplicated/NaN -> 'NaN'.
            mscoreVec.append(_fmt4(_one_mean_score(mscore, symb, 'M_Score_mean')))

            cscoreVec.append(_fmt4(_one_mean_score(cscore, symb, 'C_Score_mean')))

            # Check grossProfitMargin (needs 4 periods).  The sum spans rows [0..3] but only
            # row [0] was type-checked, so a None in ANY of rows 1-3 raised TypeError on
            # `float + None`.  NOT named in the review -- found applying its lesson to the rest
            # of the loop.  All four rows are now validated before summing.
            if len(temp_resp_fr) == 0 or 'grossProfitMargin' not in temp_resp_fr[0]:
                margin.append('NaN')
            elif len(temp_resp_fr) < 4:
                margin.append('NaN')
            else:
                _gpm = [temp_resp_fr[i].get('grossProfitMargin') for i in range(4)]
                if all(isinstance(v, (int, float)) and not isinstance(v, bool) and v == v
                       for v in _gpm):
                    margin.append(_fmt4(sum(_gpm) * 25))
                else:
                    margin.append('NaN')

            # Check DCF to Price
            if len(temp_resp_dcf) == 0:
                dcf2p.append('NaN')
            elif 'dcf' not in temp_resp_dcf[0]:
                dcf2p.append('NaN')
            elif temp_resp_dcf[0]['dcf'] is None:
                dcf2p.append('NaN')
            elif type(temp_resp_dcf[0]['dcf']) == int or type(temp_resp_dcf[0]['dcf']) == float:
                if 'Stock Price' not in temp_resp_dcf[0]:
                    dcf2p.append('NaN')
                elif temp_resp_dcf[0]['Stock Price'] is None:
                    dcf2p.append('NaN')
                elif type(temp_resp_dcf[0]['Stock Price']) == int or type(temp_resp_dcf[0]['Stock Price']) == float:
                    dcf2p.append("{:.4f}".format(temp_resp_dcf[0]['dcf']/(temp_resp_dcf[0]['Stock Price'])))
                else:
                    dcf2p.append('NaN')
            else:
                dcf2p.append('NaN')
        except Exception as _row_err:
            gdg.bar_print('WARNING: %s row degraded to NaN in the AggScore CSV (%s: %s) -- the'
                          ' remaining names and every later deliverable still run.'
                          % (symb, type(_row_err).__name__, _row_err))
            _rows_degraded.append(symb)
        finally:
            for _v in _row_vectors:
                del _v[_want:]              # discard a partial append, then pad to length
                while len(_v) < _want:
                    _v.append('NaN')
        # `len(_rows_degraded)` is the SAME list the summary below reports -- nothing computed
        # for display, so the live view and the summary cannot disagree.  refresh=False: the
        # update on the next line renders it.
        pbar.set_postfix_str('degraded=%d' % len(_rows_degraded), refresh=False)
        pbar.update(n=1)
    # THE BAR IS STILL OPEN from here down -- `pbar.close()` is at the END of the function, so
    # every summary/warning between the loop and it lands on the bar's un-terminated line and
    # gets glued to the right-hand end of it.  Routed through the bar-safe writer, same text.
    if _rows_degraded:
        gdg.bar_print('AGGSCORE-CSV DEGRADED-ROW SUMMARY: %d of %d name(s) had their API-sourced '
                      'columns written as NaN: %s'
                      % (len(_rows_degraded), len(BoComp_tocsv),
                         ', '.join(map(str, _rows_degraded))))
    #  SAY HOW MANY CELLS THE VENDOR STILL SUPPLIED.  A silently-mixed column is the thing
    #  this change exists to stop, so the count of fallback cells is printed rather than left
    #  to be inferred from the numbers.
    if _pe_vendor_fallback:
        gdg.bar_print(
            "PE-ratio: %d of %d name(s) fell back to FMP's `priceEarningsRatio` because our "
            "own panel could not answer (no positive earningsYield on the newest row): %s. "
            "Every OTHER cell is COMPUTED as 1/(rpy x earningsYield) from the run's own "
            "panel -- see postBo._pe_ratio_from_panel."
            % (len(_pe_vendor_fallback), len(BoComp_tocsv),
               ', '.join(map(str, _pe_vendor_fallback[:20]))))
    #  currentRatio: the SAME two lines the P/E gets, and for the same reason.  The first
    #  reports how many cells did NOT come from our own balance sheet; the second reports
    #  values the vendor supplied and this REFUSED, which a bare 'NaN' cannot distinguish
    #  from a vendor that never answered.
    if _cr_vendor_fallback:
        gdg.bar_print(
            "currentRatio: %d of %d name(s) fell back to FMP's `currentRatio` because our "
            "own panel could not answer (no usable totalCurrentAssets/totalCurrentLiabilities "
            "on the newest row): %s. Every OTHER cell is COMPUTED as "
            "totalCurrentAssets/totalCurrentLiabilities from the run's own panel -- see "
            "postBo._current_ratio_panel_table."
            % (len(_cr_vendor_fallback), len(BoComp_tocsv),
               ', '.join(map(str, _cr_vendor_fallback[:20]))))
    if _cr_refused:
        gdg.bar_print(
            "currentRatio: REFUSED %d impossible vendor value(s) (<=0 or non-finite), written "
            "as NaN rather than shipped: %s. Before 2026-08-24 a vendor 0 printed as '0.0000' "
            "and the deck flagged it as a solvency red flag -- 092730.KQ at RANK 7 on the "
            "2026-08-22 run."
            % (len(_cr_refused), ', '.join(map(str, _cr_refused[:20]))))
    if not _cr_panel:
        gdg.bar_print(
            "currentRatio: THE PANEL TABLE IS EMPTY -- every cell in this column came from "
            "the vendor's live `v3/ratios` call. Not a finding that the panel is clean; the "
            "computed basis did not run at all (missing totalCurrentAssets / "
            "totalCurrentLiabilities columns, or a failure inside "
            "postBo._current_ratio_panel_table).")
    #  SAY WHICH CELLS CAME OFF THE FALLBACK BASIS, exactly as the P/E line above does.  A
    #  column carrying two bases must name the rows, or it is a silently-mixed column.
    if _gtp_panel_fallback:
        gdg.bar_print(
            "GrahamNumberToPrice: %d of %d name(s) could not be published from the SCORED "
            "metric (Stage-2's window imputed it) and fall back to the newest panel row's "
            "grahamNumber/price: %s. Both bases are FX-free (statement currency over "
            "statement currency); neither is the vendor's FX-mixed ratio. These names also "
            "carry a high `imputed_weight_share` in this same CSV."
            % (len(_gtp_panel_fallback), len(BoComp_tocsv),
               ', '.join(map(str, _gtp_panel_fallback[:20]))))
    #  THE SYMMETRIC LINE, AND IT IS THE ONE THAT MATTERS MORE (reviewer L-2).  The block
    #  above reports the FALLBACK; this reports ABSENCE.  Both helpers behind
    #  `_graham_to_price_published` return `{}` on any exception, so a wholesale failure --
    #  a renamed panel column, a bad frame -- writes an ALL-NaN column and, without this,
    #  says nothing at all.  A silently empty column is worse than the wrong one it replaced,
    #  because nobody goes looking for a number that was never there.
    _gtp_missing = [s for s in symblist if s not in _gtp_vals]
    if _gtp_missing:
        gdg.bar_print(
            "GrahamNumberToPrice: %d of %d name(s) have NO published value -- neither the "
            "scored metric nor the newest panel row could answer: %s.%s"
            % (len(_gtp_missing), len(BoComp_tocsv),
               ', '.join(map(str, _gtp_missing[:20])),
               ("  THE WHOLE COLUMN IS EMPTY: this is a FAILURE of the computation, not a "
                "property of these companies -- check postBo._graham_to_price_published."
                if len(_gtp_missing) == len(symblist) else '')))
    _dy_missing = [s for s in symblist if s not in _div_ttm]
    if _dy_missing:
        gdg.bar_print(
            "dividendYield: %d of %d name(s) have NO trailing-twelve-month yield -- the panel "
            "does not hold a COMPLETE rpy-row dividend window for them, and a partial sum "
            "would understate the yield while looking like a real number: %s."
            % (len(_dy_missing), len(BoComp_tocsv), ', '.join(map(str, _dy_missing[:20]))))
    BoComp_tocsv['price'] = priceVec
    #  IMMEDIATELY BESIDE `price`, deliberately: column order is what a human reads, and the
    #  defect was that 1,443,000 and 76.21 sat in one column with nothing to distinguish them.
    BoComp_tocsv['priceCurrency'] = ccyVec
    BoComp_tocsv['PE-ratio'] = pEratioVec
    BoComp_tocsv['beta'] = betaVec
    BoComp_tocsv['sector'] = sectorVec
    BoComp_tocsv['rating_fmp'] = ratingVec_fmp
    BoComp_tocsv['currentRatio'] = crVec
    #  THESE TWO ARE ASSIGNED FROM THE PANEL, OUTSIDE THE ROW GUARD (2026-08-13).
    #  They were per-row vectors while they were vendor fields, and they had to be: the value
    #  arrived with the API response for that name.  Now that both are COMPUTED from the run's
    #  own panel they need no response at all, and leaving them inside the loop would keep
    #  them in the blast radius of the pad-to-length guard -- so a throttled `key-metrics`
    #  call on one name would blank a number we already held offline.  A degraded row must
    #  cost only what actually came from the vendor, which is what the DEGRADED-ROW SUMMARY
    #  above claims it costs.
    #  POSITIONAL over `symblist` is safe HERE and only here: this runs before the flag_df
    #  merge, so the frame is still exactly one row per requested name in order (the same
    #  reason the CycleHeat block below can do it, and the reason the volAvg block further
    #  down maps on `source` instead).
    BoComp_tocsv['dividendYield'] = [_fmt4(_div_ttm.get(s)) for s in symblist]
    BoComp_tocsv['GrahamNumberToPrice'] = [_fmt4(_gtp_vals.get(s)) for s in symblist]
    #  WHICH BASIS EACH CELL CAME FROM (reviewer H-2).  The column carries two -- the Stage-2
    #  scored metric for 97 of the 2026-08-13 top-100, the newest panel row for 3 -- and until
    #  now that was stated only in the run log.  It is published as a COLUMN because this same
    #  change added `dollarVolume_basis` for exactly this reason and `volAvg_asof` already sat
    #  beside `volAvg_report`: leaving Graham as the one two-basis column with no basis field
    #  is an inconsistency inside one diff, not a considered difference.
    #  An earlier version of this argued the mix was self-evident because the fallback names
    #  carry a high `imputed_weight_share` in the row beside it.  That is TRUE -- the fallback
    #  set is exactly the >=90%-imputed set -- but it asks the reader to cross-reference two
    #  columns AND know the rule connecting them.  A basis column just says it.
    BoComp_tocsv['GrahamNumberToPrice_basis'] = [
        _gtp_basis.get(s, 'unavailable') for s in symblist]
    BoComp_tocsv['GrossProfitMargin_ttm'] = margin
    BoComp_tocsv['DCF-to-Price'] = dcf2p
    BoComp_tocsv['M-Score'] = mscoreVec
    BoComp_tocsv['C-Score'] = cscoreVec
    # CycleHeat -- PUBLISHED FROM THE RAW FRAME, NOT FROM postRank (fix, 2026-07-29).
    #
    # THE DEFECT THIS REPLACES.  `fb_df` is resdic['postRank'], and postBoRank multiplies
    # every metric column by its weight BEFORE assembling that frame (the weighting loop in
    # postBoRank.postBoScoreRanking, just above its getAggScore call), so
    # postRank['CycleHeat'] is `z x (-0.080)` -- NOT the metric.
    # Because CycleHeat is winsor-EXEMPT (bounded/discrete), its z is an exact affine
    # function of the raw value, so multiplying by a NEGATIVE weight inverts it EXACTLY:
    # measured on the 2026-07-17 panel, corr(published, true) = -1.000000 and median ratio
    # -0.027282.  The published column's MINIMUM was therefore the HOTTEST name in the pool.
    #
    # SCOPE -- ONLY THIS CSV WAS AFFECTED.  The HTML deck was NOT: it sources CycleHeat from a
    # recomputed raw value and matches RawMetricsTop100 to 4.44e-16 (corr +1.000000).  An
    # earlier version of this comment claimed "the deck's cyclicality read was inverted for
    # every name" -- that was FALSE and is corrected here, because a wrong scope in a permanent
    # comment is how a wrong scope reaches a report.
    #
    # THE CLASS, swept exhaustively (2026-07-29): all 21 metric columns in postRank are
    # `z x w`, none are raw, and THREE sites copied metric-named columns out of it:
    #   * postBo.writeBoAggToCSV (here)          -- the defect above, fixed;
    #   * backtest_outputs.save_stock_picks      -- 7 columns, now renamed *_weighted_contrib;
    #   * backtest_unified / backtest_ols_analysis -- regressed on `z x w`, so CycleHeat's
    #     coefficient was reported sign-flipped; both now un-weight first.
    #   * moatScore -- SAFE, and deliberately left as-is: merged into postRank by Sbocker
    #     (`resdic['postRank'].merge(moat_merge, ...)`, after the postBoWrapper call), i.e.
    #     AFTER getAggScore has run, so never weighted, never summed into AggScore, and not a
    #     weight_series key.  Verified: integral values.  Pinned by
    #     test_published_columns.test_moatScore_is_raw_because_it_is_merged_after_scoring,
    #     which asserts the ORDER of those two Sbocker statements.
    # A first inspection found only two of the three sites; the frozen inventory in
    # baseline_tools/test_published_columns.py now fails on any unreviewed reader, because
    # "every metric column here is z x w" means a new reader is wrong BY DEFAULT.
    # That is why this now reads from `raw_df` and REFUSES to fall back to postRank.
    if raw_df is not None and 'CycleHeat' in getattr(raw_df, 'columns', []):
        # STRUCTURAL, not just by variable name: raw_df must DECLARE the raw basis
        # (postBoRank stamps it), so pointing this at a weighted frame is now caught here
        # instead of publishing an exactly sign-inverted column again.  An UNDECLARED frame
        # passes -- attrs are dropped by concat/merge -- which is why the frozen inventory in
        # baseline_tools/test_published_columns.py stays as the backstop.
        pbr.assert_metric_basis(raw_df, pbr.BASIS_RAW, label='writeBoAggToCSV raw_df')
        _cyc = raw_df.set_index('source')['CycleHeat']
        BoComp_tocsv['CycleHeat'] = [_cyc.get(s, np.nan) for s in symblist]
    elif 'CycleHeat' in fbdf_tocsv.columns:
        # Do NOT silently publish the weighted column again.  Emit the gap loudly instead:
        # a missing column is recoverable, a sign-inverted one is not detectable downstream.
        gdg.bar_print('!' * 78)
        gdg.bar_print('!!! CycleHeat OMITTED from %s: postScoreMetric_raw was not supplied, and the\n'
                      '!!! only other source (postRank) holds the WEIGHTED z (w=-0.080), which is\n'
                      '!!! SIGN-INVERTED against the metric. Publishing a gap, not a wrong number.'
                      % fname_AggScoretop)
        gdg.bar_print('!' * 78)
    # moatScore -- raw by construction (merged post-weighting by Sbocker, after the
    # postBoWrapper call); see above.
    if 'moatScore' in fbdf_tocsv.columns:
        BoComp_tocsv['moatScore'] = fbdf_tocsv['moatScore'].values

    # Merge in the forensic-flag decision-support columns (offline-computed; no API).
    # These make the M/C flags, their drivers, Sloan accruals, the financial-invalid
    # indicator and the summary tag visible alongside the aggregate scores.
    if flag_df is not None and not flag_df.empty:
        # Cross-check the offline (pickle-based) financial classification against the
        # API `sector` just fetched (sectorVec is aligned to symblist): if EITHER
        # source says bank/insurer/REIT, the name is forensic-invalid (conservative).
        flag_df = ff.applySectorFallback(flag_df, dict(zip(symblist, sectorVec)))
        #  `M_abstain_reason` RIDES THE SAME MERGE (CEO, 2026-08-16), for the same reason
        #  `forensicTag` does: the HTML deck reads this CSV, not the forensic one, so a reason
        #  that stopped here would explain the abstention in the artifact the CEO opens least.
        forensic_cols = ['source', 'isFinancial', 'financialKind', 'forensicValid',
                         'M_flag_gt_-1.78', 'M_drivers', 'M_abstain_reason', 'C_flag_ge_4',
                         'C_flags_fired', 'sloanAccruals',
                         'sloan_worstQuintile_inShortlist', 'forensicTag']
        keep = [c for c in forensic_cols if c in flag_df.columns]
        BoComp_tocsv = BoComp_tocsv.merge(flag_df[keep], on='source', how='left')

    # AVERAGE VOLUME -- REPORTED, NEVER SCREENED ON (register J-1, CEO 2026-08-06).
    # Two APPENDED columns (`volAvg_report`, `volAvg_asof`) beside the names the CEO actually
    # reads.  The CEO's ruling is report-only: no liquidity floor, no exclusion, no effect on
    # ordering -- so this runs AFTER `fbdf_tocsv = fb_df.head(ntopagg)` has already fixed both
    # the membership and the order, and it only ever ADDS columns to a frame whose rows are
    # settled.  It cannot reach the dedup survivor pick either: that happens far upstream in
    # carveOut, and the tiebreak's own reader (`_volavg_liquidity_term`) is untouched.
    # OFFLINE: the values come from the volavgdic pickle, so this adds ZERO API calls to a
    # stage that already makes ~4-5 per name.
    # Absence reads as absence, not as zero -- see carveOut.volavg_report_frame; every pickle
    # written before 2026-08-06 has no volume map at all, so `not-captured` is the expected
    # value on current data and an empty numeric cell must not be read as an illiquid name.
    try:
        import carveOut as _co_vol
        # MAPPED ON `source`, NOT POSITIONAL.  This is the first column assignment AFTER the
        # flag_df merge, and that merge is a `how='left'` on a table that can carry a
        # duplicated source (issuer clones are a known phenomenon here), so the frame is not
        # guaranteed to still be len(symblist).  A positional `.values` assignment would then
        # raise -- or, worse under a future refactor, misalign a volume onto the wrong name.
        _vol = _co_vol.volavg_report_frame(list(dict.fromkeys(BoComp_tocsv['source'])))
        _vol.index = list(dict.fromkeys(BoComp_tocsv['source']))
        BoComp_tocsv['volAvg_report'] = BoComp_tocsv['source'].map(_vol['volAvg_report'])
        BoComp_tocsv['volAvg_asof'] = BoComp_tocsv['source'].map(_vol['volAvg_asof'])
    except Exception as _ve:
        gdg.bar_print(f'WARNING: volAvg report columns not added to {fname_AggScoretop} '
                      f'({type(_ve).__name__}: {_ve}); CSV otherwise unaffected. The columns are '
                      f'REPORT-ONLY, so nothing selected or ranked is affected either.')

    # TRADED VALUE PER DAY -- the CEO's own named example (2026-08-13).  REPORT-ONLY on
    # exactly the same terms as `volAvg_report` above: appended to a frame whose membership
    # and order are already settled, no floor, no exclusion, no effect on ranking.
    # WHY IT IS ADDED AT ALL when `volAvg_report` is already here: `volAvg` is a SHARE count
    # and is therefore not comparable across listings -- 45.7M shares of a $154 line and 45.7M
    # shares of a KRW 1.5M line differ by four orders of magnitude in money.  Traded VALUE is
    # the comparable quantity, and it is now computable for 100% of the universe offline
    # (measured 2026-08-13: 3,144 of 3,145 names, the one gap being a name with no volume
    # reading at all).  See carveOut.dollar_volume_frame for the currency contract -- it
    # converts with the TRADING currency, which is the opposite of `marketcap_usd_series` and
    # correct for the same reason.
    # ZERO API CALLS: volAvg, price and currency all come from one volavgdic entry sharing
    # one `asof`, and the rate comes from the run's live FX table.
    try:
        import carveOut as _co_dv
        _uniq = list(dict.fromkeys(BoComp_tocsv['source']))
        #  THE CLONE MARKERS, from this run's own contamination artifact (reviewer H-5).
        #  Guarded separately from the frame itself: a missing/unreadable flags CSV must cost
        #  the QUALIFIER, never the number -- an unmarked traded value is a smaller loss than
        #  no traded value, and `clone_counterparts` already returns {} rather than raising.
        try:
            import vendor_contamination as _vc
            _clone = _vc.clone_counterparts(run_date=run_date)
        except Exception:
            _clone = {}
        _dv = _co_dv.dollar_volume_frame(_uniq, clone_map=_clone)
        _dv.index = _uniq
        BoComp_tocsv['dollarVolume_usd'] = BoComp_tocsv['source'].map(_dv['dollarVolume_usd'])
        BoComp_tocsv['dollarVolume_basis'] = BoComp_tocsv['source'].map(_dv['dollarVolume_basis'])
    except Exception as _de:
        gdg.bar_print(f'WARNING: dollarVolume columns not added to {fname_AggScoretop} '
                      f'({type(_de).__name__}: {_de}); CSV otherwise unaffected. REPORT-ONLY, '
                      f'so nothing selected or ranked is affected either.')

    # HOW MUCH OF THIS NAME'S SCORE IS FILL (register N-4, CEO 2026-08-13).
    # The run ALREADY computes this -- `postBoRank.missing_data_fill_report` writes it to
    # `MissingDataFillReport_<date>.csv` -- and it reached no artifact a shortlist reader
    # opens.  MEASURED on the shipped 2026-08-13 top-100: STRT (rank 55) is 93.2% fill, ENS
    # (rank 64) 93.2%, PET.TO (rank 71) 90.0%.  STRT and PET.TO carry `forensicTag = clean`;
    # ENS is tagged `data-incomplete: dig-deeper`, so the forensic layer caught one of the
    # three and not the other two -- which is the point, the two tags measure different things.  That tag is
    # not lying -- it describes the FORENSIC checks and nothing else -- but nothing in the CSV
    # told a reader the score was almost entirely imputed.  Top-100 MEDIAN is 0.0000, so this
    # is a three-name tail, not a systemic problem, and it is surfaced rather than acted on:
    # nothing here drops or re-orders a name.
    # A FRACTION OF WEIGHT, NOT A COUNT OF COLUMNS: the columns carry weights spanning ~20x,
    # so "16 of 19 columns" and "90% of the weight" are different statements and only the
    # second one says how much of the score is guesswork.
    try:
        if missing_fill_df is not None and not missing_fill_df.empty \
                and 'source' in missing_fill_df.columns:
            _imp = missing_fill_df.drop_duplicates('source').set_index('source')
            BoComp_tocsv['imputed_weight_share'] = BoComp_tocsv['source'].map(
                pd.to_numeric(_imp['imputed_weight_share'], errors='coerce'))
        else:
            #  ABSENCE READS AS ABSENCE.  A pre-2026-08-13 resdic carries no per-name fill
            #  table, and an all-zero column would assert "nothing was imputed" -- which is
            #  the one thing this column exists to stop anyone believing by default.
            gdg.bar_print(
                'NOTE: `imputed_weight_share` not added to %s -- this run carries no per-name '
                'missing-data fill table (resdic["missing_fill_by_name"]). The column is '
                'OMITTED rather than written as zeros, because a zero here would assert that '
                'nothing was imputed.' % fname_AggScoretop)
    except Exception as _ie:
        gdg.bar_print(f'WARNING: imputed_weight_share not added to {fname_AggScoretop} '
                      f'({type(_ie).__name__}: {_ie}); CSV otherwise unaffected.')

    # UNIVERSE STAMP IN THE RANKED CSV ITSELF (2026-08-03), not only in the sidecar.
    # This is the artifact most often read standalone, and its FILENAME carries only a
    # universe name whose meaning changed on 2026-08-02 -- so a reader comparing two
    # AggScore CSVs had nothing to compare bases on. Two constant columns, APPENDED, so
    # no existing consumer that selects columns by name is affected. Guarded: a stamp
    # must never cost the deliverable.
    try:
        BoComp_tocsv['universe'] = universe_stamp.get('universe') if universe_stamp else None
        BoComp_tocsv['universe_fingerprint'] = (
            (universe_stamp or {}).get('universe_fingerprint') or 'unknown-not-stamped')
    except Exception as _se:
        gdg.bar_print(f'WARNING: universe stamp not added to {fname_AggScoretop} '
                      f'({type(_se).__name__}: {_se}); CSV otherwise unaffected.')

    BoComp_tocsv.to_csv(fname_AggScoretop)
    pbar.close()
    return flag_df

def createPresentation(finalBoRank_df, mscore, cscore, baseurl, api_key, topn, fname, years, flag_df=None,
                       bands=None):
    #test
    #fname = fname_spreadSheet
    #topn = 20
    #years = 10
    # Market-cap banding (ADDITIVE): see generalTopN -- the General band (>$300M) when currency
    # data is present, unbanded postRank.head(topn) otherwise (byte-identical to pre-banding).
    # TODO: emit each sub-band's top-5 as its own labelled sheet block once the field flows
    # (deferred: the HTML presentation already carries the full banded view; adding sheets
    # here multiplies the per-symbol live API calls).
    _general_df = generalTopN(finalBoRank_df, bands, topn)
    symblist = list(_general_df['source'].head(topn))

    # AVERAGE VOLUME ON THE XLSX FACE -- REPORTED, NEVER SCREENED ON (register J-1; CEO
    # 2026-08-06, extending the AggScore-CSV columns to the XLSX because the XLSX is where he
    # actually reviews the top 20).
    #
    # WHY THIS CANNOT AFFECT SELECTION OR ORDERING, structurally rather than by intention:
    # `symblist` is FIXED on the two lines above, before this lookup exists, and nothing below
    # rebinds it, re-sorts it, or filters it -- the loop is `for symb in symblist[::-1]` and the
    # only thing done with a volume here is `ws.cell(...).value = ...`. There is no threshold,
    # no comparison and no floor anywhere in this block, so there is no code path by which a
    # volume reading can drop a page or move one. The dedup survivor tiebreak that DOES read
    # volume (`carveOut._volavg_liquidity_term`) ran far upstream and is untouched.
    #
    # OFFLINE, DELIBERATELY. Resolved ONCE here from the local volavgdic pickle -- NOT per
    # symbol and NOT from the profile response inside the loop, even though `pr[0]` happens to
    # carry a `volAvg`. Two reasons: this loop already fires ~7 live calls per name straight
    # after a 12-hour fetch and must not gain more work, and the profile's `volAvg` is UNDATED,
    # so using it would silently break the "the date travels with the number" rule that the CSV
    # columns keep. Same source, same three absence markers, same as-of date as the CSV -- one
    # number the CEO can compare across both artifacts.
    _vol_x = {}
    try:
        import carveOut as _co_volx
        _vf = _co_volx.volavg_report_frame(list(dict.fromkeys(symblist)))
        _vf.index = list(dict.fromkeys(symblist))
        _vol_x = {s: (_vf['volAvg_report'].get(s), _vf['volAvg_asof'].get(s))
                  for s in _vf.index}
    except Exception as _vex:
        # The volume cells are REPORT-ONLY, so their absence costs a number on a page and
        # nothing else. Never let it cost the XLSX -- but never let it pass quietly either:
        # an empty volume cell must be attributable to a named failure, not guessed at.
        print(f'WARNING: average-volume cells not added to {fname} '
              f'({type(_vex).__name__}: {_vex}); every page and every other deliverable is '
              f'unaffected, and nothing selected or ranked is affected either.', flush=True)

    #eyVec = []
    #quote_full = pd.DataFrame(requests.get(f'{baseurl}v3/quote/{symblist}?&apikey={api_key}').json())

    almago = datetime.today() - timedelta(weeks=5)
    if almago.day >= 5:
        lastlast_5th = almago.replace(day=5)
    else:
        lastlast_5th = (almago - timedelta(days=almago.day)).replace(day=5)
    ll5 = lastlast_5th.strftime('%Y-%d-%m')
    wb = openpyxl.Workbook()
    print(f'Writing top {topn} stocks to an .xlsx file for presentation')
    # Total from `symblist`, the list the loop iterates, not from the REQUESTED `topn`:
    # `generalTopN(...).head(topn)` returns fewer names whenever the banded general pool is
    # smaller than the ask, and the bar then stops short of its total for good.  Display only.
    pbar = tqdm(total=len(symblist), desc='XLSX presentation', unit='page',
                smoothing=0.05, dynamic_ncols=True)
    _pages_skipped = []
    for symb in symblist[::-1]:
        # HARDENED (fix, 2026-07-31): seven bare `requests.get(...).json()` calls per name,
        # x topn names, run immediately after 12+ hours of sustained API load and with no
        # timeout, no retry and `.json()` chained to the call.  Now routed through
        # gdg.safe_json_list -- 10s timeout, 3 retries with backoff, and [] instead of a
        # JSONDecodeError on a throttled 200 carrying an HTML body.
        km = pd.DataFrame(gdg.safe_json_list(
            f'{baseurl}v3/key-metrics/{symb}?period=annual&limit={years}&apikey={api_key}',
            label='key-metrics %s' % symb))
        fr = pd.DataFrame(gdg.safe_json_list(
            f'{baseurl}v3/ratios/{symb}?period=annual&limit={years}&apikey={api_key}',
            label='ratios %s' % symb))
        pr = gdg.safe_json_list(
            f'{baseurl}v3/profile/{symb}?apikey={api_key}', label='profile %s' % symb)
        rating = gdg.safe_json_list(
            f'{baseurl}v3/rating/{symb}?apikey={api_key}', label='rating %s' % symb)
        #target = requests.get(
        # f'https://financialmodelingprep.com/api/v4/price-target-consensus?symbol{symb}&apikey={api_key}').json()
        sp = gdg.safe_json_list(
            f'{baseurl}v4/stock_peers?symbol={symb}&apikey={api_key}', label='peers %s' % symb)
        cf = pd.DataFrame(gdg.safe_json_list(
            f'{baseurl}v3/cash-flow-statement/{symb}?period=annual&limit={years}&apikey={api_key}',
            label='cash-flow %s' % symb))
        dcf_resp = gdg.safe_json_list(
            f'{baseurl}v3/discounted-cash-flow/{symb}?apikey={api_key}', label='dcf %s' % symb)
        dcf = pd.DataFrame.from_dict(dcf_resp) if dcf_resp else pd.DataFrame()

        # PAGE-LEVEL DEGRADATION (fix, 2026-07-31).  pr / rating / sp / dcf / nspe are each
        # empty-guarded below, but km, fr and cf are NOT: every column of the sheet indexes
        # `km.<field>` / `fr['<field>']` / `cf.freeCashFlow` directly, so an empty response for
        # any of those three is an AttributeError/KeyError that propagates out of the loop and
        # costs the WHOLE XLSX -- and, before writeResWrapper guarded this call, every
        # deliverable after it too.  Hardening the transport does not fix that on its own: a
        # genuinely empty FMP response produces the same empty frame as a failed call.  So skip
        # THIS PAGE, name it, and report the total at the end.
        if km.empty or fr.empty or cf.empty:
            _missing = ', '.join(n for n, d in (('key-metrics', km), ('ratios', fr),
                                                ('cash-flow', cf)) if d.empty)
            gdg.bar_print(f'  WARNING: {symb} presentation page SKIPPED -- empty/failed {_missing}. '
                          f'The remaining pages and every later deliverable still run.')
            _pages_skipped.append(symb)
            pbar.set_postfix_str('skipped=%d' % len(_pages_skipped), refresh=False)
            pbar.update(n=1)
            continue

        # Check if DCF has required columns, use fallback if empty
        dcf_has_data = not dcf.empty and 'Stock Price' in dcf.columns and 'dcf' in dcf.columns

        NYSEspe = gdg.safe_json_list(
            f'https://financialmodelingprep.com/api/v4/sector_price_earning_ratio?date={ll5}&exchange=NYSE&apikey={api_key}',
            label='sector-PE')
        nspe_df = pd.DataFrame(NYSEspe) if isinstance(NYSEspe, list) and len(NYSEspe) > 0 else pd.DataFrame()
        nspe_has_data = not nspe_df.empty and 'sector' in nspe_df.columns and 'pe' in nspe_df.columns
        symb_df = pd.DataFrame(
            columns=['Symbol', 'Date', 'Earnings yield', 'PE-ratio', 'Price-to-book', 'Current ratio',
                     'Dividend yield', 'Price-to-fair value', 'Price'])
        symb_df['Symbol'] = fr['symbol']
        symb_df['Date'] = fr['date']
        symb_df['Earnings yield'] = (km.earningsYield * 100).apply(format_num)
        symb_df['PE-ratio'] = km.peRatio.apply(format_num)
        symb_df['Price-to-book'] = km.ptbRatio.apply(format_num)
        symb_df['Current ratio'] = km.currentRatio.apply(format_num)
        symb_df['Dividend yield'] = (km.dividendYield.fillna(0)*100).apply(format_num)
        symb_df['Price-to-fair value'] = fr.priceFairValue.apply(format_num)
        # Handle empty DCF data (common for non-US stocks)
        if dcf_has_data:
            price = dcf['Stock Price'].apply(format_num)
            symb_df['Price'] = price
            symb_df['Graham number to price'] = (km.grahamNumber/dcf['Stock Price']).apply(format_num)
        else:
            symb_df['Price'] = 'N/A'
            symb_df['Graham number to price'] = 'N/A'

        fcf = cf.freeCashFlow
        shares = fcf/km.freeCashFlowPerShare
        #qdDCFperPrice = quickDCF(fcf,0.12,0,km.interestDebtPerShare*shares,shares,price)

        if symb not in wb.sheetnames:
            ws = wb.create_sheet(symb, 0)
            ws.title = symb
        else:
            ws = wb[symb]
        wb.active = ws

        for r in dataframe_to_rows(symb_df, index=False, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        bold_font = Font(bold=True)
        psdf_col = len(symb_df.columns)+2
        psdf_row = 1
        
        # Check if profile data is available
        pr_has_data = pr and len(pr) > 0 and isinstance(pr[0], dict)
        
        ws.cell(row=psdf_row, column=psdf_col).font = bold_font
        ws.cell(row=psdf_row, column=psdf_col).value = 'Company'
        ws.cell(row=psdf_row+1, column=psdf_col).value = pr[0].get('companyName', 'N/A') if pr_has_data else 'N/A'

        ws.cell(row=psdf_row, column=psdf_col + 1).font = bold_font
        ws.cell(row=psdf_row, column=psdf_col+1).value = 'beta'
        if pr_has_data and pr[0].get('beta') is not None:
            ws.cell(row=psdf_row+1, column=psdf_col+1).value = "{:.4f}".format(pr[0]['beta'])
        else:
            ws.cell(row=psdf_row+1, column=psdf_col+1).value = 'N/A'

        ws.cell(row=psdf_row, column=psdf_col + 2).font = bold_font
        ws.cell(row=psdf_row, column=psdf_col+2).value = 'Market Cap'
        if pr_has_data and pr[0].get('mktCap') is not None:
            ws.cell(row=psdf_row+1, column=psdf_col+2).value = "{:,.2f}".format(pr[0]['mktCap']/1000000) + " million"
        else:
            ws.cell(row=psdf_row+1, column=psdf_col+2).value = 'N/A'

        ws.cell(row=psdf_row, column=psdf_col + 3).font = bold_font
        ws.cell(row=psdf_row, column=psdf_col+3).value = 'Industry & Sector'
        if pr_has_data and 'industry' in pr[0]:
            ws.cell(row=psdf_row+1, column=psdf_col+3).value = pr[0]['industry']
            ws.cell(row=psdf_row+2, column=psdf_col+3).value = pr[0].get('sector', 'N/A')
        else:
            ws.cell(row=psdf_row+1, column=psdf_col+3).value = 'N/A'
            ws.cell(row=psdf_row+2, column=psdf_col+3).value = 'N/A'

        ws.cell(row=psdf_row, column=psdf_col + 4).font = bold_font
        ws.cell(row=psdf_row, column=psdf_col+4).value = 'Sector Average PE-ratio'
        # Handle sector PE ratio - may not be available on all subscriptions
        if nspe_has_data and pr_has_data and 'sector' in pr[0]:
            sector_match = nspe_df[nspe_df['sector'] == pr[0]['sector']]
            if not sector_match.empty:
                secpe = sector_match.pe.iloc[0]
                ws.cell(row=psdf_row+1, column=psdf_col + 4).value = str(round(float(secpe), 4))
            else:
                ws.cell(row=psdf_row+1, column=psdf_col + 4).value = 'N/A'
        else:
            ws.cell(row=psdf_row+1, column=psdf_col + 4).value = 'N/A'

        # --- Average volume (REPORT-ONLY; resolved offline before the loop) ---------------
        # THE DATE TRAVELS WITH THE NUMBER, and the THREE KINDS OF ABSENCE stay distinguishable
        # on the sheet face exactly as they are in the CSV -- 'not-captured' (symbol absent from
        # the volume map), 'no-reading' (present but null/0/non-finite, FMP's usual answer for a
        # thin line) and 'undated-capture' (a real value from the pre-dating pickle shape). The
        # NUMBER cell is left EMPTY rather than 0 when there is no reading, because a 0 in a
        # volume column is a finding and an empty one is a gap; the status goes in the 'as of'
        # cell beneath, so a blank number is never ambiguous. This is why the marker strings are
        # read from carveOut rather than re-spelled here.
        ws.cell(row=psdf_row, column=psdf_col + 5).font = bold_font
        ws.cell(row=psdf_row, column=psdf_col + 5).value = 'Average volume'
        _v, _vasof = _vol_x.get(symb, (None, None))
        try:
            _vnum = float(_v)
        except (TypeError, ValueError):
            _vnum = float('nan')
        if _vnum == _vnum and _vnum > 0:                      # finite and positive
            ws.cell(row=psdf_row + 1, column=psdf_col + 5).value = "{:,.0f}".format(_vnum)
        else:
            #  No usable reading: leave the NUMBER cell blank (absence is not zero).
            ws.cell(row=psdf_row + 1, column=psdf_col + 5).value = None
        ws.cell(row=psdf_row + 2, column=psdf_col + 5).value = (
            'as of %s' % _vasof if _vasof else 'as of unknown -- lookup unavailable')

        ws.cell(row=psdf_row + 5, column=psdf_col).font = bold_font
        ws.cell(row=psdf_row + 5, column=psdf_col).value = 'Rating Recommendation'
        # Handle empty rating response
        if rating and len(rating) > 0 and 'ratingRecommendation' in rating[0]:
            ws.cell(row=psdf_row + 6, column=psdf_col).value = rating[0]['ratingRecommendation']
        else:
            ws.cell(row=psdf_row + 6, column=psdf_col).value = 'N/A'

        ws.cell(row=psdf_row + 5, column=psdf_col+2).font = bold_font
        ws.cell(row=psdf_row + 5, column=psdf_col+2).value = '(QD?) DCF per price'
        if dcf_has_data:
            ws.cell(row=psdf_row + 6, column=psdf_col+2).value = str(round(float(dcf['dcf']/dcf['Stock Price']), 4))
        else:
            ws.cell(row=psdf_row + 6, column=psdf_col+2).value = 'N/A'
        #ws.cell(row=psdf_row + 6, column=psdf_col+2).value = qdDCFperPrice

        ws.cell(row=psdf_row+5, column=psdf_col + 4).font = bold_font
        ws.cell(row=psdf_row + 5, column=psdf_col + 4).value = 'List of peers'
        # Handle empty stock peers response
        if sp and len(sp) > 0 and isinstance(sp[0], dict) and 'peersList' in sp[0]:
            peerslist = sp[0]['peersList']
            for i, peer in enumerate(peerslist):
                ws.cell(row=psdf_row + 6 + i, column=psdf_col + 4).value = peer
        else:
            ws.cell(row=psdf_row + 6, column=psdf_col + 4).value = 'N/A'

        # --- Forensic decision-support block (offline; from precomputed flag_df) ---
        # FLAGS, NOT VERDICTS. Surfaces the M/C flags + drivers, Sloan accruals, the
        # financial-invalid indicator and the summary tag so the CEO's manual review
        # sees the risk and its driver. No API call here.
        if flag_df is not None and not flag_df.empty and (flag_df['source'] == symb).any():
            frow = flag_df[flag_df['source'] == symb].iloc[0]
            fcol = psdf_col
            frow0 = psdf_row + 13
            ws.cell(row=frow0, column=fcol).font = bold_font
            ws.cell(row=frow0, column=fcol).value = 'FORENSIC FLAGS (guidance, not a drop)'
            if not bool(frow.get('forensicValid', True)):
                ws.cell(row=frow0, column=fcol + 1).value = (
                    f"INVALID for {frow.get('financialKind', 'financial')} — use financial lens")
            forensic_items = [
                ('Summary tag', frow.get('forensicTag', '')),
                ('Beneish M > -1.78?', 'FLAG' if frow.get('M_flag_gt_-1.78') else 'no'),
                #  WHY THERE IS NO M AT ALL (CEO, 2026-08-16).  Blank for a name that HAS an
                #  M, so the line appears exactly where it is needed; an abstention with a
                #  reason reads as a refusal, an abstention without one reads as a bug.
                ('  Why no M-score', frow.get('M_abstain_reason', '') or '-'),
                ('  M drivers', frow.get('M_drivers', '') or '-'),
                ('Montier C >= 4?', 'FLAG' if frow.get('C_flag_ge_4') else 'no'),
                ('  C flags fired', frow.get('C_flags_fired', '') or '-'),
                ('Sloan accruals', frow.get('sloanAccruals', '')),
                ('  Sloan worst-quintile (within shortlist)?',
                 'FLAG' if frow.get('sloan_worstQuintile_inShortlist') else 'no'),
                ('Financial (bank/insurer/REIT)?', 'YES' if frow.get('isFinancial') else 'no'),
            ]
            for i, (label, val) in enumerate(forensic_items, start=1):
                ws.cell(row=frow0 + i, column=fcol).font = bold_font
                ws.cell(row=frow0 + i, column=fcol).value = label
                ws.cell(row=frow0 + i, column=fcol + 1).value = ('' if val is None else str(val))

        resize_columns(ws)

        pbar.update(n=1)

    # Still inside the bar's lifetime -- `pbar.close()` is below the workbook save.
    if _pages_skipped:
        gdg.bar_print('PRESENTATION SKIPPED-PAGE SUMMARY: %d of %d page(s) omitted for empty/failed '
                      'API data: %s'
                      % (len(_pages_skipped), len(symblist),
                         ', '.join(map(str, _pages_skipped))))

    wb.save(fname)
    wb.close()
    pbar.close()

    return None

# Function to resize columns in an Excel sheet
def resize_columns(ws):
    for column_cells in ws.iter_cols():
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

def format_num(x):
    return "{:.4f}".format(x)

def moatIdentifier(symblist, cdx_df, n=20, freq_map=None):
    """Per-name 0-11 moat criteria count.  DISPLAY-ONLY (merged into postRank AFTER
    getAggScore, so it is never weighted and never summed into AggScore) -- but it reaches
    the CEO as an absolute count, so its windows matter.

    RUN-KILLER FIXED 2026-08-02 (CEO-authorised error-path change).  The per-symbol re-sort
    uses rp.ON_BAD_DATE_RAISE, and the production call site
    (Sbocker: `pb.moatIdentifier(resdic['BoScore_df']['source'], resdic['cdx_df'])`) is
    UNGUARDED and runs over the FULL universe (~7.7k names).  One unparseable date on ANY
    name used to raise out of this function and take down everything after it in Sbocker --
    detectManipulation, writeResWrapper, and with it every deliverable of a 12-hour fetch.
    Same shape as the `_FailedResponse.text` outage, and invisible to every bit-identity
    check, because it fires only on a rare data shape that no saved panel contains.

    THE FIX IS PER-NAME CONTAINMENT, NOT A LOOSER DATE POLICY and NOT a guard at the call
    site.  Each name is scored inside its own try; a failure yields that name's row with
    every criterion and moatScore NaN, and the loop continues.  The distinction is
    deliberate:
      * relaxing ON_BAD_DATE_RAISE would score the name off a NaT-sorted window -- a wrong
        number instead of a missing one, on a figure the CEO reads as an absolute count;
      * a try/except around the Sbocker call would save the run but silently drop the moat
        score for the ENTIRE universe.  This drops ONE name and names it.
    Failures are reported at the end of the loop -- loudly with every name when there are
    any, and as an affirmative "0 per-name failures" line when there are none, so a dead
    guard is distinguishable from a healthy one.
    """
    #moatdf = pd.DataFrame(columns=['source','moatScore'])
    #for symb in symbollist:
    # calculate FCFyield (>5-10%), Gross Margin (GrossProfit/Revenue>30%)
    # Calculate sales/Assets (>0.75)
    # Calculate RoE (>15%) and RoA(>10%) and ROIC (>15%)
    # High earnings even when market earnings are down (beta < 1?)
    # Calculate SG&A/GrossProfit (<15%) [sale, general and administrative expenses]
    # Calculate Depreciation/GrossProfit (<10%)
    # Calculate InterestExpenses/OperatingIncome (<15%)
    # Calculate NetMargin [NetIncome/Revenue] (>20%)
    # Calculate Capex/NetIncome (<25%)
    # Calculate TotalLiabilities/ShareholderEquity (<0.8)
    #moatdf = pd.DataFrame(columns=['source','moatScore','FCFyield','GrossMargin','RevtoASS','RoE','RoA','ROIC','SGAtoGP','DeptoGP',
    #                               'InteresttoOI', 'NetMargin','CapExtoEarnings','TLtoEquity'])
    moatdf = pd.DataFrame(columns=['source','moatScore','FCFyield','GrossMargin','RevtoASS','RoE','RoA','ROIC','SGAtoGP','DeptoGP',
                                   'NetMargin','CapExtoEarnings','TLtoEquity'])
    nan_dict = {'source': np.nan,  'moatScore': np.nan, 'FCFyield': np.nan, 'GrossMargin': np.nan, 'RevtoASS': np.nan,
                'RoE': np.nan, 'RoA': np.nan, 'ROIC': np.nan, 'SGAtoGP': np.nan, 'DeptoGP': np.nan,
                'NetMargin': np.nan, 'CapExtoEarnings': np.nan, 'TLtoEquity': np.nan}
    # Every criterion below is a head(n) mean over the name's own history, so `n` is
    # scaled per source to its rows-per-year: n=20 is 5 years of quarters but TEN years
    # of halves, which graded semi-annual filers on twice the history (row-based site
    # NOT on the audit's list -- found in the 2026-07-25 sweep). Display-only metric,
    # but it reaches the CEO as an absolute 0-11 count.
    if freq_map is None:
        freq_map = rp.frequency_by_source(cdx_df)
    tempdf_orig = pd.concat([moatdf, pd.DataFrame([nan_dict])], ignore_index=True)
    # The 11 moat criteria that ARE the score -- each column is already expressed as
    # (value - threshold), so "> 0" means "passes".  Counting these EXPLICITLY (rather
    # than every numeric column) is half of the audit M2 fix below.
    moat_criteria = ['FCFyield', 'GrossMargin', 'RevtoASS', 'RoE', 'RoA', 'ROIC',
                     'SGAtoGP', 'DeptoGP', 'NetMargin', 'CapExtoEarnings', 'TLtoEquity']
    _n_quarterly = n            # the caller's window, expressed in QUARTERS
    #  PER-NAME CONTAINMENT (fix, 2026-08-02) -- see the docstring.  One company's bad data
    #  must cost THAT COMPANY's moat row, never the other ~7,700 names and everything
    #  Sbocker runs after this call.
    _failed = []
    for symb in symblist:
        # .copy() -- audit M2 fix (2026-07-19).  `tempdf = tempdf_orig` bound the SAME
        # one-row frame every iteration, so the PREVIOUS ticker's moatScore was still
        # sitting in it when `select_dtypes(include='number') > 0` counted the criteria,
        # and any prior score > 0 was counted as a 12th criterion.  Effect on the shipped
        # 2026-07-17 universe: moatScore too high by exactly +1 on 7,751 of 7,752 names
        # (only the first-processed ticker, whose carried value was still NaN, was right),
        # and the observed minimum was 1 on an 0-11 scale where 0 must be attainable --
        # 496 names shown as "1" are really 0.  The number is display-only (moatScore is
        # merged AFTER Stage-2, so AggScore never saw it), but it reaches the CEO as an
        # absolute count.
        tempdf = tempdf_orig.copy()
        # NEWEST-FIRST BEFORE head(n) (ship-gate BLOCKER, fixed 2026-07-27).
        # Every comparator below is a `.head(n).mean()`, i.e. a RECENCY window -- but
        # cdx_df arrives OLDEST-first (data_quality sorts ascending; verified 7,752 of
        # 7,752 sources), so head(n) was reading the name's OLDEST rows.  Median lag of
        # the window from the newest filing was 1.00 year for a quarterly filer and
        # 7.00 years for a semi-annual one: the 2026-07-25 window scaling (head(20) ->
        # head(10) for rpy=2) AMPLIFIED the pre-existing defect by anchoring a shorter
        # window at the wrong end.  moatScore differed on 50.2% of 400 real names, by up
        # to +-7 points on an 0-11 scale.
        # Goes through the SHARED row-order boundary (reporting_period, "ROW ORDER") with the
        # policy stated in the call, rather than a fresh sort or a bare helper name: there is
        # ONE definition of "newest-first" in the pipeline and one place that decides what an
        # unparseable date means.  It sorts by PARSED date, so it is robust to however the
        # rows happen to arrive.
        #
        # ON_BAD_DATE_RAISE is deliberately KEPT (2026-08-02).  moatScore's comparators are
        # all `head(n)` recency windows, so a date that will not parse means THIS NAME's
        # window cannot be trusted -- the strict policy is right, and relaxing it here would
        # silently score the name off a NaT-sorted window.  What was wrong was the BLAST
        # RADIUS, not the strictness: the raise escaped an unguarded call site and cost every
        # other name.  It is now contained per name (see the `except` at the end of this
        # loop), so the strict verdict costs exactly the company it is about.
        try:
            cdx_temp = rp.to_newest_first(cdx_df[cdx_df['source'] == symb],
                                          rp.ON_BAD_DATE_RAISE)
            _rpy = rp.rows_per_year(freq_map, symb)
            n = rp.scale_window(_n_quarterly, _rpy)
            # ANNUALIZE the FLOW-over-STOCK comparators (specialist ruling, 2026-07-25).
            # Every bar below is an ANNUAL rule of thumb -- FCF yield > 10%, sales/assets >
            # 0.75, RoE > 15%, RoA > 10%, ROIC > 15% -- but the ratios were per-PERIOD, so a
            # quarterly filer was being asked to earn a full YEAR's return in three months.
            # Five of the eleven comparators were therefore near-unpassable by construction.
            # `af` is a TRUE annualisation (x4 quarterly, x2 semi-annual) because these
            # thresholds are ABSOLUTE; it deliberately changes quarterly names too, and
            # moatScores rise materially as a result.
            # NOT annualized (flow/flow or stock/stock -- scale cancels): GrossMargin,
            # SGAtoGP, DeptoGP, NetMargin, CapExtoEarnings, TLtoEquity.
            af = rp.annualize_factor(_rpy)
            tempdf['source'] = symb
            #  ---- THE WINDOW IS POSITIONAL: MASK TO NaN, NEVER FILTER BEFORE head(n) -------
            #  (fetch-depth audit, 2026-08-14.)  Four of the eleven comparators used to DROP
            #  their inadmissible rows and THEN take head(n) -- `series[mask].head(n)`.  A
            #  filtered-then-windowed series has NO CALENDAR BOUND: head(20) over a series
            #  with the loss-making / zero-denominator rows removed reaches back until it has
            #  found 20 SURVIVING rows, however far that is.  So the window's calendar span
            #  was a function of the FETCH DEPTH for exactly the names most likely to have
            #  inadmissible rows -- and at `-nrperiods 80` it could reach back TWENTY YEARS to
            #  fill a window whose own comment calls it "5 years of quarters".
            #  MEASURED: on the 2026-08-13 CUR3K panel, re-deepening 24 -> 80 moved moatScore
            #  by a full point on 5 of 1,884 quarterly names with nothing else changed
            #  (baseline_tools/depth_sensitivity.py) -- and it is display-only but the CEO
            #  reads it as an absolute 0-11 count.
            #
            #  THE ON-PANEL COST OF THE FIX ITSELF, AT TODAY'S DEPTH, WITH THE COUNTING RULE
            #  STATED (2,629 sources; review item M-6 asked for exactly this, because a single
            #  "flip count" merges three different events and two readers got two answers).
            #  moatScore is `sum(criterion > 0)` and `NaN > 0` is False, so what moves the
            #  score is a change in the `> 0` BOOLEAN:
            #        value -> value sign flips ....... 47
            #        NaN   -> pass  (score +1) ......   0
            #        pass  -> NaN   (score -1) ......  24
            #        NaN transitions moving nothing ..  25   (not counted)
            #        --------------------------------------
            #        SCORE-RELEVANT flips ...........  71   of 10,516 evaluations
            #  The VALUES move far more often than the verdicts: CapExtoEarnings moves on
            #  54.2% of sources (loss quarters are common, and the old window reached PAST
            #  them), FCFyield 4.4%, SGAtoGP 4.0%, DeptoGP 3.4%.
            #  WHAT THE CEO ACTUALLY SEES: 68 of 2,629 names (2.59%) change their 0-11 count,
            #  by -1 on 33, +1 on 33, +2 on 1 and -3 on 1.  Max movement 3 points.
            #  THE FIX IS THE PIPELINE'S OWN CONVENTION, not a new one: every other window
            #  here -- `nan_policy.window_verdict`, `calcScore.calcByTier` -- keeps an
            #  inadmissible row IN PLACE as NaN and windows POSITIONALLY, then lets `.mean()`
            #  skip it.  The value becomes "the mean over the computable rows INSIDE the
            #  newest n", which is the honest reading and is fetch-depth invariant.
            #  BEHAVIOUR CHANGE, on this panel too, for any name with an inadmissible row
            #  inside its window -- declared, not silent.  It only ever moves the window
            #  NEWER.
            _pfcf = pd.to_numeric(cdx_temp['pfcfRatio'], errors='coerce')
            fcfyield_filter = _pfcf.where(_pfcf != 0)
            tempdf['FCFyield'] = (1/fcfyield_filter).head(n).mean()*af-0.1
            tempdf['GrossMargin'] = cdx_temp['grossProfitMargin'].head(n).mean()-0.3
            tempdf['RevtoASS'] = (cdx_temp['revenue']/cdx_temp['totalAssets']).head(n).mean()*af-0.75
            tempdf['RoE'] = cdx_temp['returnOnEquity'].head(n).mean()*af-0.15
            tempdf['RoA'] = cdx_temp['returnOnAssets'].head(n).mean()*af-0.1
            tempdf['ROIC'] = cdx_temp['returnOnCapitalEmployed'].head(n).mean()*af - 0.15
            #  Masked to NaN IN PLACE, not filtered out -- see the FCFyield note above.
            _gp = pd.to_numeric(cdx_temp['grossProfit'], errors='coerce')
            gp_filter = _gp.where(_gp != 0)
            tempdf['SGAtoGP'] = 0.15-(cdx_temp['sellingGeneralAndAdministrativeExpenses']/gp_filter).head(n).mean()
            tempdf['DeptoGP'] = 0.1 - (cdx_temp['depreciationAndAmortization']/gp_filter).head(n).mean()
            #tempdf['InteresttoOI'] = 0.15 - (cdx_df['interestExpense']/cdx_df['operatingIncome']).head(n).mean()
            tempdf['NetMargin'] = cdx_temp['netProfitMargin'].head(n).mean() - 0.2
            # CapEx/Earnings: |capex| / |NI| < 0.20, GATED ON NI > 0 (domain review S8, fixed
            # 2026-07-26).  It was `0.2 - mean(capexPerShare/netIncomePerShare)`, which
            # free-passed 100.0% of loss-makers vs 44.5% of profitable names.
            # NOTE the audit's stated MECHANISM was wrong and would have produced the wrong fix:
            # it blamed "FMP capexPerShare negative", but on this panel capexPerShare is POSITIVE
            # on 86.5% of 176,604 rows and negative on 0.0000% (median +0.084).  The real cause is
            # the DENOMINATOR -- for NI < 0 the ratio is negative on 83.4% of rows, so
            # `0.2 - negative > 0` is an automatic tick.
            # "Capex is a small share of earnings" is UNDEFINED when there are no earnings, so a
            # loss-making period is NOT-COMPUTABLE (NaN), not a pass: NaN > 0 is False, so it
            # neither ticks the box nor counts against the other ten comparators.
            #  `_prof` is applied with `.where`, NOT as a row filter -- see the FCFyield note
            #  above.  The gate itself (NI > 0 is required for the ratio to mean anything) is
            #  UNCHANGED; only the window stops sliding backwards to replace the gated rows.
            _ni_ps = pd.to_numeric(cdx_temp['netIncomePerShare'], errors='coerce')
            _cx_ps = pd.to_numeric(cdx_temp['capexPerShare'], errors='coerce')
            _prof = _ni_ps > 0
            tempdf['CapExtoEarnings'] = (0.2 - (_cx_ps.abs() / _ni_ps.abs())
                                         .where(_prof).head(n).mean())
            tempdf['TLtoEquity'] = 0.8 - (cdx_temp['totalLiabilities']/cdx_temp['totalStockholdersEquity']).head(n).mean()
            # Count ONLY the 11 criteria -- never moatScore itself, and never any column that
            # happens to be numeric.  NaN > 0 is False, so a non-computable criterion does not
            # pass (an unchanged property of the old code).
            mask = tempdf[moat_criteria].apply(pd.to_numeric, errors='coerce') > 0
            tempdf['moatScore'] = mask.sum(axis=1)
        except Exception as _e:
            # PER-NAME CONTAINMENT (fix, 2026-08-02).  This loop runs over the FULL universe
            # (~7.7k names) from an UNGUARDED Sbocker call site, so an exception here used to
            # propagate out and take down detectManipulation, writeResWrapper and every
            # deliverable of a 12-hour fetch -- one company's malformed date destroying the
            # expensive work done for every other company.  That is the same failure shape as
            # `_FailedResponse.text` aborting Stage-2 for all six pools.
            #
            # CONTAINED PER NAME, NOT AT THE CALL SITE, and the difference matters: a guard
            # around the call in Sbocker would also save the run, but it would silently lose
            # the moat score for the WHOLE UNIVERSE.  This loses ONE name and says which.
            #
            # The row is rebuilt from the pristine NaN template so a PARTIALLY-written tempdf
            # can never be emitted as if it were scored: every criterion is NaN, and because
            # `NaN > 0` is False the name would score 0 anyway -- so moatScore is left NaN
            # rather than 0, which is the honest 'not computable' and is what the downstream
            # `merge(..., how='left')` already produces for an absent name.
            tempdf = tempdf_orig.copy()
            tempdf['source'] = symb
            _failed.append((symb, '%s: %s' % (type(_e).__name__, _e)))

        moatdf = pd.concat([moatdf, tempdf]).reset_index(drop=True)

    # ZERO IS REPORTED AS A RESULT, NOT AS SILENCE -- the same standard as the
    # reporting-frequency watchdog (reporting_period.log_conflicts), and for the same reason:
    # a guard that is silent when healthy cannot be told apart from a guard that is dead, and
    # a SILENTLY skipped name is exactly how a defect hides.  One line on the happy path, a
    # banner naming every casualty otherwise.
    if _failed:
        print('!' * 78, flush=True)
        print('!!! moatIdentifier: %d of %d name(s) could NOT be scored and carry moatScore =\n'
              '!!! NaN. The run CONTINUES -- these are per-name failures, contained so one\n'
              '!!! company\'s data cannot cost the other %d. Most likely cause: a date that will\n'
              '!!! not parse (the recency window is unusable, so the name is not scored).'
              % (len(_failed), len(moatdf), max(0, len(moatdf) - len(_failed))), flush=True)
        for _s, _why in _failed[:40]:
            print('!!!   %-16s %s' % (_s, _why), flush=True)
        if len(_failed) > 40:
            print('!!!   ... (+%d more)' % (len(_failed) - 40), flush=True)
        print('!' * 78, flush=True)
    else:
        print('moatIdentifier: %d name(s) scored, 0 per-name failures.' % len(moatdf),
              flush=True)

    moatdf.sort_values(by='moatScore', ascending=False, inplace=True)

    return moatdf


#CLEAN THIS UP
def findHighestOfEachSector(resdic):
    #sectorlist = resdic['sectorlist']
    #sectorlist =  ['all', 'Unspecified', 'Basic Materials', 'Healthcare', 'Financial Services',
    #              'Energy', 'Consumer Cyclical', 'Consumer Defensive', 'Industrials',
    #              'Communication Services', 'Technology', 'Real Estate', 'Utilities','Biotechnology']
    #sectorlist.remove('all')
    #sectorlist.remove('Unspecified')
    sectordic = pd.read_pickle('sectorsdic_fmp.pickle')
    sectorlist = list(sectordic.keys())
    baseurl = resdic['baseurl']
    api_key = resdic['api_key']
    sectorsfound = []
    sectorsnotfound = sectorlist
    bsdf = resdic['BoScore_df'].reset_index(drop=True)
    bsdf['score'] = bsdf['score'].astype(float)
    bshs = pd.DataFrame(columns=['source', 'score', 'sector'])
    symblist = bsdf['source']

    for sector in sectorlist:
        sslist = sectordic[sector]
        if len(bsdf.loc[bsdf['source'].isin(sslist), 'score']) > 0:
            highest_rowid = bsdf.loc[bsdf['source'].isin(sslist), 'score'].idxmax()
            newrow = {'source': bsdf.loc[highest_rowid]['source'], 'score': bsdf.loc[highest_rowid]['score'], 'sector': sector}
            tempdf = pd.DataFrame([newrow])
            bshs = pd.concat([bshs, tempdf], ignore_index=True)
    bshs = bshs.reset_index(drop=True)

    #bshs['sector'] = sectorlist

    # Removed second loop that was searching for sectors not in sectordic
    # The first loop (lines 388-394) already finds the highest-scoring stock for each sector
    # that exists in the current dataset. The second loop was inefficiently searching through
    # all symbols for sectors that don't exist due to sector mapping/consolidation.

    resdic = {**resdic, **{'BoScore_highsectors': bshs}}

    return resdic

#see sector performance
#https://financialmodelingprep.com/api/v3/stock/sector-performance?apikey=YOUR_API_KEY
#https://financialmodelingprep.com/api/v3/historical-sectors-performance?limit=50&apikey={api_key}

## Inspect Owner's Earnings

## Get historical dividends via
#https://financialmodelingprep.com/api/v3/historical-price-full/stock_dividend/AAPL?apikey=YOUR_API_KEY

## see analyst estimates: https://financialmodelingprep.com/api/v3/analyst-estimates/AAPL?limit=30&apikey=YOUR_API_KEY
